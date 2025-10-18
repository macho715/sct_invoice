#!/usr/bin/env python3
"""특정 행의 색상 확인"""

import openpyxl


def check_specific_colors():
    try:
        wb = openpyxl.load_workbook("HVDC WAREHOUSE_HITACHI(HE).debug.xlsx")
        ws = wb.active

        # Row 3241 확인 (ETA/ATD, ETD/ATD 변경됨)
        print("Row 3241 색상 확인:")
        print(
            f"  Col 30 (ETD/ATD): {ws.cell(3241, 30).value} - Fill: {ws.cell(3241, 30).fill.start_color.rgb if ws.cell(3241, 30).fill.start_color else 'No Fill'}"
        )
        print(
            f"  Col 31 (ETA/ATA): {ws.cell(3241, 31).value} - Fill: {ws.cell(3241, 31).fill.start_color.rgb if ws.cell(3241, 31).fill.start_color else 'No Fill'}"
        )

        # Row 4069 확인
        print("\nRow 4069 색상 확인:")
        print(
            f"  Col 30 (ETD/ATD): {ws.cell(4069, 30).value} - Fill: {ws.cell(4069, 30).fill.start_color.rgb if ws.cell(4069, 30).fill.start_color else 'No Fill'}"
        )
        print(
            f"  Col 31 (ETA/ATA): {ws.cell(4069, 31).value} - Fill: {ws.cell(4069, 31).fill.start_color.rgb if ws.cell(4069, 31).fill.start_color else 'No Fill'}"
        )

        # Row 4070 확인
        print("\nRow 4070 색상 확인:")
        print(
            f"  Col 30 (ETD/ATD): {ws.cell(4070, 30).value} - Fill: {ws.cell(4070, 30).fill.start_color.rgb if ws.cell(4070, 30).fill.start_color else 'No Fill'}"
        )
        print(
            f"  Col 31 (ETA/ATA): {ws.cell(4070, 31).value} - Fill: {ws.cell(4070, 31).fill.start_color.rgb if ws.cell(4070, 31).fill.start_color else 'No Fill'}"
        )

        # 전체적으로 색상이 있는 셀 찾기
        print("\n전체 색상 셀 검색:")
        colored_count = 0
        for row in range(1, min(ws.max_row + 1, 5000)):  # 처음 5000행만 확인
            for col in range(1, min(ws.max_column + 1, 50)):  # 처음 50열만 확인
                cell = ws.cell(row=row, column=col)
                if (
                    cell.fill
                    and cell.fill.start_color
                    and cell.fill.start_color.rgb != "00000000"
                ):
                    colored_count += 1
                    if colored_count <= 10:  # 처음 10개만 출력
                        print(
                            f"  Row {row}, Col {col}: {cell.value} - {cell.fill.start_color.rgb}"
                        )

        print(f"총 {colored_count}개 셀에 색상 적용됨")

    except Exception as e:
        print(f"오류: {e}")


if __name__ == "__main__":
    check_specific_colors()
