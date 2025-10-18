#!/usr/bin/env python3
"""
v2.9 동기화 검증 스크립트

CASE 280753의 MIR 날짜가 Master 값으로 업데이트되었는지 자동 검증합니다.
"""

import argparse
import json
import os
import sys
from datetime import datetime
from typing import Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def load_excel_data(
    file_path: str, case_id: str
) -> Tuple[Optional[Any], Optional[str]]:
    """Excel 파일에서 특정 CASE ID의 MIR 값을 찾아 반환"""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 헤더 찾기 (첫 10행 스캔)
        header_row = None
        case_col = None
        mir_col = None

        for row in range(1, 11):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    if "case" in cell_lower and (
                        "no" in cell_lower or cell_lower == "case"
                    ):
                        header_row = row
                        case_col = col
                    elif cell_lower in ["mir", "mir date", "mir_date"]:
                        mir_col = col

            if header_row and case_col and mir_col:
                break

        if not all([header_row, case_col, mir_col]):
            return (
                None,
                f"헤더를 찾을 수 없습니다: header_row={header_row}, case_col={case_col}, mir_col={mir_col}",
            )

        # CASE ID 검색
        for row in range(header_row + 1, ws.max_row + 1):
            case_value = ws.cell(row=row, column=case_col).value
            if str(case_value).strip() == case_id:
                mir_value = ws.cell(row=row, column=mir_col).value
                return mir_value, None

        return None, f"CASE {case_id}를 찾을 수 없습니다"

    except Exception as e:
        return None, f"파일 읽기 오류: {e}"


def check_cell_highlight(
    file_path: str, case_id: str, expected_color: str = "FFC000"
) -> Tuple[bool, Optional[str]]:
    """특정 CASE의 MIR 셀이 예상 색상으로 하이라이트되었는지 확인"""
    try:
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active

        # 헤더 찾기
        header_row = None
        case_col = None
        mir_col = None

        for row in range(1, 11):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    if "case" in cell_lower and (
                        "no" in cell_lower or cell_lower == "case"
                    ):
                        header_row = row
                        case_col = col
                    elif cell_lower in ["mir", "mir date", "mir_date"]:
                        mir_col = col

            if header_row and case_col and mir_col:
                break

        if not all([header_row, case_col, mir_col]):
            return False, "헤더를 찾을 수 없습니다"

        # CASE ID 검색 및 하이라이트 확인
        for row in range(header_row + 1, ws.max_row + 1):
            case_value = ws.cell(row=row, column=case_col).value
            if str(case_value).strip() == case_id:
                mir_cell = ws.cell(row=row, column=mir_col)
                fill = mir_cell.fill

                if hasattr(fill, "start_color") and fill.start_color:
                    actual_color = fill.start_color.rgb
                    if actual_color and actual_color.upper() == expected_color.upper():
                        return True, f"하이라이트 확인됨: {actual_color}"
                    else:
                        return (
                            False,
                            f"하이라이트 색상 불일치: 예상={expected_color}, 실제={actual_color}",
                        )
                else:
                    return False, "하이라이트가 적용되지 않음"

        return False, f"CASE {case_id}를 찾을 수 없습니다"

    except Exception as e:
        return False, f"하이라이트 확인 오류: {e}"


def analyze_stats_file(stats_path: str) -> Dict[str, Any]:
    """통계 JSON 파일 분석"""
    try:
        with open(stats_path, "r", encoding="utf-8") as f:
            stats = json.load(f)

        analysis = {
            "valid": True,
            "issues": [],
            "summary": {
                "updates": stats.get("updates", 0),
                "appends": stats.get("appends", 0),
                "new_case_count": stats.get("new_case_count", 0),
                "wh_dupe_keys": stats.get("wh_dupe_keys", 0),
                "master_rows": stats.get("master_rows", 0),
                "wh_rows": stats.get("wh_rows", 0),
            },
        }

        # 검증 로직
        if analysis["summary"]["updates"] == 0:
            analysis["issues"].append("업데이트된 셀이 0개입니다")

        if (
            analysis["summary"]["appends"] == 0
            and analysis["summary"]["new_case_count"] == 0
        ):
            analysis["issues"].append("신규 케이스가 0개입니다")

        if analysis["summary"]["wh_dupe_keys"] > 0:
            ambiguous_sample = stats.get("ambiguous_keys_sample", [])
            analysis["summary"]["ambiguous_sample"] = ambiguous_sample[:5]  # 최대 5개만

        return analysis

    except Exception as e:
        return {"valid": False, "issues": [f"통계 파일 읽기 오류: {e}"], "summary": {}}


def main():
    parser = argparse.ArgumentParser(description="v2.9 동기화 결과 검증")
    parser.add_argument("--master", required=True, help="Master Excel 파일 경로")
    parser.add_argument("--synced", required=True, help="동기화된 Excel 파일 경로")
    parser.add_argument(
        "--case-id", default="280753", help="검증할 CASE ID (기본값: 280753)"
    )
    parser.add_argument("--stats", help="통계 JSON 파일 경로 (자동 추론 시 생략 가능)")

    args = parser.parse_args()

    print("=" * 60)
    print("v2.9 동기화 검증 시작")
    print("=" * 60)

    # 1. Master 파일에서 MIR 값 확인
    print(f"\n1. Master 파일에서 CASE {args.case_id} MIR 값 확인...")
    master_mir, master_error = load_excel_data(args.master, args.case_id)
    if master_error:
        print(f"❌ Master 파일 오류: {master_error}")
        return 1

    print(f"✅ Master MIR 값: {master_mir} ({type(master_mir).__name__})")

    # 2. 동기화된 파일에서 MIR 값 확인
    print(f"\n2. 동기화된 파일에서 CASE {args.case_id} MIR 값 확인...")
    synced_mir, synced_error = load_excel_data(args.synced, args.case_id)
    if synced_error:
        print(f"❌ 동기화 파일 오류: {synced_error}")
        return 1

    print(f"✅ 동기화된 MIR 값: {synced_mir} ({type(synced_mir).__name__})")

    # 3. 값 비교
    print(f"\n3. MIR 값 동기화 확인...")
    if master_mir == synced_mir:
        print("✅ MIR 값이 정상적으로 동기화되었습니다")
    else:
        print(f"❌ MIR 값 동기화 실패: Master={master_mir}, Synced={synced_mir}")

    # 4. 하이라이트 확인
    print(f"\n4. 날짜 셀 하이라이트 확인...")
    highlighted, highlight_msg = check_cell_highlight(
        args.synced, args.case_id, "FFC000"
    )
    if highlighted:
        print(f"✅ {highlight_msg}")
    else:
        print(f"⚠️  {highlight_msg}")

    # 5. 통계 파일 분석
    stats_path = args.stats
    if not stats_path:
        # 자동 추론
        base_path = os.path.splitext(args.synced)[0]
        stats_path = f"{base_path}.sync_stats.json"

    if os.path.exists(stats_path):
        print(f"\n5. 통계 파일 분석: {stats_path}")
        stats_analysis = analyze_stats_file(stats_path)

        if stats_analysis["valid"]:
            summary = stats_analysis["summary"]
            print("✅ 통계 파일 분석 결과:")
            print(f"   - 업데이트된 셀: {summary['updates']}개")
            print(f"   - 신규 추가 케이스: {summary['appends']}개")
            print(f"   - 신규 케이스 총 개수: {summary['new_case_count']}개")
            print(f"   - 중복 CASE NO: {summary['wh_dupe_keys']}개")
            print(f"   - Master 행 수: {summary['master_rows']}개")
            print(f"   - Warehouse 행 수: {summary['wh_rows']}개")

            if "ambiguous_sample" in summary:
                print(f"   - 중복 키 샘플: {summary['ambiguous_sample']}")

            if stats_analysis["issues"]:
                print("\n⚠️  발견된 이슈:")
                for issue in stats_analysis["issues"]:
                    print(f"   - {issue}")
        else:
            print(f"❌ 통계 파일 분석 실패:")
            for issue in stats_analysis["issues"]:
                print(f"   - {issue}")
    else:
        print(f"\n⚠️  통계 파일을 찾을 수 없습니다: {stats_path}")

    print("\n" + "=" * 60)
    print("검증 완료")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
