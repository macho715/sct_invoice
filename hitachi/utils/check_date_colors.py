#!/usr/bin/env python3
"""날짜 변경 색상 확인"""

import openpyxl


def check_date_colors():
    try:
        wb = openpyxl.load_workbook("HVDC WAREHOUSE_HITACHI(HE).synced.xlsx")
        ws = wb.active

        # 날짜 컬럼 (30-44: ETD/ATD ~ AGI)
        date_cols = list(range(30, 45))

        print("날짜 컬럼에서 색상이 있는 셀 찾기...")
        colored_cells = []

        for row in range(2, min(ws.max_row + 1, 6000)):  # 처음 6000행 확인
            for col in date_cols:
                cell = ws.cell(row=row, column=col)
                if (
                    cell.fill
                    and cell.fill.start_color
                    and cell.fill.start_color.rgb not in ["00000000", "00FFFF00"]
                ):  # 검정, 노랑 제외
                    colored_cells.append(
                        (row, col, cell.value, cell.fill.start_color.rgb)
                    )

        print(f"총 {len(colored_cells)}개 날짜 셀에 색상 적용됨")

        if colored_cells:
            print("\n처음 20개:")
            for row, col, value, color in colored_cells[:20]:
                col_name = ws.cell(1, col).value
                print(f"  Row {row}, Col {col} ({col_name}): {value} - {color}")
        else:
            print("\n주황색 날짜 셀이 없습니다.")

        # 신규 케이스 확인
        print(f"\n신규 케이스 (노란색) 확인:")
        yellow_rows = []
        for row in range(2, min(ws.max_row + 1, 6000)):
            cell = ws.cell(row, 1)
            if (
                cell.fill
                and cell.fill.start_color
                and cell.fill.start_color.rgb == "00FFFF00"
            ):
                yellow_rows.append((row, cell.value))

        print(f"총 {len(yellow_rows)}개 신규 케이스 행")
        if yellow_rows:
            print("처음 10개:")
            for row, value in yellow_rows[:10]:
                print(f"  Row {row}: Case {value}")
            print(f"마지막 5개:")
            for row, value in yellow_rows[-5:]:
                print(f"  Row {row}: Case {value}")

    except Exception as e:
        print(f"오류: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    check_date_colors()
