#!/usr/bin/env python3
"""동기화된 파일의 색상 확인"""

import openpyxl


def check_synced_colors():
    try:
        wb = openpyxl.load_workbook("HVDC WAREHOUSE_HITACHI(HE).synced.xlsx")
        ws = wb.active

        print(f"총 행 수: {ws.max_row}")
        print(f"총 열 수: {ws.max_column}")

        # 헤더 확인
        print("\n헤더 (첫 번째 행):")
        for i, cell in enumerate(ws[1], 1):
            if cell.value:
                print(f"  Col {i}: {cell.value}")

        # 색상이 있는 셀 찾기
        print("\n색상이 적용된 셀들:")
        colored_cells = []
        for row in range(2, min(ws.max_row + 1, 100)):  # 처음 100행만 확인
            for col in range(1, min(ws.max_column + 1, 20)):  # 처음 20열만 확인
                cell = ws.cell(row=row, column=col)
                if (
                    cell.fill
                    and cell.fill.start_color
                    and cell.fill.start_color.rgb != "00000000"
                ):
                    colored_cells.append(
                        (row, col, cell.value, cell.fill.start_color.rgb)
                    )

        if colored_cells:
            print(f"  총 {len(colored_cells)}개 셀에 색상 적용됨")
            for row, col, value, color in colored_cells[:10]:  # 처음 10개만 출력
                print(f"    Row {row}, Col {col}: {value} - {color}")
        else:
            print("  색상이 적용된 셀이 없습니다.")

        # 마지막 몇 행 확인
        print(f"\n마지막 5개 행 색상:")
        for i in range(max(2, ws.max_row - 4), ws.max_row + 1):
            cell = ws.cell(i, 1)
            fill_color = (
                cell.fill.start_color.rgb
                if cell.fill and cell.fill.start_color
                else "No Fill"
            )
            print(f"  Row {i}: {cell.value} - Fill: {fill_color}")

    except Exception as e:
        print(f"오류: {e}")


if __name__ == "__main__":
    check_synced_colors()
