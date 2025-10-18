import openpyxl

# Excel 파일 열기
wb = openpyxl.load_workbook("HVDC WAREHOUSE_HITACHI(HE).xlsx")
ws = wb.active

print("=== 색상 확인 테스트 ===")
print(f"총 행 수: {ws.max_row}")
print()

# 마지막 10개 행의 색상 확인
print("마지막 10개 행의 색상:")
for row in range(ws.max_row - 9, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    fill_color = "None"

    if cell.fill and cell.fill.start_color:
        fill_color = (
            cell.fill.start_color.rgb
            if hasattr(cell.fill.start_color, "rgb")
            else str(cell.fill.start_color)
        )

    case_no = cell.value
    print(f"Row {row} (CASE {case_no}): Fill = {fill_color}")

print()

# 중간 부분 샘플 확인 (5550-5560)
print("중간 부분 샘플 (Row 5552-5562):")
for row in range(5552, 5563):
    cell = ws.cell(row=row, column=1)
    fill_color = "None"

    if cell.fill and cell.fill.start_color:
        fill_color = (
            cell.fill.start_color.rgb
            if hasattr(cell.fill.start_color, "rgb")
            else str(cell.fill.start_color)
        )

    case_no = cell.value
    print(f"Row {row} (CASE {case_no}): Fill = {fill_color}")
