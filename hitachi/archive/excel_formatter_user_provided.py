
"""
ExcelFormatter v2.9
- Colors changed date cells ORANGE (FFC000) using change_tracker entries (change_type == "date_update")
- Colors whole rows for new records YELLOW (FFFF00) using change_type == "new_record" and/or change_tracker.new_cases with row_index
- Tolerant to both dict- and object-style change entries
"""

from __future__ import annotations
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelFormatter:
    def __init__(self, change_tracker, orange_hex: str = "FFC000", yellow_hex: str = "FFFF00") -> None:
        self.ct = change_tracker
        self.orange = PatternFill(start_color=orange_hex, end_color=orange_hex, fill_type="solid")
        self.yellow = PatternFill(start_color=yellow_hex, end_color=yellow_hex, fill_type="solid")

    def _get(self, obj: Any, key: str, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)

    def _build_header_map(self, ws, header_row: int) -> Dict[str, int]:
        mapping = {}
        for c_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value is None:
                continue
            mapping[str(cell.value).strip()] = c_idx
        return mapping

    def apply_formatting_inplace(self, excel_file_path: str, sheet_name: str, header_row: int = 1) -> bool:
        try:
            wb = load_workbook(excel_file_path)
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]

            header_map = self._build_header_map(ws, header_row)

            # 1) Date changes → ORANGE cell
            for ch in getattr(self.ct, "changes", []) or []:
                ctype = str(self._get(ch, "change_type", ""))
                if ctype != "date_update":
                    continue

                row_index = self._get(ch, "row_index", None)
                col_name = self._get(ch, "column_name", None)
                if row_index is None or col_name is None:
                    continue

                # pandas index (0-based) + header row offset (header_row) + 1 for first data row
                excel_row = int(row_index) + header_row + 1
                # resolve column index with case-insensitive fallback
                col_idx = header_map.get(col_name)
                if col_idx is None:
                    # simple normalize fallback
                    norm = str(col_name).strip().lower().replace(" ", "_")
                    for k, v in header_map.items():
                        if norm == str(k).strip().lower().replace(" ", "_"):
                            col_idx = v
                            break
                if col_idx is None:
                    continue

                ws.cell(row=excel_row, column=col_idx).fill = self.orange

            # 2) New records → entire row YELLOW
            # Prefer explicit new_record changes with row_index
            painted_rows = set()
            for ch in getattr(self.ct, "changes", []) or []:
                if str(self._get(ch, "change_type", "")) == "new_record":
                    row_index = self._get(ch, "row_index", None)
                    if row_index is None:
                        continue
                    excel_row = int(row_index) + header_row + 1
                    for c in ws[excel_row]:
                        c.fill = self.yellow
                    painted_rows.add(excel_row)

            # Optional safety: if new_cases exists without row_index info, skip broad painting to avoid false positives.

            wb.save(excel_file_path)
            return True
        except Exception:
            return False
