
"""
Excel IO helpers: read sheets, write patches, highlight cells and rows.
"""
from dataclasses import dataclass
from typing import Dict, List, Optional, Any, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from .sync_config import HIGHLIGHT_CHANGED_DATE_HEX, HIGHLIGHT_NEW_ROW_HEX, HeaderMap

@dataclass
class SheetAccess:
    wb: Any
    ws: Any
    header_row: int
    header_map: HeaderMap

def open_target_workbook(path: str, data_only: bool = True) -> Workbook:
    return load_workbook(path, data_only=data_only)

def ensure_sheet_for_sync(wb: Workbook) -> Any:
    # For now, use the first sheet as the sync target.
    # If you have a specific sheet name, change here.
    return wb[wb.sheetnames[0]]

def apply_cell(ws, row_idx_1b: int, col_idx_1b: int, value) -> None:
    ws.cell(row=row_idx_1b, column=col_idx_1b, value=value)

def highlight_cell(ws, row_idx_1b: int, col_idx_1b: int, hex_color: str) -> None:
    ws.cell(row=row_idx_1b, column=col_idx_1b).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def highlight_entire_row(ws, row_idx_1b: int, hex_color: str) -> None:
    for cell in ws[row_idx_1b]:
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
