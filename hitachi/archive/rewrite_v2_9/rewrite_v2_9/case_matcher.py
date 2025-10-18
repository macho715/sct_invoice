"""
Build fast O(n) indexes and compute a patch plan for Master→Warehouse sync.
"""

from dataclasses import dataclass, field
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from .sync_config import (
    DATE_CANONICAL_COLUMNS,
    ALWAYS_OVERWRITE_NONDATE,
    DEFAULT_MAX_WORKERS,
    normalize_case_no,
    is_truthy,
)
from .excel_io import highlight_cell, highlight_entire_row, apply_cell
from .sync_config import HeaderMap, ALLOW_FUZZY_COLUMN_MATCH


@dataclass
class UpdateInstruction:
    row_idx_1b: int
    col_idx_1b: int
    new_value: Any
    is_date_cell: bool = False
    highlight: bool = False


@dataclass
class AppendInstruction:
    values_by_col1b: Dict[int, Any]  # target col1b -> value
    highlight_row: bool = True


@dataclass
class PatchPlan:
    updates: List[UpdateInstruction] = field(default_factory=list)
    appends: List[AppendInstruction] = field(default_factory=list)
    stats: Dict[str, Any] = field(default_factory=dict)


def _to_datetime(val) -> Optional[datetime]:
    if val is None or val == "":
        return None
    try:
        if isinstance(val, datetime):
            return val
        return datetime.fromisoformat(str(val))
    except Exception:
        # Try generic parse
        try:
            from pandas import to_datetime

            return to_datetime(val, errors="coerce").to_pydatetime()
        except Exception:
            return None


def _need_update(master_val, wh_val, is_date: bool) -> Tuple[bool, bool]:
    """
    Returns (need_update, highlight_change)
    - For date: we update whenever Master has value and the *logical date* differs or WH empty.
      We highlight only when the logical date differs (not just formatting).
    - For non-date: update when ALWAYS_OVERWRITE_NONDATE and different, or WH empty.
    """
    if not is_truthy(master_val):
        return False, False

    if is_date:
        m_dt = _to_datetime(master_val)
        w_dt = _to_datetime(wh_val)
        if w_dt is None and m_dt is not None:
            return True, True  # filled new date
        if m_dt is None:
            # Master has non-null but unparsable date; treat string compare as last resort
            if wh_val is None or str(master_val) != str(wh_val):
                return True, True
            return False, False
        # Both parsed: logical compare
        if w_dt is None or (m_dt.date() != w_dt.date() or m_dt.time() != w_dt.time()):
            return True, True
        # Logical equal -> keep but do not update or highlight
        return False, False

    # Non-date
    if ALWAYS_OVERWRITE_NONDATE:
        if wh_val is None:
            return True, True
        if str(master_val) != str(wh_val):
            return True, True
        return False, False
    else:
        return (wh_val is None), (wh_val is None)


def build_patch_plan(
    master_ws,
    master_header_row: int,
    master_hmap: HeaderMap,
    wh_ws,
    wh_header_row: int,
    wh_hmap: HeaderMap,
    date_columns: List[str] = None,
    max_workers: int = DEFAULT_MAX_WORKERS,
) -> PatchPlan:
    date_columns = date_columns or DATE_CANONICAL_COLUMNS

    # 1) Build case index for Warehouse: canonical case_no -> first row index (1-based), duplicates tracked
    wh_case_col = wh_hmap.canonical_to_index.get("case_no")
    if wh_case_col is None:
        raise KeyError("Warehouse: CASE NO column not found")

    wh_index: Dict[str, int] = {}
    wh_dupes: Dict[str, List[int]] = {}

    for r in range(wh_header_row + 1, wh_ws.max_row + 1):
        raw = wh_ws.cell(row=r, column=wh_case_col + 1).value
        key = normalize_case_no(raw)
        if not key:
            continue
        if key in wh_index:
            wh_dupes.setdefault(key, []).append(r)
            # Keep first seen as canonical row
        else:
            wh_index[key] = r

    # 2) Compute reverse header map (canonical -> col index 1-based)
    wh_col_1b: Dict[str, int] = {
        canon: idx + 1 for canon, idx in wh_hmap.canonical_to_index.items()
    }
    master_col_1b: Dict[str, int] = {
        canon: idx + 1 for canon, idx in master_hmap.canonical_to_index.items()
    }

    # Build optional fuzzy column map: headers that are literally equal ignoring case/space
    def _norm_name(n: str) -> str:
        return "".join(str(n).strip().lower().split()) if n is not None else ""

    if ALLOW_FUZZY_COLUMN_MATCH:
        # master extras
        m_extras = {}
        for idx, name in master_hmap.index_to_name.items():
            canon = None
            for k, v in master_hmap.canonical_to_index.items():
                if v == idx:
                    canon = k
                    break
            if canon is None:
                m_extras[_norm_name(name)] = idx + 1  # col1b

        # warehouse extras
        w_extras = {}
        for idx, name in wh_hmap.index_to_name.items():
            canon = None
            for k, v in wh_hmap.canonical_to_index.items():
                if v == idx:
                    canon = k
                    break
            if canon is None:
                w_extras[_norm_name(name)] = idx + 1  # col1b

        # compute intersection and add to column maps with a synthetic canonical key
        for norm, m_col1b in m_extras.items():
            w_col1b = w_extras.get(norm)
            if w_col1b:
                synth = f"col::{norm}"
                master_col_1b[synth] = m_col1b
                wh_col_1b[synth] = w_col1b

    if "case_no" not in master_col_1b:
        raise KeyError("Master: CASE NO column not found")

    # 3) Iterate master rows and produce update/append instructions
    plan = PatchPlan()
    plan.stats.update(
        updates=0,
        appends=0,
        master_rows=0,
        wh_rows=wh_ws.max_row - wh_header_row,
        wh_dupe_keys=len(wh_dupes),
        wh_dupe_rows=sum(len(v) for v in wh_dupes.values()),
        ambiguous_keys_sample=[],
        new_case_count=0,
    )
    if wh_dupes:
        # sample up to 10 dupe keys
        plan.stats["ambiguous_keys_sample"] = list(wh_dupes.keys())[:10]

    # Build a list of master rows to process for possible parallel mapping
    master_rows = list(range(master_header_row + 1, master_ws.max_row + 1))
    plan.stats["master_rows"] = len(master_rows)

    def process_row(
        r: int,
    ) -> Tuple[List[UpdateInstruction], Optional[AppendInstruction]]:
        case_val = master_ws.cell(row=r, column=master_col_1b["case_no"]).value
        key = normalize_case_no(case_val)
        if not key:
            return [], None

        # Mid-step map of canonical->value for this master row
        master_values = {}
        for canon, col1b in master_col_1b.items():
            master_values[canon] = master_ws.cell(row=r, column=col1b).value

        if key not in wh_index:
            # NEW CASE: append a row with known columns
            # Create value by aligning canonicals -> target columns
            values_by_col1b = {}
            for canon, target_col1b in wh_col_1b.items():
                if canon in master_values:
                    values_by_col1b[target_col1b] = master_values[canon]
            app = AppendInstruction(values_by_col1b=values_by_col1b, highlight_row=True)
            return [], app

        # EXISTING: produce cell updates
        updates: List[UpdateInstruction] = []
        target_row_1b = wh_index[key]
        for canon, master_val in master_values.items():
            if canon == "case_no":
                continue
            target_col1b = wh_col_1b.get(canon)
            if not target_col1b:
                continue

            wh_val = wh_ws.cell(row=target_row_1b, column=target_col1b).value
            is_date = canon in date_columns
            need, hi = _need_update(master_val, wh_val, is_date=is_date)
            if need:
                updates.append(
                    UpdateInstruction(
                        row_idx_1b=target_row_1b,
                        col_idx_1b=target_col1b,
                        new_value=master_val,
                        is_date_cell=is_date,
                        highlight=hi,
                    )
                )
        return updates, None

    # Parallel processing for computing plan
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(process_row, r) for r in master_rows]
        for fut in as_completed(futures):
            ups, app = fut.result()
            if ups:
                plan.updates.extend(ups)
            if app:
                plan.appends.append(app)

    plan.stats["updates"] = len(plan.updates)
    plan.stats["appends"] = len(plan.appends)
    plan.stats["new_case_count"] = len(plan.appends)
    return plan


def apply_patch_plan(
    ws,
    plan: PatchPlan,
    header_row: int,
    highlight_date_hex: str,
    highlight_new_row_hex: str,
) -> None:
    # 1) Apply updates
    for ins in plan.updates:
        apply_cell(ws, ins.row_idx_1b, ins.col_idx_1b, ins.new_value)
        if ins.is_date_cell and ins.highlight:
            highlight_cell(ws, ins.row_idx_1b, ins.col_idx_1b, highlight_date_hex)

    # 2) Apply appends
    for app in plan.appends:
        new_row_idx = ws.max_row + 1
        for col1b, val in app.values_by_col1b.items():
            apply_cell(ws, new_row_idx, col1b, val)
        if app.highlight_row:
            highlight_entire_row(ws, new_row_idx, highlight_new_row_hex)
