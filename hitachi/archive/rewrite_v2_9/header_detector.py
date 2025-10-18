
"""
Dynamic header detection and header row locator.
"""
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from .sync_config import CANONICAL_HEADER_PATTERNS, HEADER_SCAN_MAX_ROWS, MIN_HEADER_HITS, HeaderMap, canonical_name_of

def find_header_row_and_map(ws) -> Tuple[int, HeaderMap]:
    """
    Scan the first HEADER_SCAN_MAX_ROWS rows to find the header row.
    Return (header_row_index_1_based, HeaderMap).
    """
    best_hits = -1
    best_row = None
    best_map = None

    for r in range(1, HEADER_SCAN_MAX_ROWS + 1):
        values = [c.value if c.value is not None else "" for c in ws[r]]
        hits = 0
        hmap = HeaderMap()
        for idx, raw in enumerate(values):
            name = str(raw).strip()
            canon = canonical_name_of(name)
            if canon:
                hits += 1
                hmap.canonical_to_index[canon] = idx  # 0-based index
                hmap.index_to_name[idx] = name
        if hits > best_hits:
            best_hits = hits
            best_row = r
            best_map = hmap

    if best_hits < MIN_HEADER_HITS:
        raise RuntimeError(f"Failed to detect header row (hits={best_hits}, need >= {MIN_HEADER_HITS}).")

    return best_row, best_map


def score_sheet(ws) -> tuple:
    \"\"\"Return (hits, row_idx, HeaderMap).\"\"\"
    best_hits = -1
    best_row = None
    best_map = None
    for r in range(1, HEADER_SCAN_MAX_ROWS + 1):
        values = [c.value if c.value is not None else "" for c in ws[r]]
        hits = 0
        hmap = HeaderMap()
        for idx, raw in enumerate(values):
            name = str(raw).strip()
            canon = canonical_name_of(name)
            if canon:
                hits += 1
                hmap.canonical_to_index[canon] = idx
                hmap.index_to_name[idx] = name
        if hits > best_hits:
            best_hits = hits
            best_row = r
            best_map = hmap
    return best_hits, best_row, best_map


def pick_best_sheet(wb):
    \"\"\"Scan all sheets and return (ws, header_row, header_map).\"\"\"
    best = (-1, None, None, None)  # (hits, ws, row, map)
    for name in wb.sheetnames:
        ws = wb[name]
        hits, row, hmap = score_sheet(ws)
        if hits > best[0]:
            best = (hits, ws, row, hmap)
    hits, ws, row, hmap = best[0], best[1], best[2], best[3]
    if hits < MIN_HEADER_HITS:
        raise RuntimeError(f\"Failed to detect header on any sheet (best hits={hits}).\")
    return ws, row, hmap
