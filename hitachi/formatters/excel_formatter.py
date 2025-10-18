"""
Excel 파일 색깔 표시 및 서식 적용 시스템

이 모듈은 HVDC 데이터 동기화 후 변경사항을 Excel 파일에서
시각적으로 확인할 수 있도록 색깔과 서식을 적용하는 기능을 제공합니다.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Set, Tuple
import os
import logging
from datetime import datetime

try:
    from ..validators.change_tracker import ChangeTracker
    from .header_matcher import HeaderMatcher
    from ..core.parallel_processor import ParallelProcessor
except ImportError:
    # 직접 실행 시 fallback
    from validators.change_tracker import ChangeTracker
    from header_matcher import HeaderMatcher
    from core.parallel_processor import ParallelProcessor


class ExcelFormatter:
    """Excel 파일에 변경사항을 색깔로 표시하는 포맷터"""

    def __init__(self, change_tracker: ChangeTracker, max_workers: int = None):
        self.change_tracker = change_tracker
        self.header_matcher = HeaderMatcher()
        self.parallel_processor = ParallelProcessor(max_workers)
        self.logger = logging.getLogger(__name__)

        # 색상 정의 (openpyxl 형식)
        self.colors = {
            "new_case": {
                "fill": PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                ),  # 노란색 (사용자 요구사항)
                "font": Font(color="000000", bold=True),  # 검은색, 볼드
            },
            "high_priority_date": {
                "fill": PatternFill(
                    start_color="FFC000", end_color="FFC000", fill_type="solid"
                ),  # 주황색 (사용자 요구사항)
                "font": Font(color="000000", bold=True),  # 검은색, 볼드
            },
            "medium_priority_date": {
                "fill": PatternFill(
                    start_color="F0FFF0", end_color="F0FFF0", fill_type="solid"
                ),  # 연한 초록색
                "font": Font(color="32CD32", bold=True),  # 초록색, 볼드
            },
            "low_priority_date": {
                "fill": PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                ),  # 연한 회색
                "font": Font(color="808080"),  # 회색
            },
            "case_no_changed": {
                "fill": PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                ),  # 노란색
                "font": Font(color="000000", bold=True),  # 검은색, 볼드
            },
        }

        # 테두리 스타일
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def apply_formatting_to_excel(
        self,
        excel_file_path: str,
        case_column: str = "Case No.",
        sheet_name: Optional[str] = None,
    ) -> str:
        """
        Excel 파일에 변경사항 색깔 표시 적용 (병렬 처리)

        Args:
            excel_file_path: Excel 파일 경로
            case_column: Case No. 컬럼명 (동적 헤더 매칭 사용)
            sheet_name: 대상 시트명 (None이면 첫 번째 시트)

        Returns:
            색깔이 적용된 새 Excel 파일 경로
        """
        try:
            # Excel 파일 로드
            workbook = load_workbook(excel_file_path)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    self.logger.error(f"시트 '{sheet_name}'를 찾을 수 없습니다.")
                    return excel_file_path
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            self.logger.info(
                f"Excel 서식 적용 시작: {os.path.basename(excel_file_path)}"
            )

            # 데이터프레임으로 읽어서 케이스 번호와 행 매핑
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name or 0)

            # 동적 헤더 매칭으로 CASE NO 컬럼 찾기
            if not case_column or case_column == "Case No.":
                case_column = self.header_matcher.find_column(df.columns, "case_no")
                if not case_column:
                    self.logger.error("CASE NO 컬럼을 찾을 수 없습니다.")
                    return excel_file_path

            case_to_row = self._create_case_to_row_mapping(df, case_column)

            # 병렬로 색깔 적용
            colored_cases = self._apply_formatting_parallel(worksheet, df, case_to_row)

            # 범례 추가
            self._add_legend(worksheet, df.shape[0] + 5)

            # 새 파일명 생성 (색깔 적용됨을 표시)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(excel_file_path)[0]
            colored_file_path = f"{base_name}_colored_{timestamp}.xlsx"

            # 저장
            workbook.save(colored_file_path)
            workbook.close()

            self.logger.info(f"색깔 적용 완료: {os.path.basename(colored_file_path)}")

            # 결과 요약 출력
            self._print_formatting_summary(colored_cases)

            return colored_file_path

        except Exception as e:
            self.logger.error(f"Excel 서식 적용 오류: {str(e)}")
            return excel_file_path

    def apply_formatting_inplace(
        self,
        excel_file_path: str,
        case_column: str = "Case No.",
        sheet_name: Optional[str] = None,
    ) -> bool:
        """
        Excel 파일에 변경사항 색상 표시 적용 (원본 파일 직접 수정)

        Args:
            excel_file_path: Excel 파일 경로
            case_column: Case No. 컬럼명
            sheet_name: 대상 시트명

        Returns:
            성공 여부
        """
        try:
            # Excel 파일 로드
            workbook = load_workbook(excel_file_path)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    self.logger.error(f"시트 '{sheet_name}'를 찾을 수 없습니다.")
                    return False
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            self.logger.info(
                f"Excel 서식 적용 시작 (원본 수정): {os.path.basename(excel_file_path)}"
            )

            # 데이터프레임으로 읽어서 케이스 번호와 행 매핑
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name or 0)

            # 동적 헤더 매칭으로 CASE NO 컬럼 찾기
            if not case_column or case_column == "Case No.":
                print(f"[DEBUG] CASE NO 컬럼 찾기 중...")
                print(f"[DEBUG] 사용 가능한 컬럼: {list(df.columns)[:10]}...")
                case_column = self.header_matcher.find_column(df.columns, "case_no")
                print(f"[DEBUG] 찾은 CASE NO 컬럼: {case_column}")
                if not case_column:
                    self.logger.error("CASE NO 컬럼을 찾을 수 없습니다.")
                    return False

            case_to_row = self._create_case_to_row_mapping(df, case_column)

            # 색상 적용
            colored_cases = self._apply_formatting_parallel(worksheet, df, case_to_row)

            # 범례 추가
            self._add_legend(worksheet, df.shape[0] + 5)

            # 원본 파일에 직접 저장
            workbook.save(excel_file_path)
            workbook.close()

            self.logger.info(f"색상 적용 완료: {os.path.basename(excel_file_path)}")
            self._print_formatting_summary(colored_cases)

            return True

        except Exception as e:
            self.logger.error(f"Excel 서식 적용 오류: {str(e)}")
            return False

    def _apply_formatting_parallel(self, worksheet, df, case_to_row):
        """병렬로 색깔 적용"""

        def create_style_batch(changes_batch):
            """스타일 정보 생성 (병렬)"""
            styles = []

            for change in changes_batch:
                col_idx = self._find_column_index(df.columns, change.column_name)
                if col_idx:
                    # case_to_row 매핑을 사용하여 실제 Excel 행 번호 찾기
                    if change.case_no in case_to_row:
                        row_num = (
                            case_to_row[change.case_no] + 2
                        )  # Excel은 1부터 시작, 헤더 고려
                    else:
                        # 매핑이 없으면 DataFrame 인덱스 사용
                        row_num = change.row_index + 2
                    print(
                        f"[DEBUG] 스타일 생성: {change.case_no} - {change.column_name} (Row {row_num}, Col {col_idx}) - {change.change_type}"
                    )

                    if change.change_type == "date_update":
                        styles.append(
                            {
                                "row": row_num,
                                "col": col_idx,
                                "fill": PatternFill(
                                    start_color="FFC000",  # 주황색
                                    end_color="FFC000",
                                    fill_type="solid",
                                ),
                                "font": Font(bold=True),
                            }
                        )
                    elif change.change_type == "field_update":
                        # 일반 필드 업데이트는 연한 파란색
                        styles.append(
                            {
                                "row": row_num,
                                "col": col_idx,
                                "fill": PatternFill(
                                    start_color="E6F3FF",  # 연한 파란색
                                    end_color="E6F3FF",
                                    fill_type="solid",
                                ),
                            }
                        )
                    elif change.change_type == "new_record":
                        # 신규 레코드는 전체 행을 노란색으로
                        for c_idx in range(1, len(df.columns) + 1):
                            styles.append(
                                {
                                    "row": row_num,
                                    "col": c_idx,
                                    "fill": PatternFill(
                                        start_color="FFFF00",  # 노란색
                                        end_color="FFFF00",
                                        fill_type="solid",
                                    ),
                                }
                            )

            return styles

        # 디버그: ChangeTracker 상태 확인
        print(f"[DEBUG] ChangeTracker 상태:")
        print(f"  - 총 변경사항: {len(self.change_tracker.changes)}")
        print(f"  - 신규 케이스: {len(self.change_tracker.get_new_cases())}")
        print(f"  - 날짜 변경: {len(self.change_tracker.date_changes)}")

        if self.change_tracker.changes:
            print(f"  - 첫 번째 변경사항: {self.change_tracker.changes[0]}")

        # 변경사항을 배치로 나누어 병렬 처리
        batch_size = max(
            100,
            len(self.change_tracker.changes)
            // (self.parallel_processor.max_workers * 2),
        )

        style_batches = self.parallel_processor.process_batches(
            self.change_tracker.changes,
            batch_size,
            create_style_batch,
            use_threads=True,
        )

        # 스타일 적용 (순차 - openpyxl은 thread-safe하지 않음)
        colored_cases = {
            "new_cases": 0,
            "high_priority": 0,
            "medium_priority": 0,
            "low_priority": 0,
        }

        print(f"[DEBUG] 스타일 배치 수: {len(style_batches)}")
        total_styles = sum(len(batch) for batch in style_batches)
        print(f"[DEBUG] 총 스타일 수: {total_styles}")

        for batch_idx, batch in enumerate(style_batches):
            print(f"[DEBUG] 배치 {batch_idx}: {len(batch)}개 스타일")
            for style_idx, style_info in enumerate(batch):
                if style_idx < 3:  # 처음 3개만 디버그 출력
                    print(
                        f"[DEBUG] 스타일 적용: Row {style_info['row']}, Col {style_info['col']}, Fill {style_info['fill'].start_color.rgb}"
                    )
                cell = worksheet.cell(row=style_info["row"], column=style_info["col"])
                cell.fill = style_info["fill"]
                if "font" in style_info:
                    cell.font = style_info["font"]

                # 통계 업데이트
                if style_info["fill"].start_color == "FFFF00":
                    colored_cases["new_cases"] += 1
                elif style_info["fill"].start_color == "FFC000":
                    colored_cases["high_priority"] += 1

        return colored_cases

    def _find_column_index(self, columns, target_col):
        """대소문자 무관 컬럼 인덱스 찾기"""
        if target_col in columns:
            return columns.get_loc(target_col) + 1  # Excel은 1부터 시작

        target_normalized = self.header_matcher.normalize_header(target_col)
        for i, col in enumerate(columns):
            if self.header_matcher.normalize_header(col) == target_normalized:
                return i + 1
        return None

    def _create_case_to_row_mapping(
        self, df: pd.DataFrame, case_column: str
    ) -> Dict[str, int]:
        """케이스 번호와 Excel 행 번호 매핑 생성"""
        case_to_row = {}

        for idx, case_no in enumerate(df[case_column]):
            if pd.notna(case_no):
                # Excel에서는 1부터 시작, 헤더 고려하여 +2
                case_to_row[str(case_no).strip()] = idx + 2

        # 디버그: 매핑 상태 확인
        print(f"[DEBUG] case_to_row 매핑 생성: {len(case_to_row)}개")
        if case_to_row:
            sample_cases = list(case_to_row.items())[:5]
            print(f"[DEBUG] 샘플 매핑: {sample_cases}")

        return case_to_row

    def _apply_case_colors(
        self, worksheet, case_to_row: Dict[str, int]
    ) -> Dict[str, int]:
        """케이스별 색깔 적용"""
        colored_cases = {
            "new_cases": 0,
            "high_priority": 0,
            "medium_priority": 0,
            "low_priority": 0,
        }

        # 신규 케이스 - 파란색
        new_cases = self.change_tracker.get_new_cases()
        for case_no in new_cases:
            if case_no in case_to_row:
                row_num = case_to_row[case_no]
                self._apply_row_style(worksheet, row_num, "new_case")
                colored_cases["new_cases"] += 1

        # 날짜 변경된 케이스 - 우선순위별 색상
        for case_no, changes in self.change_tracker.date_changes.items():
            if case_no in case_to_row and case_no not in new_cases:
                row_num = case_to_row[case_no]

                # 가장 높은 우선순위 결정
                priority = self._determine_case_priority(changes)
                color_key = f"{priority}_date"

                self._apply_row_style(worksheet, row_num, color_key)

                if priority == "high_priority":
                    colored_cases["high_priority"] += 1
                elif priority == "medium_priority":
                    colored_cases["medium_priority"] += 1
                else:
                    colored_cases["low_priority"] += 1

        return colored_cases

    def _apply_column_colors(
        self, worksheet, df: pd.DataFrame, case_to_row: Dict[str, int]
    ):
        """특정 컬럼의 변경사항에 대해 개별 셀 색깔 적용"""
        for change in self.change_tracker.changes:
            case_no = change.case_no
            if case_no in case_to_row:
                row_num = case_to_row[case_no]

                # 컬럼명으로 컬럼 인덱스 찾기
                try:
                    col_idx = df.columns.get_loc(change.column_name)
                    col_letter = get_column_letter(col_idx + 1)

                    cell = worksheet[f"{col_letter}{row_num}"]

                    # 변경 타입에 따른 추가 스타일 적용
                    if change.change_type == "date_update":
                        # 테두리 강조
                        cell.border = Border(
                            left=Side(style="thick", color="FF8C00"),
                            right=Side(style="thick", color="FF8C00"),
                            top=Side(style="thick", color="FF8C00"),
                            bottom=Side(style="thick", color="FF8C00"),
                        )

                except (KeyError, ValueError):
                    # 컬럼을 찾을 수 없는 경우 무시
                    pass

    def _apply_row_style(self, worksheet, row_num: int, style_key: str):
        """행 전체에 스타일 적용"""
        if style_key not in self.colors:
            return

        style = self.colors[style_key]

        # 현재 행의 모든 셀에 스타일 적용
        for cell in worksheet[row_num]:
            if cell.value is not None:  # 값이 있는 셀만
                cell.fill = style["fill"]
                cell.font = style["font"]
                cell.border = self.border

    def _determine_case_priority(self, changes: List) -> str:
        """케이스의 변경사항들 중 가장 높은 우선순위 결정"""
        priorities = ["high_priority", "medium_priority", "low_priority"]

        for priority in priorities:
            if any(change.priority == priority for change in changes):
                return priority

        return "low_priority"

    def _add_legend(self, worksheet, start_row: int):
        """범례 추가"""
        legend_data = [
            ("색상 범례", "설명"),
            ("신규 케이스", "새로 추가된 케이스 (파란색)"),
            ("고우선순위 날짜변경", "창고/현장별 중요 날짜 변경 (주황색)"),
            ("중우선순위 날짜변경", "일반 날짜 변경 (초록색)"),
            ("저우선순위 날짜변경", "기타 날짜 변경 (회색)"),
        ]

        # 범례 헤더
        worksheet.cell(row=start_row, column=1, value="🎨 변경사항 색상 범례")
        worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14)

        # 범례 항목들
        for i, (category, description) in enumerate(legend_data):
            row = start_row + i + 2
            worksheet.cell(row=row, column=1, value=category)
            worksheet.cell(row=row, column=2, value=description)

            # 해당 색상 적용
            if i > 0:  # 헤더 제외
                color_keys = [
                    "new_case",
                    "high_priority_date",
                    "medium_priority_date",
                    "low_priority_date",
                ]
                if i - 1 < len(color_keys):
                    style = self.colors[color_keys[i - 1]]
                    worksheet.cell(row=row, column=1).fill = style["fill"]
                    worksheet.cell(row=row, column=1).font = style["font"]

    def _print_formatting_summary(self, colored_cases: Dict[str, int]):
        """서식 적용 결과 요약 출력"""
        total_colored = sum(colored_cases.values())

        print("\n" + "=" * 60)
        print("🎨 Excel 색깔 표시 완료")
        print("=" * 60)
        print(f"📊 색깔 적용된 케이스: {total_colored:,}개")
        print(f"  🔵 신규 케이스: {colored_cases['new_cases']:,}개")
        print(f"  🟠 고우선순위 날짜변경: {colored_cases['high_priority']:,}개")
        print(f"  🟢 중우선순위 날짜변경: {colored_cases['medium_priority']:,}개")
        print(f"  ⚪ 저우선순위 날짜변경: {colored_cases['low_priority']:,}개")
        print("=" * 60)

    def create_change_summary_sheet(
        self, excel_file_path: str, output_path: Optional[str] = None
    ) -> str:
        """
        변경사항 요약 시트를 별도로 생성

        Args:
            excel_file_path: 원본 Excel 파일 경로
            output_path: 출력 파일 경로 (None이면 자동 생성)

        Returns:
            생성된 요약 파일 경로
        """
        try:
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_name = os.path.splitext(excel_file_path)[0]
                output_path = f"{base_name}_change_summary_{timestamp}.xlsx"

            # 변경사항 데이터 준비
            changes_data = []
            for change in self.change_tracker.changes:
                changes_data.append(
                    {
                        "Case No.": change.case_no,
                        "Column": change.column_name,
                        "Before": change.old_value,
                        "After": change.new_value,
                        "Change Type": change.change_type,
                        "Priority": change.priority,
                        "Timestamp": change.timestamp,
                    }
                )

            # 데이터프레임 생성
            changes_df = pd.DataFrame(changes_data)

            # 요약 통계
            summary_stats = self.change_tracker.generate_summary()

            # Excel 파일로 저장
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # 변경사항 상세 시트
                changes_df.to_excel(writer, sheet_name="변경사항 상세", index=False)

                # 요약 통계 시트
                summary_df = pd.DataFrame([summary_stats])
                summary_df.to_excel(writer, sheet_name="요약 통계", index=False)

                # 새 케이스 목록 시트
                new_cases_df = pd.DataFrame(
                    {"New Cases": list(self.change_tracker.get_new_cases())}
                )
                new_cases_df.to_excel(writer, sheet_name="신규 케이스", index=False)

            self.logger.info(
                f"변경사항 요약 파일 생성: {os.path.basename(output_path)}"
            )
            return output_path

        except Exception as e:
            self.logger.error(f"요약 시트 생성 오류: {str(e)}")
            return ""


def apply_hvdc_formatting(
    excel_file_path: str,
    change_tracker: ChangeTracker,
    case_column: str = "Case No.",
    create_summary: bool = True,
) -> Tuple[str, str]:
    """
    HVDC Excel 파일에 변경사항 색깔 표시 적용 (편의 함수)

    Args:
        excel_file_path: Excel 파일 경로
        change_tracker: 변경사항 추적기
        case_column: Case No. 컬럼명
        create_summary: 요약 시트 생성 여부

    Returns:
        (색깔_적용된_파일_경로, 요약_파일_경로)
    """
    formatter = ExcelFormatter(change_tracker)

    # 색깔 적용
    colored_file = formatter.apply_formatting_to_excel(
        excel_file_path=excel_file_path, case_column=case_column
    )

    # 요약 시트 생성
    summary_file = ""
    if create_summary:
        summary_file = formatter.create_change_summary_sheet(excel_file_path)

    return colored_file, summary_file
