#!/usr/bin/env python3
"""
Excel 시트 구조 확인 스크립트
"""

import openpyxl


def check_sheet_structure(file_path: str):
    """Excel 파일의 시트 구조 확인"""
    wb = openpyxl.load_workbook(file_path)

    print(f"File: {file_path}")
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")

        # 첫 5행 출력
        for row in range(1, min(6, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(11, sheet.max_column + 1)):  # 첫 10컬럼만
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value)[:20] if cell_value else "")
            print(f"Row {row}: {row_data}")


if __name__ == "__main__":
    check_sheet_structure("report_input.xlsx")
