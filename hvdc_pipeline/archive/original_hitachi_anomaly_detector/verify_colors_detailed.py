# -*- coding: utf-8 -*-
"""
색상 적용 상태 상세 검증 스크립트
- Excel 파일의 색상 적용 상태를 정확히 분석
- 색상별 행 수, Case ID 매칭 상태 확인
"""
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import argparse
import json
from pathlib import Path


def verify_colors(excel_file: str, sheet_name: str = "Case List", json_file: str = None):
    """색상 적용 상태 상세 검증"""
    print(f"🔍 색상 검증 시작: {excel_file}")
    print(f"📋 시트: {sheet_name}")
    
    # Excel 파일 로드
    try:
        wb = openpyxl.load_workbook(excel_file)
        if sheet_name not in wb.sheetnames:
            print(f"❌ 시트 '{sheet_name}'을 찾을 수 없습니다.")
            print(f"사용 가능한 시트: {wb.sheetnames}")
            return
        ws = wb[sheet_name]
    except Exception as e:
        print(f"❌ Excel 파일 로드 실패: {e}")
        return
    
    print(f"📊 총 행 수: {ws.max_row}")
    print(f"📊 총 열 수: {ws.max_column}")
    
    # 색상별 행 수 카운트
    color_counts = defaultdict(int)
    colored_rows = []
    case_id_colors = {}  # Case ID별 색상 매핑
    
    # Case NO 컬럼 찾기
    case_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and "case" in str(cell_value).lower():
            case_col_idx = col_idx
            print(f"🔍 Case NO 컬럼 발견: {col_idx}번째 ('{cell_value}')")
            break
    
    if case_col_idx is None:
        print("❌ Case NO 컬럼을 찾을 수 없습니다.")
        return
    
    # 각 행의 색상 분석
    for row_num in range(2, ws.max_row + 1):
        row_colors = set()
        has_color = False
        
        # Case ID 가져오기
        case_id = ws.cell(row=row_num, column=case_col_idx).value
        case_id_str = str(case_id).strip() if case_id else ""
        
        # 각 셀의 색상 확인
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.fill:
                # 안전 파서: fg/start/end 중 하나라도 RGB면 포착
                color = getattr(cell.fill, "fgColor", None)
                rgb = getattr(color, "rgb", None) if color else None
                if not rgb and getattr(cell.fill, "start_color", None):
                    rgb = cell.fill.start_color.rgb
                if not rgb and getattr(cell.fill, "end_color", None):
                    rgb = cell.fill.end_color.rgb
                if rgb and isinstance(rgb, str):
                    color_str = rgb.upper()
                    # 6자리면 FF 접두로 보강
                    if len(color_str) == 6:
                        color_str = "FF" + color_str
                    row_colors.add(color_str)
                    has_color = True
        
        if has_color:
            colored_rows.append((row_num, row_colors, case_id_str))
            for color in row_colors:
                color_counts[color] += 1
            
            # Case ID별 색상 매핑
            if case_id_str:
                case_id_colors[case_id_str] = list(row_colors)
    
    # 결과 출력
    print(f"\n📊 색상 검증 결과:")
    print(f"  - 총 색칠된 행: {len(colored_rows)}개")
    print(f"  - 전체 행 대비: {len(colored_rows)}/{ws.max_row-1} ({len(colored_rows)/(ws.max_row-1)*100:.1f}%)")
    
    print(f"\n🎨 색상별 카운트:")
    # ARGB(FF/00) 동시 허용
    color_names = {
        "FFFF0000": "빨강 (시간 역전)",
        "00FF0000": "빨강 (시간 역전)",
        "FFFFC000": "주황 (ML 이상치-높음)",
        "00FFC000": "주황 (ML 이상치-높음)",
        "FFFFFF00": "노랑 (ML 이상치-보통)",
        "00FFFF00": "노랑 (ML 이상치-보통)",
        "FFCC99FF": "보라 (데이터 품질)",
        "00CC99FF": "보라 (데이터 품질)",
    }
    
    for color, count in sorted(color_counts.items()):
        color_name = color_names.get(color, f"기타 ({color})")
        print(f"  - {color}: {count}개 ({color_name})")
    
    # 색칠된 행 샘플 출력
    print(f"\n📋 색칠된 행 샘플 (처음 10개):")
    for i, (row_num, colors, case_id) in enumerate(colored_rows[:10]):
        color_list = [color_names.get(c, c) for c in colors]
        print(f"  - 행 {row_num}: Case ID '{case_id}' → {', '.join(color_list)}")
    
    if len(colored_rows) > 10:
        print(f"  ... (총 {len(colored_rows)}개 행 중 10개만 표시)")
    
    # JSON 파일과 비교 (제공된 경우)
    if json_file and Path(json_file).exists():
        print(f"\n🔍 JSON 파일과 비교: {json_file}")
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                anomalies = json.load(f)
            
            print(f"  - JSON 이상치 수: {len(anomalies)}건")
            print(f"  - Excel 색칠 행 수: {len(colored_rows)}건")
            
            # Case ID 매칭 확인
            json_case_ids = set()
            for anomaly in anomalies:
                case_id = anomaly.get('Case_ID', '')
                if case_id:
                    json_case_ids.add(str(case_id).strip().upper())
            
            excel_case_ids = set()
            for _, _, case_id in colored_rows:
                if case_id:
                    excel_case_ids.add(case_id.upper())
            
            matched = json_case_ids.intersection(excel_case_ids)
            print(f"  - 매칭된 Case ID: {len(matched)}개")
            print(f"  - JSON만 있는 Case ID: {len(json_case_ids - excel_case_ids)}개")
            print(f"  - Excel만 있는 Case ID: {len(excel_case_ids - json_case_ids)}개")
            
            if len(json_case_ids - excel_case_ids) > 0:
                print(f"  - 매칭 실패 Case ID 샘플: {list(json_case_ids - excel_case_ids)[:5]}")
            
        except Exception as e:
            print(f"❌ JSON 파일 분석 실패: {e}")
    
    # 문제 진단
    print(f"\n🔍 문제 진단:")
    
    if len(colored_rows) == 0:
        print("  ❌ 색상이 전혀 적용되지 않았습니다.")
        print("  💡 해결방법: anomaly_detector.py --visualize 옵션으로 재실행")
    elif len(colored_rows) == ws.max_row - 1:
        print("  ⚠️ 모든 행이 색칠되었습니다. (범례가 데이터에 영향을 준 가능성)")
        print("  💡 해결방법: add_color_legend()를 별도 시트에 작성하도록 변경 필요")
    elif len(colored_rows) < 500:
        print("  ⚠️ 예상보다 적은 행이 색칠되었습니다.")
        print("  💡 해결방법: Case ID 매칭 로직 확인 필요")
    else:
        print("  ✅ 색상 적용이 정상적으로 보입니다.")
    
    return {
        "total_rows": ws.max_row - 1,
        "colored_rows": len(colored_rows),
        "color_counts": dict(color_counts),
        "colored_cases": case_id_colors
    }


def main():
    parser = argparse.ArgumentParser(description="Excel 파일 색상 적용 상태 검증")
    parser.add_argument("--excel", required=True, help="Excel 파일 경로")
    parser.add_argument("--sheet", default="Case List", help="시트명")
    parser.add_argument("--json", help="비교할 JSON 파일 경로")
    
    args = parser.parse_args()
    
    result = verify_colors(args.excel, args.sheet, args.json)
    
    if result:
        print(f"\n✅ 검증 완료!")
    else:
        print(f"\n❌ 검증 실패!")


if __name__ == "__main__":
    main()

색상 적용 상태 상세 검증 스크립트
- Excel 파일의 색상 적용 상태를 정확히 분석
- 색상별 행 수, Case ID 매칭 상태 확인
"""
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import argparse
import json
from pathlib import Path


def verify_colors(excel_file: str, sheet_name: str = "Case List", json_file: str = None):
    """색상 적용 상태 상세 검증"""
    print(f"🔍 색상 검증 시작: {excel_file}")
    print(f"📋 시트: {sheet_name}")
    
    # Excel 파일 로드
    try:
        wb = openpyxl.load_workbook(excel_file)
        if sheet_name not in wb.sheetnames:
            print(f"❌ 시트 '{sheet_name}'을 찾을 수 없습니다.")
            print(f"사용 가능한 시트: {wb.sheetnames}")
            return
        ws = wb[sheet_name]
    except Exception as e:
        print(f"❌ Excel 파일 로드 실패: {e}")
        return
    
    print(f"📊 총 행 수: {ws.max_row}")
    print(f"📊 총 열 수: {ws.max_column}")
    
    # 색상별 행 수 카운트
    color_counts = defaultdict(int)
    colored_rows = []
    case_id_colors = {}  # Case ID별 색상 매핑
    
    # Case NO 컬럼 찾기
    case_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and "case" in str(cell_value).lower():
            case_col_idx = col_idx
            print(f"🔍 Case NO 컬럼 발견: {col_idx}번째 ('{cell_value}')")
            break
    
    if case_col_idx is None:
        print("❌ Case NO 컬럼을 찾을 수 없습니다.")
        return
    
    # 각 행의 색상 분석
    for row_num in range(2, ws.max_row + 1):
        row_colors = set()
        has_color = False
        
        # Case ID 가져오기
        case_id = ws.cell(row=row_num, column=case_col_idx).value
        case_id_str = str(case_id).strip() if case_id else ""
        
        # 각 셀의 색상 확인
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.fill:
                # 안전 파서: fg/start/end 중 하나라도 RGB면 포착
                color = getattr(cell.fill, "fgColor", None)
                rgb = getattr(color, "rgb", None) if color else None
                if not rgb and getattr(cell.fill, "start_color", None):
                    rgb = cell.fill.start_color.rgb
                if not rgb and getattr(cell.fill, "end_color", None):
                    rgb = cell.fill.end_color.rgb
                if rgb and isinstance(rgb, str):
                    color_str = rgb.upper()
                    # 6자리면 FF 접두로 보강
                    if len(color_str) == 6:
                        color_str = "FF" + color_str
                    row_colors.add(color_str)
                    has_color = True
        
        if has_color:
            colored_rows.append((row_num, row_colors, case_id_str))
            for color in row_colors:
                color_counts[color] += 1
            
            # Case ID별 색상 매핑
            if case_id_str:
                case_id_colors[case_id_str] = list(row_colors)
    
    # 결과 출력
    print(f"\n📊 색상 검증 결과:")
    print(f"  - 총 색칠된 행: {len(colored_rows)}개")
    print(f"  - 전체 행 대비: {len(colored_rows)}/{ws.max_row-1} ({len(colored_rows)/(ws.max_row-1)*100:.1f}%)")
    
    print(f"\n🎨 색상별 카운트:")
    # ARGB(FF/00) 동시 허용
    color_names = {
        "FFFF0000": "빨강 (시간 역전)",
        "00FF0000": "빨강 (시간 역전)",
        "FFFFC000": "주황 (ML 이상치-높음)",
        "00FFC000": "주황 (ML 이상치-높음)",
        "FFFFFF00": "노랑 (ML 이상치-보통)",
        "00FFFF00": "노랑 (ML 이상치-보통)",
        "FFCC99FF": "보라 (데이터 품질)",
        "00CC99FF": "보라 (데이터 품질)",
    }
    
    for color, count in sorted(color_counts.items()):
        color_name = color_names.get(color, f"기타 ({color})")
        print(f"  - {color}: {count}개 ({color_name})")
    
    # 색칠된 행 샘플 출력
    print(f"\n📋 색칠된 행 샘플 (처음 10개):")
    for i, (row_num, colors, case_id) in enumerate(colored_rows[:10]):
        color_list = [color_names.get(c, c) for c in colors]
        print(f"  - 행 {row_num}: Case ID '{case_id}' → {', '.join(color_list)}")
    
    if len(colored_rows) > 10:
        print(f"  ... (총 {len(colored_rows)}개 행 중 10개만 표시)")
    
    # JSON 파일과 비교 (제공된 경우)
    if json_file and Path(json_file).exists():
        print(f"\n🔍 JSON 파일과 비교: {json_file}")
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                anomalies = json.load(f)
            
            print(f"  - JSON 이상치 수: {len(anomalies)}건")
            print(f"  - Excel 색칠 행 수: {len(colored_rows)}건")
            
            # Case ID 매칭 확인
            json_case_ids = set()
            for anomaly in anomalies:
                case_id = anomaly.get('Case_ID', '')
                if case_id:
                    json_case_ids.add(str(case_id).strip().upper())
            
            excel_case_ids = set()
            for _, _, case_id in colored_rows:
                if case_id:
                    excel_case_ids.add(case_id.upper())
            
            matched = json_case_ids.intersection(excel_case_ids)
            print(f"  - 매칭된 Case ID: {len(matched)}개")
            print(f"  - JSON만 있는 Case ID: {len(json_case_ids - excel_case_ids)}개")
            print(f"  - Excel만 있는 Case ID: {len(excel_case_ids - json_case_ids)}개")
            
            if len(json_case_ids - excel_case_ids) > 0:
                print(f"  - 매칭 실패 Case ID 샘플: {list(json_case_ids - excel_case_ids)[:5]}")
            
        except Exception as e:
            print(f"❌ JSON 파일 분석 실패: {e}")
    
    # 문제 진단
    print(f"\n🔍 문제 진단:")
    
    if len(colored_rows) == 0:
        print("  ❌ 색상이 전혀 적용되지 않았습니다.")
        print("  💡 해결방법: anomaly_detector.py --visualize 옵션으로 재실행")
    elif len(colored_rows) == ws.max_row - 1:
        print("  ⚠️ 모든 행이 색칠되었습니다. (범례가 데이터에 영향을 준 가능성)")
        print("  💡 해결방법: add_color_legend()를 별도 시트에 작성하도록 변경 필요")
    elif len(colored_rows) < 500:
        print("  ⚠️ 예상보다 적은 행이 색칠되었습니다.")
        print("  💡 해결방법: Case ID 매칭 로직 확인 필요")
    else:
        print("  ✅ 색상 적용이 정상적으로 보입니다.")
    
    return {
        "total_rows": ws.max_row - 1,
        "colored_rows": len(colored_rows),
        "color_counts": dict(color_counts),
        "colored_cases": case_id_colors
    }


def main():
    parser = argparse.ArgumentParser(description="Excel 파일 색상 적용 상태 검증")
    parser.add_argument("--excel", required=True, help="Excel 파일 경로")
    parser.add_argument("--sheet", default="Case List", help="시트명")
    parser.add_argument("--json", help="비교할 JSON 파일 경로")
    
    args = parser.parse_args()
    
    result = verify_colors(args.excel, args.sheet, args.json)
    
    if result:
        print(f"\n✅ 검증 완료!")
    else:
        print(f"\n❌ 검증 실패!")


if __name__ == "__main__":
    main()
