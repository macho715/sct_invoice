#!/usr/bin/env python3
import json

# 이상치 데이터 로드
with open("hvdc_anomaly_report_new.json", "r", encoding="utf-8") as f:
    data = json.load(f)

print("Sample Case IDs from anomaly data:")
for i, item in enumerate(data[:5]):
    print(f"  {i}: {item['Case_ID']}")

print(f"\nTotal anomalies: {len(data)}")

# Excel 데이터 확인
import openpyxl

wb = openpyxl.load_workbook("report_input.xlsx")
sheet = None
for sheet_name in wb.sheetnames:
    if "통합_원본데이터_Fixed" in sheet_name:
        sheet = wb[sheet_name]
        break

if sheet:
    print(f"\nSample Case IDs from Excel (first 5 rows):")
    for row in range(2, 7):
        case_id = sheet.cell(row=row, column=1).value
        print(f"  Row {row}: {case_id}")
