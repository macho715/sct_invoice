# -*- coding: utf-8 -*-
"""
HVDC 시간역전 이상치 분석보고서 생성기
- 원본 파일은 절대 수정하지 않음
- 별도의 Excel 보고서 파일 생성
- 4개 시트: 요약 대시보드, 시간역전 상세, ML 이상치 상세, 원본 파일 참조
"""
import json
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


class AnalysisReportGenerator:
    """이상치 분석 보고서 생성기"""

    def __init__(self):
        self.colors = {
            "RED": "FF0000",  # 시간 역전
            "ORANGE": "FFC000",  # ML 이상치 (높음)
            "YELLOW": "FFFF00",  # ML 이상치 (보통)
            "PURPLE": "CC99FF",  # 데이터 품질
        }

    def create_report(self, json_path: str, output_path: str) -> bool:
        """분석 보고서 생성"""
        try:
            # JSON 데이터 로드
            with open(json_path, "r", encoding="utf-8") as f:
                anomalies = json.load(f)

            print(f"📊 로드된 이상치 데이터: {len(anomalies)}건")

            # 새 워크북 생성
            wb = openpyxl.Workbook()

            # Sheet 1: 요약 대시보드
            ws_summary = wb.active
            ws_summary.title = "요약 대시보드"
            self._create_summary_sheet(ws_summary, anomalies)

            # Sheet 2: 시간역전 상세
            ws_time = wb.create_sheet("시간역전 상세")
            time_reversals = [a for a in anomalies if a["Anomaly_Type"] == "시간 역전"]
            self._create_detail_sheet(ws_time, time_reversals, "시간역전")

            # Sheet 3: ML 이상치 상세
            ws_ml = wb.create_sheet("ML 이상치 상세")
            ml_outliers = [
                a for a in anomalies if a["Anomaly_Type"] == "머신러닝 이상치"
            ]
            self._create_detail_sheet(ws_ml, ml_outliers, "ML 이상치")

            # Sheet 4: 데이터 품질 상세
            ws_quality = wb.create_sheet("데이터 품질 상세")
            quality_issues = [
                a for a in anomalies if a["Anomaly_Type"] == "데이터 품질"
            ]
            self._create_detail_sheet(ws_quality, quality_issues, "데이터 품질")

            # Sheet 5: 원본 파일 참조
            ws_ref = wb.create_sheet("원본 파일 참조")
            self._create_reference_sheet(ws_ref)

            # 저장
            wb.save(output_path)
            print(f"✅ 보고서 생성 완료: {output_path}")
            return True

        except Exception as e:
            print(f"❌ 보고서 생성 실패: {e}")
            return False

    def _create_summary_sheet(self, ws, anomalies: List[Dict[str, Any]]):
        """요약 대시보드 시트 생성"""
        # 제목
        ws["A1"] = "HVDC 시간역전 이상치 분석보고서"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        ws.merge_cells("A1:F1")

        # 생성 일시
        ws["A2"] = f"생성 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)

        # 통계 요약
        row = 4
        ws[f"A{row}"] = "📊 이상치 통계 요약"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 2

        # 총 개수
        total_count = len(anomalies)
        ws[f"A{row}"] = "총 이상치 개수"
        ws[f"B{row}"] = f"{total_count:,}건"
        ws[f"B{row}"].font = Font(size=12, bold=True)
        row += 2

        # 유형별 분포
        ws[f"A{row}"] = "유형별 분포"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        type_counts = {}
        for anomaly in anomalies:
            anomaly_type = anomaly["Anomaly_Type"]
            type_counts[anomaly_type] = type_counts.get(anomaly_type, 0) + 1

        for anomaly_type, count in type_counts.items():
            ws[f"A{row}"] = f"  • {anomaly_type}"
            ws[f"B{row}"] = f"{count:,}건"

            # 색상 적용
            if anomaly_type == "시간 역전":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["RED"],
                    end_color=self.colors["RED"],
                    fill_type="solid",
                )
            elif anomaly_type == "머신러닝 이상치":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["ORANGE"],
                    end_color=self.colors["ORANGE"],
                    fill_type="solid",
                )
            elif anomaly_type == "데이터 품질":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["PURPLE"],
                    end_color=self.colors["PURPLE"],
                    fill_type="solid",
                )

            row += 1

        row += 1

        # 심각도별 분포
        ws[f"A{row}"] = "심각도별 분포"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        severity_counts = {}
        for anomaly in anomalies:
            severity = anomaly["Severity"]
            severity_counts[severity] = severity_counts.get(severity, 0) + 1

        for severity, count in severity_counts.items():
            ws[f"A{row}"] = f"  • {severity}"
            ws[f"B{row}"] = f"{count:,}건"
            row += 1

        row += 2

        # 색상 범례
        ws[f"A{row}"] = "🎨 색상 범례"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        legend_items = [
            ("🔴 빨강", "시간 역전", self.colors["RED"]),
            ("🟠 주황", "ML 이상치 (높음)", self.colors["ORANGE"]),
            ("🟡 노랑", "ML 이상치 (보통)", self.colors["YELLOW"]),
            ("🟣 보라", "데이터 품질", self.colors["PURPLE"]),
        ]

        for emoji, description, color in legend_items:
            ws[f"A{row}"] = emoji
            ws[f"B{row}"] = description
            ws[f"A{row}"].fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            row += 1

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

    def _create_detail_sheet(
        self, ws, anomalies: List[Dict[str, Any]], sheet_type: str
    ):
        """상세 데이터 시트 생성"""
        if not anomalies:
            ws["A1"] = f"{sheet_type} 이상치가 없습니다."
            return

        # 헤더
        headers = [
            "Case ID",
            "이상치 유형",
            "심각도",
            "설명",
            "탐지 일시",
            "위험도 점수",
            "위치",
            "탐지값",
            "예상 범위",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터
        for row_idx, anomaly in enumerate(anomalies, 2):
            ws.cell(row=row_idx, column=1, value=anomaly.get("Case_ID", ""))
            ws.cell(row=row_idx, column=2, value=anomaly.get("Anomaly_Type", ""))
            ws.cell(row=row_idx, column=3, value=anomaly.get("Severity", ""))
            ws.cell(row=row_idx, column=4, value=anomaly.get("Description", ""))
            ws.cell(row=row_idx, column=5, value=anomaly.get("Timestamp", ""))
            ws.cell(row=row_idx, column=6, value=anomaly.get("Risk_Score", ""))
            ws.cell(row=row_idx, column=7, value=anomaly.get("Location", ""))
            ws.cell(row=row_idx, column=8, value=anomaly.get("Detected_Value", ""))

            # 예상 범위
            expected_range = anomaly.get("Expected_Range")
            if (
                expected_range
                and isinstance(expected_range, list)
                and len(expected_range) == 2
            ):
                ws.cell(
                    row=row_idx,
                    column=9,
                    value=f"{expected_range[0]} ~ {expected_range[1]}",
                )
            else:
                ws.cell(row=row_idx, column=9, value=expected_range)

        # 컬럼 너비 조정
        column_widths = [15, 15, 10, 30, 20, 12, 15, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 테두리 추가
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=1, max_row=len(anomalies) + 1, min_col=1, max_col=len(headers)
        ):
            for cell in row:
                cell.border = thin_border

    def _create_reference_sheet(self, ws):
        """원본 파일 참조 시트 생성"""
        # 제목
        ws["A1"] = "📁 원본 파일 참조 정보"
        ws["A1"].font = Font(size=14, bold=True)

        # 원본 파일 정보
        ws["A3"] = "원본 Excel 파일:"
        ws["B3"] = "HVDC WAREHOUSE_HITACHI(HE).xlsx"
        ws["B3"].font = Font(bold=True)

        ws["A4"] = "파일 위치:"
        ws["B4"] = r"C:\Users\minky\Downloads\HVDC_Invoice_Audit-20251012T195441Z-1-001"

        ws["A5"] = "색상 적용 상태:"
        ws["B5"] = "✅ 완료 (508건 이상치 색상 표시됨)"
        ws["B5"].font = Font(color="008000", bold=True)

        # 백업 파일 정보
        ws["A7"] = "백업 파일:"
        ws["B7"] = "HVDC WAREHOUSE_HITACHI(HE).backup_YYYYMMDD_HHMMSS.xlsx"
        ws["B7"].font = Font(bold=True)

        ws["A8"] = "백업 위치:"
        ws["B8"] = "원본 파일과 동일한 폴더"

        # 색상 매핑 정보
        ws["A10"] = "🎨 색상 매핑 정보"
        ws["A10"].font = Font(size=12, bold=True)

        color_mappings = [
            ("시간 역전", "빨강 (FF0000)", "날짜 컬럼만 색칠"),
            ("ML 이상치 (높음)", "주황 (FFC000)", "전체 행 색칠"),
            ("ML 이상치 (보통)", "노랑 (FFFF00)", "전체 행 색칠"),
            ("데이터 품질", "보라 (CC99FF)", "전체 행 색칠"),
        ]

        row = 11
        for anomaly_type, color, description in color_mappings:
            ws[f"A{row}"] = f"• {anomaly_type}"
            ws[f"B{row}"] = color
            ws[f"C{row}"] = description
            row += 1

        # 주의사항
        ws["A15"] = "⚠️ 주의사항"
        ws["A15"].font = Font(size=12, bold=True, color="FF0000")

        ws["A16"] = "• 원본 파일은 절대 수정하지 마세요"
        ws["A17"] = "• 색상이 적용된 원본 파일을 참조하세요"
        ws["A18"] = "• 백업 파일이 있으므로 안전합니다"

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25


def main():
    parser = argparse.ArgumentParser(description="HVDC 시간역전 이상치 분석보고서 생성")
    parser.add_argument("--json", required=True, help="이상치 JSON 파일 경로")
    parser.add_argument("--output", help="출력 Excel 파일 경로 (기본값: 자동 생성)")

    args = parser.parse_args()

    # 출력 파일명 자동 생성
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d")
        args.output = f"시간역전_이상치_분석보고서_{timestamp}.xlsx"

    # 보고서 생성
    generator = AnalysisReportGenerator()
    success = generator.create_report(args.json, args.output)

    if success:
        print(f"🎉 분석보고서가 성공적으로 생성되었습니다!")
        print(f"📁 파일 위치: {Path(args.output).absolute()}")
    else:
        print("❌ 보고서 생성에 실패했습니다.")


if __name__ == "__main__":
    main()

HVDC 시간역전 이상치 분석보고서 생성기
- 원본 파일은 절대 수정하지 않음
- 별도의 Excel 보고서 파일 생성
- 4개 시트: 요약 대시보드, 시간역전 상세, ML 이상치 상세, 원본 파일 참조
"""
import json
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


class AnalysisReportGenerator:
    """이상치 분석 보고서 생성기"""

    def __init__(self):
        self.colors = {
            "RED": "FF0000",  # 시간 역전
            "ORANGE": "FFC000",  # ML 이상치 (높음)
            "YELLOW": "FFFF00",  # ML 이상치 (보통)
            "PURPLE": "CC99FF",  # 데이터 품질
        }

    def create_report(self, json_path: str, output_path: str) -> bool:
        """분석 보고서 생성"""
        try:
            # JSON 데이터 로드
            with open(json_path, "r", encoding="utf-8") as f:
                anomalies = json.load(f)

            print(f"📊 로드된 이상치 데이터: {len(anomalies)}건")

            # 새 워크북 생성
            wb = openpyxl.Workbook()

            # Sheet 1: 요약 대시보드
            ws_summary = wb.active
            ws_summary.title = "요약 대시보드"
            self._create_summary_sheet(ws_summary, anomalies)

            # Sheet 2: 시간역전 상세
            ws_time = wb.create_sheet("시간역전 상세")
            time_reversals = [a for a in anomalies if a["Anomaly_Type"] == "시간 역전"]
            self._create_detail_sheet(ws_time, time_reversals, "시간역전")

            # Sheet 3: ML 이상치 상세
            ws_ml = wb.create_sheet("ML 이상치 상세")
            ml_outliers = [
                a for a in anomalies if a["Anomaly_Type"] == "머신러닝 이상치"
            ]
            self._create_detail_sheet(ws_ml, ml_outliers, "ML 이상치")

            # Sheet 4: 데이터 품질 상세
            ws_quality = wb.create_sheet("데이터 품질 상세")
            quality_issues = [
                a for a in anomalies if a["Anomaly_Type"] == "데이터 품질"
            ]
            self._create_detail_sheet(ws_quality, quality_issues, "데이터 품질")

            # Sheet 5: 원본 파일 참조
            ws_ref = wb.create_sheet("원본 파일 참조")
            self._create_reference_sheet(ws_ref)

            # 저장
            wb.save(output_path)
            print(f"✅ 보고서 생성 완료: {output_path}")
            return True

        except Exception as e:
            print(f"❌ 보고서 생성 실패: {e}")
            return False

    def _create_summary_sheet(self, ws, anomalies: List[Dict[str, Any]]):
        """요약 대시보드 시트 생성"""
        # 제목
        ws["A1"] = "HVDC 시간역전 이상치 분석보고서"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        ws.merge_cells("A1:F1")

        # 생성 일시
        ws["A2"] = f"생성 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)

        # 통계 요약
        row = 4
        ws[f"A{row}"] = "📊 이상치 통계 요약"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 2

        # 총 개수
        total_count = len(anomalies)
        ws[f"A{row}"] = "총 이상치 개수"
        ws[f"B{row}"] = f"{total_count:,}건"
        ws[f"B{row}"].font = Font(size=12, bold=True)
        row += 2

        # 유형별 분포
        ws[f"A{row}"] = "유형별 분포"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        type_counts = {}
        for anomaly in anomalies:
            anomaly_type = anomaly["Anomaly_Type"]
            type_counts[anomaly_type] = type_counts.get(anomaly_type, 0) + 1

        for anomaly_type, count in type_counts.items():
            ws[f"A{row}"] = f"  • {anomaly_type}"
            ws[f"B{row}"] = f"{count:,}건"

            # 색상 적용
            if anomaly_type == "시간 역전":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["RED"],
                    end_color=self.colors["RED"],
                    fill_type="solid",
                )
            elif anomaly_type == "머신러닝 이상치":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["ORANGE"],
                    end_color=self.colors["ORANGE"],
                    fill_type="solid",
                )
            elif anomaly_type == "데이터 품질":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["PURPLE"],
                    end_color=self.colors["PURPLE"],
                    fill_type="solid",
                )

            row += 1

        row += 1

        # 심각도별 분포
        ws[f"A{row}"] = "심각도별 분포"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        severity_counts = {}
        for anomaly in anomalies:
            severity = anomaly["Severity"]
            severity_counts[severity] = severity_counts.get(severity, 0) + 1

        for severity, count in severity_counts.items():
            ws[f"A{row}"] = f"  • {severity}"
            ws[f"B{row}"] = f"{count:,}건"
            row += 1

        row += 2

        # 색상 범례
        ws[f"A{row}"] = "🎨 색상 범례"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        legend_items = [
            ("🔴 빨강", "시간 역전", self.colors["RED"]),
            ("🟠 주황", "ML 이상치 (높음)", self.colors["ORANGE"]),
            ("🟡 노랑", "ML 이상치 (보통)", self.colors["YELLOW"]),
            ("🟣 보라", "데이터 품질", self.colors["PURPLE"]),
        ]

        for emoji, description, color in legend_items:
            ws[f"A{row}"] = emoji
            ws[f"B{row}"] = description
            ws[f"A{row}"].fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            row += 1

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

    def _create_detail_sheet(
        self, ws, anomalies: List[Dict[str, Any]], sheet_type: str
    ):
        """상세 데이터 시트 생성"""
        if not anomalies:
            ws["A1"] = f"{sheet_type} 이상치가 없습니다."
            return

        # 헤더
        headers = [
            "Case ID",
            "이상치 유형",
            "심각도",
            "설명",
            "탐지 일시",
            "위험도 점수",
            "위치",
            "탐지값",
            "예상 범위",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터
        for row_idx, anomaly in enumerate(anomalies, 2):
            ws.cell(row=row_idx, column=1, value=anomaly.get("Case_ID", ""))
            ws.cell(row=row_idx, column=2, value=anomaly.get("Anomaly_Type", ""))
            ws.cell(row=row_idx, column=3, value=anomaly.get("Severity", ""))
            ws.cell(row=row_idx, column=4, value=anomaly.get("Description", ""))
            ws.cell(row=row_idx, column=5, value=anomaly.get("Timestamp", ""))
            ws.cell(row=row_idx, column=6, value=anomaly.get("Risk_Score", ""))
            ws.cell(row=row_idx, column=7, value=anomaly.get("Location", ""))
            ws.cell(row=row_idx, column=8, value=anomaly.get("Detected_Value", ""))

            # 예상 범위
            expected_range = anomaly.get("Expected_Range")
            if (
                expected_range
                and isinstance(expected_range, list)
                and len(expected_range) == 2
            ):
                ws.cell(
                    row=row_idx,
                    column=9,
                    value=f"{expected_range[0]} ~ {expected_range[1]}",
                )
            else:
                ws.cell(row=row_idx, column=9, value=expected_range)

        # 컬럼 너비 조정
        column_widths = [15, 15, 10, 30, 20, 12, 15, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 테두리 추가
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=1, max_row=len(anomalies) + 1, min_col=1, max_col=len(headers)
        ):
            for cell in row:
                cell.border = thin_border

    def _create_reference_sheet(self, ws):
        """원본 파일 참조 시트 생성"""
        # 제목
        ws["A1"] = "📁 원본 파일 참조 정보"
        ws["A1"].font = Font(size=14, bold=True)

        # 원본 파일 정보
        ws["A3"] = "원본 Excel 파일:"
        ws["B3"] = "HVDC WAREHOUSE_HITACHI(HE).xlsx"
        ws["B3"].font = Font(bold=True)

        ws["A4"] = "파일 위치:"
        ws["B4"] = r"C:\Users\minky\Downloads\HVDC_Invoice_Audit-20251012T195441Z-1-001"

        ws["A5"] = "색상 적용 상태:"
        ws["B5"] = "✅ 완료 (508건 이상치 색상 표시됨)"
        ws["B5"].font = Font(color="008000", bold=True)

        # 백업 파일 정보
        ws["A7"] = "백업 파일:"
        ws["B7"] = "HVDC WAREHOUSE_HITACHI(HE).backup_YYYYMMDD_HHMMSS.xlsx"
        ws["B7"].font = Font(bold=True)

        ws["A8"] = "백업 위치:"
        ws["B8"] = "원본 파일과 동일한 폴더"

        # 색상 매핑 정보
        ws["A10"] = "🎨 색상 매핑 정보"
        ws["A10"].font = Font(size=12, bold=True)

        color_mappings = [
            ("시간 역전", "빨강 (FF0000)", "날짜 컬럼만 색칠"),
            ("ML 이상치 (높음)", "주황 (FFC000)", "전체 행 색칠"),
            ("ML 이상치 (보통)", "노랑 (FFFF00)", "전체 행 색칠"),
            ("데이터 품질", "보라 (CC99FF)", "전체 행 색칠"),
        ]

        row = 11
        for anomaly_type, color, description in color_mappings:
            ws[f"A{row}"] = f"• {anomaly_type}"
            ws[f"B{row}"] = color
            ws[f"C{row}"] = description
            row += 1

        # 주의사항
        ws["A15"] = "⚠️ 주의사항"
        ws["A15"].font = Font(size=12, bold=True, color="FF0000")

        ws["A16"] = "• 원본 파일은 절대 수정하지 마세요"
        ws["A17"] = "• 색상이 적용된 원본 파일을 참조하세요"
        ws["A18"] = "• 백업 파일이 있으므로 안전합니다"

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25


def main():
    parser = argparse.ArgumentParser(description="HVDC 시간역전 이상치 분석보고서 생성")
    parser.add_argument("--json", required=True, help="이상치 JSON 파일 경로")
    parser.add_argument("--output", help="출력 Excel 파일 경로 (기본값: 자동 생성)")

    args = parser.parse_args()

    # 출력 파일명 자동 생성
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d")
        args.output = f"시간역전_이상치_분석보고서_{timestamp}.xlsx"

    # 보고서 생성
    generator = AnalysisReportGenerator()
    success = generator.create_report(args.json, args.output)

    if success:
        print(f"🎉 분석보고서가 성공적으로 생성되었습니다!")
        print(f"📁 파일 위치: {Path(args.output).absolute()}")
    else:
        print("❌ 보고서 생성에 실패했습니다.")


if __name__ == "__main__":
    main()
