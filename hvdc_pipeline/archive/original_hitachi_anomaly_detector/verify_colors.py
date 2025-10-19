# -*- coding: utf-8 -*-
"""
Excel 파일의 색상 적용 확인 스크립트
"""

import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict


def verify_colors_in_excel(
    file_path="../HVDC WAREHOUSE_HITACHI(HE).xlsx", sheet_name="Case List"
):
    """Excel 파일에서 색상 적용 확인"""

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        print(f"📊 Excel 파일 분석: {file_path}")
        print(f"📊 시트: {sheet_name}")
        print(f"📊 총 행 수: {ws.max_row}")
        print(f"📊 총 열 수: {ws.max_column}")

        # 색상별 카운트
        color_counts = defaultdict(int)
        colored_cells = []

        # 헤더 행 건너뛰고 데이터 행만 확인
        for row_num in range(2, ws.max_row + 1):
            row_colors = []

            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)

                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    color_rgb = cell.fill.start_color.rgb
                    color_counts[color_rgb] += 1
                    row_colors.append((col_num, color_rgb))

            if row_colors:
                # Case NO 가져오기 (11번째 컬럼)
                case_no = ws.cell(row=row_num, column=11).value
                colored_cells.append(
                    {"row": row_num, "case_no": case_no, "colors": row_colors}
                )

        print(f"\n🎨 색상 적용 결과:")
        print(f"  - 총 색칠된 행: {len(colored_cells)}개")

        for color_rgb, count in color_counts.items():
            color_name = get_color_name(color_rgb)
            print(f"  - {color_name} ({color_rgb}): {count}개 셀")

        # 처음 10개 색칠된 행 상세 정보
        print(f"\n🔍 처음 10개 색칠된 행:")
        for i, cell_info in enumerate(colored_cells[:10]):
            print(f"  {i+1}. 행 {cell_info['row']}, Case NO: {cell_info['case_no']}")
            for col_num, color_rgb in cell_info["colors"]:
                color_name = get_color_name(color_rgb)
                col_letter = openpyxl.utils.get_column_letter(col_num)
                print(
                    f"     - {col_letter}{cell_info['row']}: {color_name} ({color_rgb})"
                )

        # 색상별 행 수 계산
        red_rows = set()
        orange_rows = set()
        yellow_rows = set()
        purple_rows = set()

        for cell_info in colored_cells:
            for col_num, color_rgb in cell_info["colors"]:
                if color_rgb == "00FF0000":  # 빨강
                    red_rows.add(cell_info["row"])
                elif color_rgb == "00FFC000":  # 주황
                    orange_rows.add(cell_info["row"])
                elif color_rgb == "00FFFF00":  # 노랑
                    yellow_rows.add(cell_info["row"])
                elif color_rgb == "00CC99FF":  # 보라
                    purple_rows.add(cell_info["row"])

        print(f"\n📊 색상별 행 수:")
        print(f"  - 빨강 행: {len(red_rows)}개 (시간 역전)")
        print(f"  - 주황 행: {len(orange_rows)}개 (ML 이상치 - 높음)")
        print(f"  - 노랑 행: {len(yellow_rows)}개 (ML 이상치 - 보통)")
        print(f"  - 보라 행: {len(purple_rows)}개 (데이터 품질)")

        return {
            "total_colored_rows": len(colored_cells),
            "color_counts": dict(color_counts),
            "red_rows": len(red_rows),
            "orange_rows": len(orange_rows),
            "yellow_rows": len(yellow_rows),
            "purple_rows": len(purple_rows),
        }

    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        return None


def get_color_name(color_rgb):
    """RGB 색상 코드를 색상명으로 변환"""
    color_map = {
        "00FF0000": "빨강 (RED)",
        "00FFC000": "주황 (ORANGE)",
        "00FFFF00": "노랑 (YELLOW)",
        "00CC99FF": "보라 (PURPLE)",
        "0000FF00": "초록 (GREEN)",
        "000000FF": "파랑 (BLUE)",
    }
    return color_map.get(color_rgb, f"기타 ({color_rgb})")


if __name__ == "__main__":
    result = verify_colors_in_excel()
    if result:
        print(f"\n✅ 색상 검증 완료!")
        print(f"📊 요약: 총 {result['total_colored_rows']}개 행에 색상 적용됨")
"""
Excel 파일의 색상 적용 확인 스크립트
"""

import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict


def verify_colors_in_excel(
    file_path="../HVDC WAREHOUSE_HITACHI(HE).xlsx", sheet_name="Case List"
):
    """Excel 파일에서 색상 적용 확인"""

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        print(f"📊 Excel 파일 분석: {file_path}")
        print(f"📊 시트: {sheet_name}")
        print(f"📊 총 행 수: {ws.max_row}")
        print(f"📊 총 열 수: {ws.max_column}")

        # 색상별 카운트
        color_counts = defaultdict(int)
        colored_cells = []

        # 헤더 행 건너뛰고 데이터 행만 확인
        for row_num in range(2, ws.max_row + 1):
            row_colors = []

            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)

                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    color_rgb = cell.fill.start_color.rgb
                    color_counts[color_rgb] += 1
                    row_colors.append((col_num, color_rgb))

            if row_colors:
                # Case NO 가져오기 (11번째 컬럼)
                case_no = ws.cell(row=row_num, column=11).value
                colored_cells.append(
                    {"row": row_num, "case_no": case_no, "colors": row_colors}
                )

        print(f"\n🎨 색상 적용 결과:")
        print(f"  - 총 색칠된 행: {len(colored_cells)}개")

        for color_rgb, count in color_counts.items():
            color_name = get_color_name(color_rgb)
            print(f"  - {color_name} ({color_rgb}): {count}개 셀")

        # 처음 10개 색칠된 행 상세 정보
        print(f"\n🔍 처음 10개 색칠된 행:")
        for i, cell_info in enumerate(colored_cells[:10]):
            print(f"  {i+1}. 행 {cell_info['row']}, Case NO: {cell_info['case_no']}")
            for col_num, color_rgb in cell_info["colors"]:
                color_name = get_color_name(color_rgb)
                col_letter = openpyxl.utils.get_column_letter(col_num)
                print(
                    f"     - {col_letter}{cell_info['row']}: {color_name} ({color_rgb})"
                )

        # 색상별 행 수 계산
        red_rows = set()
        orange_rows = set()
        yellow_rows = set()
        purple_rows = set()

        for cell_info in colored_cells:
            for col_num, color_rgb in cell_info["colors"]:
                if color_rgb == "00FF0000":  # 빨강
                    red_rows.add(cell_info["row"])
                elif color_rgb == "00FFC000":  # 주황
                    orange_rows.add(cell_info["row"])
                elif color_rgb == "00FFFF00":  # 노랑
                    yellow_rows.add(cell_info["row"])
                elif color_rgb == "00CC99FF":  # 보라
                    purple_rows.add(cell_info["row"])

        print(f"\n📊 색상별 행 수:")
        print(f"  - 빨강 행: {len(red_rows)}개 (시간 역전)")
        print(f"  - 주황 행: {len(orange_rows)}개 (ML 이상치 - 높음)")
        print(f"  - 노랑 행: {len(yellow_rows)}개 (ML 이상치 - 보통)")
        print(f"  - 보라 행: {len(purple_rows)}개 (데이터 품질)")

        return {
            "total_colored_rows": len(colored_cells),
            "color_counts": dict(color_counts),
            "red_rows": len(red_rows),
            "orange_rows": len(orange_rows),
            "yellow_rows": len(yellow_rows),
            "purple_rows": len(purple_rows),
        }

    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        return None


def get_color_name(color_rgb):
    """RGB 색상 코드를 색상명으로 변환"""
    color_map = {
        "00FF0000": "빨강 (RED)",
        "00FFC000": "주황 (ORANGE)",
        "00FFFF00": "노랑 (YELLOW)",
        "00CC99FF": "보라 (PURPLE)",
        "0000FF00": "초록 (GREEN)",
        "000000FF": "파랑 (BLUE)",
    }
    return color_map.get(color_rgb, f"기타 ({color_rgb})")


if __name__ == "__main__":
    result = verify_colors_in_excel()
    if result:
        print(f"\n✅ 색상 검증 완료!")
        print(f"📊 요약: 총 {result['total_colored_rows']}개 행에 색상 적용됨")
