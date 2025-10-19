#!/usr/bin/env python3
"""
HVDC 종합 보고서에 이상치 색상 적용 스크립트
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from typing import Dict, List, Any


def normalize_case_id(case_id: str) -> str:
    """Case ID 정규화 (공백, 특수문자 제거)"""
    if not case_id:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(case_id).strip().upper())


def load_anomaly_data(json_file: str) -> Dict[str, Any]:
    """이상치 JSON 데이터 로드"""
    with open(json_file, "r", encoding="utf-8") as f:
        return json.load(f)


def apply_anomaly_colors(report_file: str, anomaly_json: str, output_file: str):
    """종합 보고서에 이상치 색상 적용"""

    print("Loading anomaly data...")
    anomaly_data = load_anomaly_data(anomaly_json)

    print("Loading report file...")
    wb = openpyxl.load_workbook(report_file)

    # 통합_원본데이터_Fixed 시트 찾기
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "통합_원본데이터_Fixed" in sheet_name:
            target_sheet = wb[sheet_name]
            break

    if not target_sheet:
        print("ERROR: 통합_원본데이터_Fixed 시트를 찾을 수 없습니다.")
        return False

    print(f"Found target sheet: {target_sheet.title}")

    # 색상 정의 (ARGB 8자리)
    colors = {
        "time_reversal": PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        ),  # 빨간색
        "ml_high": PatternFill(
            start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"
        ),  # 주황색
        "ml_medium": PatternFill(
            start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
        ),  # 노란색
        "data_quality": PatternFill(
            start_color="FFCC99FF", end_color="FFCC99FF", fill_type="solid"
        ),  # 보라색
    }

    # Case NO 컬럼 찾기 (Case No. 정확히 매칭)
    case_col = None
    for col in range(1, target_sheet.max_column + 1):
        cell_value = target_sheet.cell(row=1, column=col).value
        if cell_value:
            cell_str = str(cell_value)
            if "Case No." in cell_str or "CASE_NO" in cell_str.upper():
                case_col = col
                break

    if not case_col:
        print("ERROR: Case NO 컬럼을 찾을 수 없습니다.")
        return False

    print(f"Found Case NO column: {case_col}")

    # 이상치 적용
    applied_count = 0
    anomaly_counts = {
        "time_reversal": 0,
        "ml_high": 0,
        "ml_medium": 0,
        "data_quality": 0,
    }

    # 이상치 데이터 처리 (리스트 형태)
    for anomaly in anomaly_data:
        case_id = anomaly.get("Case_ID", "")
        anomaly_type = anomaly.get("Anomaly_Type", "")
        severity = anomaly.get("Severity", "")

        case_id_norm = normalize_case_id(case_id)

        # 해당 Case ID 찾기 (문자열과 정수 모두 처리)
        for row in range(2, target_sheet.max_row + 1):
            cell_value = target_sheet.cell(row=row, column=case_col).value
            if cell_value:
                # 정수인 경우 문자열로 변환 후 비교
                cell_str = str(cell_value)
                if cell_str == case_id:
                    print(f"Found match: Row {row}, Case ID: {case_id}")

                    # 이상치 유형에 따라 색상 적용 (한글 매칭)
                    if "시간 역전" in anomaly_type or "time" in anomaly_type.lower():
                        # 시간 역전: 날짜 컬럼만
                        for col in range(1, target_sheet.max_column + 1):
                            header = target_sheet.cell(row=1, column=col).value
                            if header and any(
                                keyword in str(header).lower()
                                for keyword in ["date", "날짜", "time", "시간"]
                            ):
                                target_sheet.cell(row=row, column=col).fill = colors[
                                    "time_reversal"
                                ]
                        anomaly_counts["time_reversal"] += 1
                        applied_count += 1

                    elif "머신러닝" in anomaly_type or "ml" in anomaly_type.lower():
                        # ML 이상치: 심각도에 따라 색상 선택
                        if severity in ["높음", "치명적", "high", "critical"]:
                            color = colors["ml_high"]
                            anomaly_counts["ml_high"] += 1
                        else:
                            color = colors["ml_medium"]
                            anomaly_counts["ml_medium"] += 1

                        # 전체 행에 색상 적용
                        for col in range(1, target_sheet.max_column + 1):
                            target_sheet.cell(row=row, column=col).fill = color
                        applied_count += 1

                    elif "데이터 품질" in anomaly_type or "품질" in anomaly_type:
                        # 데이터 품질: 전체 행
                        for col in range(1, target_sheet.max_column + 1):
                            target_sheet.cell(row=row, column=col).fill = colors[
                                "data_quality"
                            ]
                        anomaly_counts["data_quality"] += 1
                        applied_count += 1

                    break

    # 색상 범례 시트 추가
    legend_sheet = wb.create_sheet("색상 범례")

    legend_data = [
        ["색상", "의미", "적용 범위", "개수"],
        [
            "🔴 빨간색",
            "시간 역전 이상치",
            "날짜 컬럼만",
            str(anomaly_counts["time_reversal"]),
        ],
        [
            "🟠 주황색",
            "ML 이상치 (높음/치명적)",
            "전체 행",
            str(anomaly_counts["ml_high"]),
        ],
        [
            "🟡 노란색",
            "ML 이상치 (보통/낮음)",
            "전체 행",
            str(anomaly_counts["ml_medium"]),
        ],
        [
            "🟣 보라색",
            "데이터 품질 이상",
            "전체 행",
            str(anomaly_counts["data_quality"]),
        ],
        ["", "", "총 적용", str(applied_count)],
    ]

    for row_data in legend_data:
        legend_sheet.append(row_data)

    # 범례 시트 서식 적용
    for row in range(1, len(legend_data) + 1):
        for col in range(1, len(legend_data[0]) + 1):
            cell = legend_sheet.cell(row=row, column=col)
            if row == 1:  # 헤더
                cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin"),
            )

    print(f"Saving colored report to: {output_file}")
    wb.save(output_file)

    print(f"SUCCESS: Applied colors to {applied_count} cases")
    print(f"Color legend added to '색상 범례' sheet")
    print(f"Anomaly breakdown: {anomaly_counts}")

    return True


def main():
    import argparse

    parser = argparse.ArgumentParser(description="HVDC 종합 보고서에 이상치 색상 적용")
    parser.add_argument("--report", required=True, help="종합 보고서 파일 경로")
    parser.add_argument("--anomaly", required=True, help="이상치 JSON 파일 경로")
    parser.add_argument("--output", required=True, help="출력 파일 경로")

    args = parser.parse_args()

    success = apply_anomaly_colors(args.report, args.anomaly, args.output)

    if success:
        print("Anomaly color marking completed successfully!")
    else:
        print("Failed to apply anomaly colors.")
        exit(1)


if __name__ == "__main__":
    main()
