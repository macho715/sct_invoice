# -*- coding: utf-8 -*-
"""
누락된 이상치 디버깅 스크립트
- JSON의 933건과 Excel의 857건 차이 분석
- 어떤 이상치 유형이 누락되었는지 확인
"""
import json
import openpyxl
from collections import defaultdict
from anomaly_detector import AnomalyRecord, AnomalyType, AnomalySeverity
from datetime import datetime


def debug_missing_anomalies(json_file: str, excel_file: str, sheet_name: str = "통합_원본데이터_Fixed"):
    """누락된 이상치 분석"""
    print(f"🔍 누락된 이상치 디버깅 시작...")
    
    # 1. JSON에서 이상치 로드
    with open(json_file, 'r', encoding='utf-8') as f:
        anomaly_data = json.load(f)
    
    print(f"📊 JSON 이상치: {len(anomaly_data)}건")
    
    # 2. AnomalyRecord 객체 생성
    anomalies = []
    for item in anomaly_data:
        anomaly = AnomalyRecord(
            case_id=item["Case_ID"],
            anomaly_type=AnomalyType(item["Anomaly_Type"]),
            severity=AnomalySeverity(item["Severity"]),
            description=item["Description"],
            detected_value=item["Detected_Value"],
            expected_range=tuple(item["Expected_Range"]) if item["Expected_Range"] else None,
            location=item["Location"],
            timestamp=datetime.fromisoformat(item["Timestamp"]),
            risk_score=item["Risk_Score"]
        )
        anomalies.append(anomaly)
    
    # 3. Excel에서 Case ID 수집
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Case NO 컬럼 찾기
    case_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and "case" in str(cell.value).lower():
            case_col_idx = idx
            break
    
    if case_col_idx is None:
        print("❌ Case NO 컬럼을 찾을 수 없습니다")
        return
    
    # Excel의 Case ID 수집
    excel_case_ids = set()
    for row_num in range(2, ws.max_row + 1):
        case_value = ws.cell(row=row_num, column=case_col_idx).value
        if case_value:
            excel_case_ids.add(str(case_value).strip().upper())
    
    print(f"📊 Excel Case ID: {len(excel_case_ids)}개")
    
    # 4. 매칭 분석
    json_case_ids = set(str(a.case_id).strip().upper() for a in anomalies)
    
    matched = json_case_ids.intersection(excel_case_ids)
    missing_in_excel = json_case_ids - excel_case_ids
    extra_in_excel = excel_case_ids - json_case_ids
    
    print(f"\n📊 매칭 분석:")
    print(f"  - 매칭된 Case ID: {len(matched)}개")
    print(f"  - JSON만 있는 Case ID: {len(missing_in_excel)}개")
    print(f"  - Excel만 있는 Case ID: {len(extra_in_excel)}개")
    
    # 5. 누락된 이상치 상세 분석
    if missing_in_excel:
        print(f"\n❌ Excel에 없는 Case ID (처음 10개):")
        for case_id in list(missing_in_excel)[:10]:
            print(f"  - {case_id}")
    
    # 6. 이상치 유형별 분석
    print(f"\n📊 이상치 유형별 분석:")
    type_counts = defaultdict(int)
    matched_type_counts = defaultdict(int)
    
    for anomaly in anomalies:
        anomaly_type = anomaly.anomaly_type.value
        type_counts[anomaly_type] += 1
        
        if str(anomaly.case_id).strip().upper() in matched:
            matched_type_counts[anomaly_type] += 1
    
    for anomaly_type in type_counts:
        total = type_counts[anomaly_type]
        matched = matched_type_counts[anomaly_type]
        missing = total - matched
        print(f"  - {anomaly_type}: {total}건 (매칭: {matched}건, 누락: {missing}건)")
    
    # 7. 색상별 분석 (실제 적용된 색상)
    print(f"\n🎨 Excel 색상 분석:")
    color_counts = defaultdict(int)
    colored_rows = 0
    
    for row_num in range(2, ws.max_row + 1):
        has_color = False
        row_colors = set()
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color and color != "00000000":
                    color_str = str(color).upper()
                    row_colors.add(color_str)
                    has_color = True
        
        if has_color:
            colored_rows += 1
            for color in row_colors:
                color_counts[color] += 1
    
    print(f"  - 총 색칠된 행: {colored_rows}개")
    for color, count in sorted(color_counts.items()):
        print(f"  - {color}: {count}개")
    
    return {
        "json_total": len(anomalies),
        "excel_total": len(excel_case_ids),
        "matched": len(matched),
        "missing": len(missing_in_excel),
        "type_counts": dict(type_counts),
        "matched_type_counts": dict(matched_type_counts),
        "colored_rows": colored_rows,
        "color_counts": dict(color_counts)
    }


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="누락된 이상치 디버깅")
    parser.add_argument("--json", required=True, help="이상치 JSON 파일")
    parser.add_argument("--excel", required=True, help="Excel 파일")
    parser.add_argument("--sheet", default="통합_원본데이터_Fixed", help="시트명")
    
    args = parser.parse_args()
    
    result = debug_missing_anomalies(args.json, args.excel, args.sheet)
    
    if result:
        print(f"\n✅ 디버깅 완료!")
    else:
        print(f"\n❌ 디버깅 실패!")


if __name__ == "__main__":
    main()
"""
누락된 이상치 디버깅 스크립트
- JSON의 933건과 Excel의 857건 차이 분석
- 어떤 이상치 유형이 누락되었는지 확인
"""
import json
import openpyxl
from collections import defaultdict
from anomaly_detector import AnomalyRecord, AnomalyType, AnomalySeverity
from datetime import datetime


def debug_missing_anomalies(json_file: str, excel_file: str, sheet_name: str = "통합_원본데이터_Fixed"):
    """누락된 이상치 분석"""
    print(f"🔍 누락된 이상치 디버깅 시작...")
    
    # 1. JSON에서 이상치 로드
    with open(json_file, 'r', encoding='utf-8') as f:
        anomaly_data = json.load(f)
    
    print(f"📊 JSON 이상치: {len(anomaly_data)}건")
    
    # 2. AnomalyRecord 객체 생성
    anomalies = []
    for item in anomaly_data:
        anomaly = AnomalyRecord(
            case_id=item["Case_ID"],
            anomaly_type=AnomalyType(item["Anomaly_Type"]),
            severity=AnomalySeverity(item["Severity"]),
            description=item["Description"],
            detected_value=item["Detected_Value"],
            expected_range=tuple(item["Expected_Range"]) if item["Expected_Range"] else None,
            location=item["Location"],
            timestamp=datetime.fromisoformat(item["Timestamp"]),
            risk_score=item["Risk_Score"]
        )
        anomalies.append(anomaly)
    
    # 3. Excel에서 Case ID 수집
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Case NO 컬럼 찾기
    case_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and "case" in str(cell.value).lower():
            case_col_idx = idx
            break
    
    if case_col_idx is None:
        print("❌ Case NO 컬럼을 찾을 수 없습니다")
        return
    
    # Excel의 Case ID 수집
    excel_case_ids = set()
    for row_num in range(2, ws.max_row + 1):
        case_value = ws.cell(row=row_num, column=case_col_idx).value
        if case_value:
            excel_case_ids.add(str(case_value).strip().upper())
    
    print(f"📊 Excel Case ID: {len(excel_case_ids)}개")
    
    # 4. 매칭 분석
    json_case_ids = set(str(a.case_id).strip().upper() for a in anomalies)
    
    matched = json_case_ids.intersection(excel_case_ids)
    missing_in_excel = json_case_ids - excel_case_ids
    extra_in_excel = excel_case_ids - json_case_ids
    
    print(f"\n📊 매칭 분석:")
    print(f"  - 매칭된 Case ID: {len(matched)}개")
    print(f"  - JSON만 있는 Case ID: {len(missing_in_excel)}개")
    print(f"  - Excel만 있는 Case ID: {len(extra_in_excel)}개")
    
    # 5. 누락된 이상치 상세 분석
    if missing_in_excel:
        print(f"\n❌ Excel에 없는 Case ID (처음 10개):")
        for case_id in list(missing_in_excel)[:10]:
            print(f"  - {case_id}")
    
    # 6. 이상치 유형별 분석
    print(f"\n📊 이상치 유형별 분석:")
    type_counts = defaultdict(int)
    matched_type_counts = defaultdict(int)
    
    for anomaly in anomalies:
        anomaly_type = anomaly.anomaly_type.value
        type_counts[anomaly_type] += 1
        
        if str(anomaly.case_id).strip().upper() in matched:
            matched_type_counts[anomaly_type] += 1
    
    for anomaly_type in type_counts:
        total = type_counts[anomaly_type]
        matched = matched_type_counts[anomaly_type]
        missing = total - matched
        print(f"  - {anomaly_type}: {total}건 (매칭: {matched}건, 누락: {missing}건)")
    
    # 7. 색상별 분석 (실제 적용된 색상)
    print(f"\n🎨 Excel 색상 분석:")
    color_counts = defaultdict(int)
    colored_rows = 0
    
    for row_num in range(2, ws.max_row + 1):
        has_color = False
        row_colors = set()
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color and color != "00000000":
                    color_str = str(color).upper()
                    row_colors.add(color_str)
                    has_color = True
        
        if has_color:
            colored_rows += 1
            for color in row_colors:
                color_counts[color] += 1
    
    print(f"  - 총 색칠된 행: {colored_rows}개")
    for color, count in sorted(color_counts.items()):
        print(f"  - {color}: {count}개")
    
    return {
        "json_total": len(anomalies),
        "excel_total": len(excel_case_ids),
        "matched": len(matched),
        "missing": len(missing_in_excel),
        "type_counts": dict(type_counts),
        "matched_type_counts": dict(matched_type_counts),
        "colored_rows": colored_rows,
        "color_counts": dict(color_counts)
    }


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="누락된 이상치 디버깅")
    parser.add_argument("--json", required=True, help="이상치 JSON 파일")
    parser.add_argument("--excel", required=True, help="Excel 파일")
    parser.add_argument("--sheet", default="통합_원본데이터_Fixed", help="시트명")
    
    args = parser.parse_args()
    
    result = debug_missing_anomalies(args.json, args.excel, args.sheet)
    
    if result:
        print(f"\n✅ 디버깅 완료!")
    else:
        print(f"\n❌ 디버깅 실패!")


if __name__ == "__main__":
    main()
