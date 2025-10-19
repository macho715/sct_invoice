# -*- coding: utf-8 -*-
from __future__ import annotations
import re, shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Union

import openpyxl
from openpyxl.styles import PatternFill

# ---- ARGB 정의(불투명: FF alpha). 검증 스크립트 호환 위해 00/FF 모두 허용 ----
DEFAULT_STAGE3_SHEET = "통합_원본데이터_Fixed"

ARGB = {
    "RED":    ("FFFF0000", {"FFFF0000", "00FF0000"}),
    "ORANGE": ("FFFFC000", {"FFFFC000", "00FFC000"}),
    "YELLOW": ("FFFFFF00", {"FFFFFF00", "00FFFF00"}),
    "PURPLE": ("FFCC99FF", {"FFCC99FF", "00CC99FF"}),
}

DATE_KEYWORDS = {
    # 한글/영문 혼합 키워드(15개 이상)
    "date","day","time","dt","입고","출고","도착","출발","반출","반입","통관",
    "선적","출항","입항","검수","검품","warehouse","site"
}

def _norm_case(s: object) -> str:
    """Case ID 정규화: 공백/특수문자 제거 + 대문자."""
    return re.sub(r"[^A-Z0-9]", "", str(s).strip().upper()) if s is not None else ""

def _is_date_col(header: str, sample_vals: List[object]) -> bool:
    h = str(header).strip().lower()
    if any(k in h for k in DATE_KEYWORDS):
        return True
    # 값 기반 휴리스틱: 30% 이상이 날짜로 파싱 가능하면 날짜열로 간주
    import pandas as pd
    s = pd.to_datetime(pd.Series(sample_vals), errors="coerce")
    non_na = s.notna().mean() if len(s) else 0.0
    return non_na >= 0.3

def _fill(cell, argb: str):
    cell.fill = PatternFill(fill_type="solid", start_color=argb, end_color=argb)

class AnomalyVisualizer:
    """
    anomalies: List[dict|dataclass] — dict 키는 Anomaly_Type/Severity/Case_ID
    """
    def __init__(self, anomalies: Iterable[object]):
        self.records: List[Dict] = []
        for a in anomalies:
            if hasattr(a, "to_dict"):
                self.records.append(a.to_dict())
            elif isinstance(a, dict):
                self.records.append(a)
        # Case → anomaly 목록(중복 허용)
        self.by_case: Dict[str, List[Dict]] = {}
        for r in self.records:
            cid = _norm_case(r.get("Case_ID", ""))
            if cid:
                self.by_case.setdefault(cid, []).append(r)

    def apply_anomaly_colors(
        self,
        excel_file: Union[str, Path],
        sheet_name: str = DEFAULT_STAGE3_SHEET,
        case_col: str = "Case No.",
        create_backup: bool = True,
    ) -> Dict:
        excel_file = Path(excel_file)
        if create_backup:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak = excel_file.with_name(f"{excel_file.stem}.backup_{ts}{excel_file.suffix}")
            shutil.copyfile(excel_file, bak)

        wb = openpyxl.load_workbook(excel_file, keep_vba=excel_file.suffix.lower()==".xlsm")
        if sheet_name not in wb.sheetnames:
            return {"success": False, "message": f"시트 없음: {sheet_name}"}
        ws = wb[sheet_name]

        # 헤더 스캔 → case 컬럼 index
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        case_col_idx = None
        for c, name in enumerate(header, 1):
            if name and "case" in str(name).lower():
                case_col_idx = c; break
        if not case_col_idx:
            return {"success": False, "message": "Case NO 열을 찾지 못함"}

        # 날짜열 식별(헤더 + 샘플 기반)
        date_cols: List[int] = []
        for c, name in enumerate(header, 1):
            sample = [ws.cell(row=r, column=c).value for r in range(2, min(ws.max_row, 50)+1)]
            if _is_date_col(name, sample):
                date_cols.append(c)

        # 색칠
        cnt = {"time_reversal": 0, "ml_outlier": 0, "data_quality": 0}
        for r in range(2, ws.max_row+1):
            raw_id = ws.cell(row=r, column=case_col_idx).value
            cid = _norm_case(raw_id)
            if not cid or cid not in self.by_case:
                continue

            # 동일 Case의 다중 이상치 처리: 시간역전(날짜열) + ML/품질(행 전체) 병행
            row_anoms = self.by_case[cid]
            paint_row = None  # ORANGE/YELLOW/PURPLE 우선순위: CRITICAL/HIGH > MEDIUM > QUALITY
            for a in row_anoms:
                atype = str(a.get("Anomaly_Type","")).strip()
                sev   = str(a.get("Severity","")).strip()
                if atype == "시간 역전":
                    # 날짜 열만 빨강
                    for c in date_cols:
                        _fill(ws.cell(row=r, column=c), ARGB["RED"][0])
                    cnt["time_reversal"] += 1
                elif atype == "머신러닝 이상치":
                    # 심각도: CRITICAL/HIGH→주황, MEDIUM/LOW→노랑
                    if sev in ("치명적","높음","HIGH","CRITICAL"):
                        paint_row = "ORANGE" if paint_row != "ORANGE" else paint_row
                    else:
                        paint_row = "YELLOW" if paint_row not in ("ORANGE",) else paint_row
                elif atype == "데이터 품질":
                    paint_row = "PURPLE"

            if paint_row:
                argb = ARGB[paint_row][0]
                for c in range(1, ws.max_column+1):
                    _fill(ws.cell(row=r, column=c), argb)
                if paint_row == "PURPLE":
                    cnt["data_quality"] += 1
                else:
                    cnt["ml_outlier"] += 1

        wb.save(excel_file)
        return {
            "success": True,
            "message": f"색상 적용 완료 (시간역전={cnt['time_reversal']}, ML={cnt['ml_outlier']}, 품질={cnt['data_quality']})",
            **cnt,
            "backup_path": str(bak) if create_backup else None,
        }

    def add_color_legend(
        self, excel_file: Union[str, Path], _: str = DEFAULT_STAGE3_SHEET
    ) -> None:
        """
        ⚠️ 기존과 달리 '데이터 시트'를 건드리지 않고, 항상 별도 시트('색상 범례')에 작성.
        """
        excel_file = Path(excel_file)
        wb = openpyxl.load_workbook(excel_file, keep_vba=excel_file.suffix.lower()==".xlsm")
        name = "색상 범례"
        if name in wb.sheetnames:
            ws = wb[name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(name)

        ws["A1"] = "이상치 색상 범례"
        ws["B2"] = "시간 역전";     _fill(ws["A2"], ARGB["RED"][0])
        ws["B3"] = "ML 이상치(높음/치명적)"; _fill(ws["A3"], ARGB["ORANGE"][0])
        ws["B4"] = "ML 이상치(보통/낮음)";   _fill(ws["A4"], ARGB["YELLOW"][0])
        ws["B5"] = "데이터 품질";   _fill(ws["A5"], ARGB["PURPLE"][0])

        wb.save(excel_file)
import re, shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Union

import openpyxl
from openpyxl.styles import PatternFill

# ---- ARGB 정의(불투명: FF alpha). 검증 스크립트 호환 위해 00/FF 모두 허용 ----
ARGB = {
    "RED":    ("FFFF0000", {"FFFF0000", "00FF0000"}),
    "ORANGE": ("FFFFC000", {"FFFFC000", "00FFC000"}),
    "YELLOW": ("FFFFFF00", {"FFFFFF00", "00FFFF00"}),
    "PURPLE": ("FFCC99FF", {"FFCC99FF", "00CC99FF"}),
}

DATE_KEYWORDS = {
    # 한글/영문 혼합 키워드(15개 이상)
    "date","day","time","dt","입고","출고","도착","출발","반출","반입","통관",
    "선적","출항","입항","검수","검품","warehouse","site"
}

def _norm_case(s: object) -> str:
    """Case ID 정규화: 공백/특수문자 제거 + 대문자."""
    return re.sub(r"[^A-Z0-9]", "", str(s).strip().upper()) if s is not None else ""

def _is_date_col(header: str, sample_vals: List[object]) -> bool:
    h = str(header).strip().lower()
    if any(k in h for k in DATE_KEYWORDS):
        return True
    # 값 기반 휴리스틱: 30% 이상이 날짜로 파싱 가능하면 날짜열로 간주
    import pandas as pd
    s = pd.to_datetime(pd.Series(sample_vals), errors="coerce")
    non_na = s.notna().mean() if len(s) else 0.0
    return non_na >= 0.3

def _fill(cell, argb: str):
    cell.fill = PatternFill(fill_type="solid", start_color=argb, end_color=argb)

class AnomalyVisualizer:
    """
    anomalies: List[dict|dataclass] — dict 키는 Anomaly_Type/Severity/Case_ID
    """
    def __init__(self, anomalies: Iterable[object]):
        self.records: List[Dict] = []
        for a in anomalies:
            if hasattr(a, "to_dict"):
                self.records.append(a.to_dict())
            elif isinstance(a, dict):
                self.records.append(a)
        # Case → anomaly 목록(중복 허용)
        self.by_case: Dict[str, List[Dict]] = {}
        for r in self.records:
            cid = _norm_case(r.get("Case_ID", ""))
            if cid:
                self.by_case.setdefault(cid, []).append(r)

    def apply_anomaly_colors(
        self,
        excel_file: Union[str, Path],
        sheet_name: str = DEFAULT_STAGE3_SHEET,
        case_col: str = "Case No.",
        create_backup: bool = True,
    ) -> Dict:
        excel_file = Path(excel_file)
        if create_backup:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak = excel_file.with_name(f"{excel_file.stem}.backup_{ts}{excel_file.suffix}")
            shutil.copyfile(excel_file, bak)

        wb = openpyxl.load_workbook(excel_file, keep_vba=excel_file.suffix.lower()==".xlsm")
        if sheet_name not in wb.sheetnames:
            return {"success": False, "message": f"시트 없음: {sheet_name}"}
        ws = wb[sheet_name]

        # 헤더 스캔 → case 컬럼 index
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        case_col_idx = None
        for c, name in enumerate(header, 1):
            if name and "case" in str(name).lower():
                case_col_idx = c; break
        if not case_col_idx:
            return {"success": False, "message": "Case NO 열을 찾지 못함"}

        # 날짜열 식별(헤더 + 샘플 기반)
        date_cols: List[int] = []
        for c, name in enumerate(header, 1):
            sample = [ws.cell(row=r, column=c).value for r in range(2, min(ws.max_row, 50)+1)]
            if _is_date_col(name, sample):
                date_cols.append(c)

        # 색칠
        cnt = {"time_reversal": 0, "ml_outlier": 0, "data_quality": 0}
        for r in range(2, ws.max_row+1):
            raw_id = ws.cell(row=r, column=case_col_idx).value
            cid = _norm_case(raw_id)
            if not cid or cid not in self.by_case:
                continue

            # 동일 Case의 다중 이상치 처리: 시간역전(날짜열) + ML/품질(행 전체) 병행
            row_anoms = self.by_case[cid]
            paint_row = None  # ORANGE/YELLOW/PURPLE 우선순위: CRITICAL/HIGH > MEDIUM > QUALITY
            for a in row_anoms:
                atype = str(a.get("Anomaly_Type","")).strip()
                sev   = str(a.get("Severity","")).strip()
                if atype == "시간 역전":
                    # 날짜 열만 빨강
                    for c in date_cols:
                        _fill(ws.cell(row=r, column=c), ARGB["RED"][0])
                    cnt["time_reversal"] += 1
                elif atype == "머신러닝 이상치":
                    # 심각도: CRITICAL/HIGH→주황, MEDIUM/LOW→노랑
                    if sev in ("치명적","높음","HIGH","CRITICAL"):
                        paint_row = "ORANGE" if paint_row != "ORANGE" else paint_row
                    else:
                        paint_row = "YELLOW" if paint_row not in ("ORANGE",) else paint_row
                elif atype == "데이터 품질":
                    paint_row = "PURPLE"

            if paint_row:
                argb = ARGB[paint_row][0]
                for c in range(1, ws.max_column+1):
                    _fill(ws.cell(row=r, column=c), argb)
                if paint_row == "PURPLE":
                    cnt["data_quality"] += 1
                else:
                    cnt["ml_outlier"] += 1

        wb.save(excel_file)
        return {
            "success": True,
            "message": f"색상 적용 완료 (시간역전={cnt['time_reversal']}, ML={cnt['ml_outlier']}, 품질={cnt['data_quality']})",
            **cnt,
            "backup_path": str(bak) if create_backup else None,
        }

    def add_color_legend(
        self, excel_file: Union[str, Path], _: str = DEFAULT_STAGE3_SHEET
    ) -> None:
        """
        ⚠️ 기존과 달리 '데이터 시트'를 건드리지 않고, 항상 별도 시트('색상 범례')에 작성.
        """
        excel_file = Path(excel_file)
        wb = openpyxl.load_workbook(excel_file, keep_vba=excel_file.suffix.lower()==".xlsm")
        name = "색상 범례"
        if name in wb.sheetnames:
            ws = wb[name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(name)

        ws["A1"] = "이상치 색상 범례"
        ws["B2"] = "시간 역전";     _fill(ws["A2"], ARGB["RED"][0])
        ws["B3"] = "ML 이상치(높음/치명적)"; _fill(ws["A3"], ARGB["ORANGE"][0])
        ws["B4"] = "ML 이상치(보통/낮음)";   _fill(ws["A4"], ARGB["YELLOW"][0])
        ws["B5"] = "데이터 품질";   _fill(ws["A5"], ARGB["PURPLE"][0])

        wb.save(excel_file)