# -*- coding: utf-8 -*-
"""
HVDC 입출고 이상치 탐지 v2 (Hybrid / Plugin-based)
- Rule + Statistical + ML(PyOD/Sklearn) 3-Layer
- 헤더 정규화, 점수 캘리브레이션(ECDF), 30초 알림 임계
- Excel/JSON 리포트(선택), 배치/워커 파라미터
"""
from __future__ import annotations

import json
import logging
import math
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

# ----- Optional deps (graceful fallback) --------------------------------------
try:
    from sklearn.ensemble import IsolationForest
    from sklearn.preprocessing import StandardScaler

    SKLEARN_AVAILABLE = True
except Exception:
    SKLEARN_AVAILABLE = False
    IsolationForest = object  # type: ignore
    StandardScaler = object  # type: ignore

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    # PyOD는 다양한 비지도 이상치 알고리즘 제공(가능하면 사용)
    from pyod.models.iforest import IForest as PyODIForest  # type: ignore

    PYOD_AVAILABLE = True
except Exception:
    PYOD_AVAILABLE = False

# ----- Logging ----------------------------------------------------------------
logger = logging.getLogger("hvdc.anomaly")
if not logger.handlers:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )


# ----- Domain enums / schema ---------------------------------------------------
class AnomalyType(Enum):
    TIME_REVERSAL = "시간 역전"
    LOCATION_SKIP = "위치 스킵"
    EXCESSIVE_DWELL = "과도 체류"
    ML_OUTLIER = "머신러닝 이상치"
    DATA_QUALITY = "데이터 품질"


class AnomalySeverity(Enum):
    CRITICAL = "치명적"
    HIGH = "높음"
    MEDIUM = "보통"
    LOW = "낮음"


@dataclass(frozen=True)
class AnomalyRecord:
    case_id: str
    anomaly_type: AnomalyType
    severity: AnomalySeverity
    description: str
    detected_value: Optional[float]
    expected_range: Optional[Tuple[float, float]]
    location: Optional[str]
    timestamp: datetime
    risk_score: Optional[float] = None  # [0..1] calibrated

    def to_dict(self) -> Dict:
        return {
            "Case_ID": self.case_id,
            "Anomaly_Type": self.anomaly_type.value,
            "Severity": self.severity.value,
            "Description": self.description,
            "Detected_Value": self.detected_value,
            "Expected_Range": self.expected_range,
            "Location": self.location,
            "Timestamp": self.timestamp.isoformat(),
            "Risk_Score": (
                None if self.risk_score is None else round(float(self.risk_score), 4)
            ),
        }


# ----- Config -----------------------------------------------------------------
@dataclass
class DetectorConfig:
    # 헤더 정규화(동의어 매핑): Master > Slave
    column_map: Dict[str, str] = None
    # 창고/현장 열(정규화된 이름 사용)
    warehouse_columns: List[str] = None
    site_columns: List[str] = None

    # 통계 탐지 파라미터
    iqr_k: float = 1.5
    mad_k: float = 3.5

    # ML 탐지 파라미터
    use_pyod_first: bool = True
    contamination: float = 0.02  # 2% 가정(데이터에 따라 조절)
    random_state: int = 42

    # 배치/워커
    batch_size: int = 1000
    max_workers: int = 8  # (요구사항: 32 이하)

    # 알림
    alert_window_sec: int = 30
    min_risk_to_alert: float = 0.8  # 0.0~1.0

    def __post_init__(self):
        if self.column_map is None:
            # Master 헤더 이름으로 정규화
            self.column_map = {
                # 키 필드
                "Case No.": "CASE_NO",
                "CASE NO": "CASE_NO",
                "CASE_NO": "CASE_NO",
                # HVDC CODE
                "HVDC CODE": "HVDC_CODE",
                "HVDC Code": "HVDC_CODE",
                # 창고/현장(표기 변형 흡수)
                "DSV Indoor": "DSV_INDOOR",
                "DSV Al Markaz": "DSV_AL_MARKAZ",
                "AAA Storage": "AAA_STORAGE",
                "DSV Outdoor": "DSV_OUTDOOR",
                "MOSB": "MOSB",
                "Hauler Indoor": "HAULER_INDOOR",
                "DHL Warehouse": "DHL_WAREHOUSE",
                "DSV MZP": "DSV_MZP",
                "AGI": "AGI",
                "DAS": "DAS",
                "MIR": "MIR",
                "SHU": "SHU",
                # 금액/수량 등
                "금액": "AMOUNT",
                "수량": "QTY",
                "Pkg": "PKG",
            }
        if self.warehouse_columns is None:
            self.warehouse_columns = [
                "AAA_STORAGE",
                "DSV_AL_MARKAZ",
                "DSV_INDOOR",
                "DSV_MZP",
                "DSV_OUTDOOR",
                "HAULER_INDOOR",
                "MOSB",
                "DHL_WAREHOUSE",
            ]
        if self.site_columns is None:
            self.site_columns = ["AGI", "DAS", "MIR", "SHU"]


# ----- Utilities ---------------------------------------------------------------
class HeaderNormalizer:
    def __init__(self, column_map: Dict[str, str]):
        self.map = {k.lower(): v for k, v in column_map.items()}

    def normalize(self, df: pd.DataFrame) -> pd.DataFrame:
        new_cols = []
        for c in df.columns:
            key = str(c).strip().lower()
            new_cols.append(self.map.get(key, str(c).strip().upper().replace(" ", "_")))
        df = df.copy()
        df.columns = new_cols
        return df


class DataQualityValidator:
    """간단/빠른 정합성 검증(필요 시 Great Expectations/Pandera로 확장)"""

    HVDC_PATTERN = r"^HVDC-ADOPT-\d{3}-\d{4}$"

    def validate(self, df: pd.DataFrame) -> List[str]:
        issues: List[str] = []
        if "CASE_NO" not in df.columns:
            issues.append("필수 필드 누락: CASE_NO")
        else:
            dup = df["CASE_NO"].astype(str).duplicated().sum()
            if dup:
                issues.append(f"CASE_NO 중복 {dup}건")

        if "HVDC_CODE" in df.columns:
            bad = ~df["HVDC_CODE"].astype(str).str.match(self.HVDC_PATTERN, na=False)
            n_bad = int(bad.sum())
            if n_bad:
                issues.append(f"HVDC_CODE 형식 오류 {n_bad}건")

        # 수치형 기본 체크
        for num_col in ("AMOUNT", "QTY", "PKG"):
            if num_col in df.columns:
                nonnum = (
                    pd.to_numeric(df[num_col], errors="coerce").isna()
                    & df[num_col].notna()
                )
                if int(nonnum.sum()):
                    issues.append(f"{num_col} 비숫자 값 {int(nonnum.sum())}건")

        # 날짜 변환 가능성(창고/현장)
        for col in df.columns:
            if col in set(cfg.warehouse_columns + cfg.site_columns):
                # 허용: 결측/문자열 → to_datetime 변환 실패율만 집계
                s = pd.to_datetime(df[col], errors="coerce")
                fail_mask = (df[col].notna()) & (s.isna())
                fail = int(fail_mask.sum())
                if fail:
                    issues.append(f"{col}: 날짜 변환 실패 {fail}건")

        return issues


# ----- Feature engineering -----------------------------------------------------
class FeatureBuilder:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg

    def build(
        self, df: pd.DataFrame
    ) -> Tuple[pd.DataFrame, List[Tuple[str, str, int]]]:
        """
        반환:
          - 행 단위 피처(정규화된 CASE_NO index)
          - dwell 목록[(case_id, location, dwell_days)]
        """
        rows = []
        dwell_list: List[Tuple[str, str, int]] = []

        for _, row in df.iterrows():
            case_id = str(row.get("CASE_NO", "NA"))
            points: List[Tuple[str, pd.Timestamp]] = []
            for col in self.cfg.warehouse_columns + self.cfg.site_columns:
                if col in row.index and pd.notna(row[col]):
                    dt = pd.to_datetime(row[col], errors="coerce")
                    if pd.notna(dt):
                        points.append((col, dt))

            points.sort(key=lambda x: x[1])
            if len(points) >= 2:
                # dwell(다음 지점까지 체류일)
                for (loc_a, t_a), (loc_b, t_b) in zip(points[:-1], points[1:]):
                    dwell = max(0, (t_b - t_a).days)
                    dwell_list.append((case_id, loc_a, dwell))

            # scalar features
            n_touch = len(points)
            first_ts = points[0][1].value if n_touch else np.nan
            last_ts = points[-1][1].value if n_touch else np.nan
            total_days = np.nan
            if n_touch >= 2:
                total_days = (points[-1][1] - points[0][1]).days

            rows.append(
                dict(
                    CASE_NO=case_id,
                    TOUCH_COUNT=n_touch,
                    TOTAL_DAYS=total_days,
                    FIRST_TS=first_ts,
                    LAST_TS=last_ts,
                    AMOUNT=pd.to_numeric(row.get("AMOUNT", np.nan), errors="coerce"),
                    QTY=pd.to_numeric(row.get("QTY", np.nan), errors="coerce"),
                    PKG=pd.to_numeric(row.get("PKG", np.nan), errors="coerce"),
                )
            )

        feat = pd.DataFrame(rows).set_index("CASE_NO", drop=True)
        return feat, dwell_list


# ----- Statistical detectors ---------------------------------------------------
class StatDetector:
    def __init__(self, iqr_k: float = 1.5, mad_k: float = 3.5):
        self.iqr_k = iqr_k
        self.mad_k = mad_k

    def iqr_outliers(
        self, dwell_list: List[Tuple[str, str, int]]
    ) -> List[AnomalyRecord]:
        if not dwell_list:
            return []
        vals = np.array([d for _, _, d in dwell_list], dtype=float)
        q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
        iqr = q3 - q1
        lo, hi = q1 - self.iqr_k * iqr, q3 + self.iqr_k * iqr

        out = []
        for case_id, loc, d in dwell_list:
            if d > hi:
                sev = AnomalySeverity.HIGH if d > 2 * hi else AnomalySeverity.MEDIUM
                out.append(
                    AnomalyRecord(
                        case_id=case_id,
                        anomaly_type=AnomalyType.EXCESSIVE_DWELL,
                        severity=sev,
                        description=f"{loc}에서 {d}일 체류 (정상≈{lo:.1f}~{hi:.1f}일)",
                        detected_value=float(d),
                        expected_range=(float(lo), float(hi)),
                        location=loc,
                        timestamp=datetime.now(),
                    )
                )
        return out


# ----- Rule-based detectors ----------------------------------------------------
class RuleDetector:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg

    def time_reversal(self, row: pd.Series) -> Optional[AnomalyRecord]:
        pts: List[Tuple[str, pd.Timestamp]] = []
        for col in self.cfg.warehouse_columns + self.cfg.site_columns:
            if col in row.index and pd.notna(row[col]):
                ts = pd.to_datetime(row[col], errors="coerce")
                if pd.notna(ts):
                    pts.append((col, ts))
        if len(pts) < 2:
            return None

        # 시간 역전이 있는지 확인 (정렬 전후 비교)
        pts_sorted = sorted(pts, key=lambda x: x[1])
        original_order = [p[0] for p in pts]
        sorted_order = [p[0] for p in pts_sorted]

        # 원래 순서와 시간순 정렬된 순서가 다르면 시간 역전
        if original_order != sorted_order:
            return AnomalyRecord(
                case_id=str(row.get("CASE_NO", "NA")),
                anomaly_type=AnomalyType.TIME_REVERSAL,
                severity=AnomalySeverity.HIGH,
                description="시간 역전(순서 불일치) 발생",
                detected_value=None,
                expected_range=None,
                location=None,
                timestamp=datetime.now(),
            )
        return None

    def location_skip(self, row: pd.Series) -> Optional[AnomalyRecord]:
        """업무상 불가능한 순번 스킵(간단 규칙 예시)"""
        # 필요 시 프로젝트 룰 카테고리(E)로 강화
        return None


# ----- ML detector (with calibration) -----------------------------------------
class ECDFCalibrator:
    """점수 분포 기반 위험도 보정: 낮을수록 정상인 decision_function/score를 [0..1] 위험도로 변환"""

    def __init__(self):
        self.ref: Optional[np.ndarray] = None

    def fit(self, raw_scores: np.ndarray) -> "ECDFCalibrator":
        self.ref = np.sort(raw_scores.astype(float))
        return self

    def transform(self, raw_scores: np.ndarray) -> np.ndarray:
        if self.ref is None or len(self.ref) == 0:
            return np.clip(
                (raw_scores - raw_scores.min()) / (raw_scores.ptp() + 1e-9), 0, 1
            )
        # 원점수가 "작을수록 이상"이라고 가정 → 분위수로 위험도 산출
        # 위험도 = 1 - ECDF(x)
        idx = np.searchsorted(self.ref, raw_scores, side="right")
        ecdf = idx / float(len(self.ref))
        risk = 1.0 - ecdf
        return np.clip(risk, 0, 1)


class MLDetector:
    def __init__(
        self,
        contamination: float = 0.02,
        random_state: int = 42,
        use_pyod_first: bool = True,
    ):
        self.contamination = contamination
        self.random_state = random_state
        self.use_pyod_first = use_pyod_first and PYOD_AVAILABLE
        self.model = None
        self.scaler = None
        self.calib = ECDFCalibrator()

    def fit_predict(self, X: pd.DataFrame) -> Tuple[np.ndarray, np.ndarray]:
        """return: (y_pred[0/1], risk[0..1])"""
        if X.empty or (not SKLEARN_AVAILABLE and not PYOD_AVAILABLE):
            return np.zeros(len(X), dtype=int), np.zeros(len(X), dtype=float)

        self.scaler = StandardScaler() if SKLEARN_AVAILABLE else None
        Xs = self.scaler.fit_transform(X.values) if self.scaler else X.values

        if self.use_pyod_first:
            # PyOD IForest
            self.model = PyODIForest(
                contamination=self.contamination, random_state=self.random_state
            )
            self.model.fit(Xs)
            # PyOD의 decision_scores_: 값이 클수록 이상치
            raw = np.asarray(self.model.decision_scores_, dtype=float)
            risk = ECDFCalibrator().fit(raw).transform(raw)
            y = (risk >= (1 - self.contamination)).astype(int)
            return y, risk

        # Sklearn IsolationForest (decision_function: 클수록 정상)
        self.model = IsolationForest(
            contamination=self.contamination,
            random_state=self.random_state,
            n_estimators=256,
        )
        self.model.fit(Xs)
        dec = self.model.decision_function(Xs)  # +: 정상, -: 이상
        # 위험도 = 1 - ECDF(dec)
        risk = ECDFCalibrator().fit(dec).transform(dec)
        y = (risk >= (1 - self.contamination)).astype(int)
        return y, risk


# ----- Alert manager -----------------------------------------------------------
class AlertManager:
    def __init__(self, window_sec: int = 30, min_risk: float = 0.8):
        self.window = window_sec
        self.min_risk = min_risk
        self._last: Optional[float] = None

    def on_anomaly(self, risk: float) -> bool:
        """
        이상치가 연달아 발생하고 window 내 해소되지 않으면 True 반환(알림)
        """
        now = time.time()
        if risk < self.min_risk:
            self._last = None
            return False
        if self._last is None:
            self._last = now
            return False
        if (now - self._last) >= self.window:
            self._last = now
            return True
        return False


# ----- Orchestrator ------------------------------------------------------------
class HybridAnomalyDetector:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg
        self.normalizer = HeaderNormalizer(cfg.column_map)
        self.validator = DataQualityValidator()
        self.rule = RuleDetector(cfg)
        self.stat = StatDetector(cfg.iqr_k, cfg.mad_k)
        self.ml = MLDetector(cfg.contamination, cfg.random_state, cfg.use_pyod_first)
        self.alert = AlertManager(cfg.alert_window_sec, cfg.min_risk_to_alert)
        self._summary: Dict = {}

    def run(
        self,
        df_raw: pd.DataFrame,
        export_excel: Optional[str] = None,
        export_json: Optional[str] = None,
    ) -> Dict:
        df = self.normalizer.normalize(df_raw)
        issues = self.validator.validate(df)
        anomalies: List[AnomalyRecord] = []
        if issues:
            logger.warning(f"데이터 품질 이슈: {issues}")
            anomalies.extend(
                [
                    AnomalyRecord(
                        case_id=(
                            str(df.iloc[0].get("CASE_NO", "NA")) if len(df) else "NA"
                        ),
                        anomaly_type=AnomalyType.DATA_QUALITY,
                        severity=AnomalySeverity.MEDIUM,
                        description="; ".join(issues),
                        detected_value=None,
                        expected_range=None,
                        location=None,
                        timestamp=datetime.now(),
                    )
                ]
            )

        # Feature build
        fb = FeatureBuilder(self.cfg)
        feat, dwell_list = fb.build(df)

        # 1) Rule
        for _, row in df.iterrows():
            r1 = self.rule.time_reversal(row)
            if r1:
                anomalies.append(r1)

        # 2) Stat
        anomalies.extend(self.stat.iqr_outliers(dwell_list))

        # 3) ML
        use_cols = [
            c
            for c in ["TOUCH_COUNT", "TOTAL_DAYS", "AMOUNT", "QTY", "PKG"]
            if c in feat.columns
        ]
        X = feat[use_cols].fillna(0.0)
        y, risk = self.ml.fit_predict(X)
        if len(X):
            for i, (case_id, yi, ri) in enumerate(zip(X.index, y, risk)):
                if yi == 1:
                    sev = (
                        AnomalySeverity.CRITICAL
                        if ri >= 0.98
                        else (
                            AnomalySeverity.HIGH
                            if ri >= 0.9
                            else AnomalySeverity.MEDIUM
                        )
                    )
                    rec = AnomalyRecord(
                        case_id=str(case_id),
                        anomaly_type=AnomalyType.ML_OUTLIER,
                        severity=sev,
                        description=f"ML 이상치(위험도 {ri:.3f})",
                        detected_value=float(ri),
                        expected_range=None,
                        location=None,
                        timestamp=datetime.now(),
                        risk_score=float(ri),
                    )
                    anomalies.append(rec)
                    # 30초 알림 윈도우
                    if self.alert.on_anomaly(ri):
                        logger.error("🚨 30초 내 복구 없음: 알림 트리거")

        # Summary
        self._summary = self._build_summary(anomalies)

        # Export
        if export_excel:
            self._export_excel(Path(export_excel), anomalies, feat)
        if export_json:
            self._export_json(Path(export_json), anomalies)

        return {
            "summary": self._summary,
            "anomalies": anomalies,
            "features": feat,
        }

    def _build_summary(self, anomalies: List[AnomalyRecord]) -> Dict:
        by_type: Dict[str, int] = {}
        by_sev: Dict[str, int] = {}
        for a in anomalies:
            by_type[a.anomaly_type.value] = by_type.get(a.anomaly_type.value, 0) + 1
            by_sev[a.severity.value] = by_sev.get(a.severity.value, 0) + 1
        return {
            "total": len(anomalies),
            "by_type": by_type,
            "by_severity": by_sev,
        }

    # -------- Exporters --------
    def _export_json(self, path: Path, anomalies: List[AnomalyRecord]) -> None:
        data = [a.to_dict() for a in anomalies]
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        logger.info(f"JSON 저장: {path}")

    def _export_excel(
        self, path: Path, anomalies: List[AnomalyRecord], feat: pd.DataFrame
    ) -> None:
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl 미설치로 Excel 생략")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "HVDC 이상치 탐지 리포트"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"생성일시: {datetime.now():%Y-%m-%d %H:%M:%S}"

        # Summary block
        s = self._summary
        ws["A4"] = "총 이상치"
        ws["B4"] = s["total"]
        ws["A6"] = "유형별"
        r = 7
        for k, v in s["by_type"].items():
            ws.cell(r, 1).value = k
            ws.cell(r, 2).value = v
            r += 1
        r += 1
        ws.cell(r, 1).value = "심각도별"
        r += 1
        for k, v in s["by_severity"].items():
            ws.cell(r, 1).value = k
            ws.cell(r, 2).value = v
            r += 1

        # Anomalies
        ws2 = wb.create_sheet("Anomalies")
        cols = list(AnomalyRecord.__annotations__.keys())
        for i, c in enumerate(cols, start=1):
            ws2.cell(1, i).value = c
            ws2.cell(1, i).font = Font(bold=True)
        for ridx, a in enumerate(anomalies, start=2):
            row = a.to_dict()
            for cidx, c in enumerate(cols, start=1):
                ws2.cell(ridx, cidx).value = row.get(
                    c if c != "risk_score" else "Risk_Score"
                )

        # Features
        ws3 = wb.create_sheet("Features")
        for r in dataframe_to_rows(feat.reset_index(), index=False, header=True):
            ws3.append(r)

        wb.save(path)
        logger.info(f"Excel 저장: {path}")


# -------- convenience runner ---------------------------------------------------
cfg = DetectorConfig()
fb = FeatureBuilder(cfg)


def main():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Excel/CSV 파일 경로")
    ap.add_argument("--sheet", default=None, help="Excel 시트명(옵션)")
    ap.add_argument("--excel-out", default=None, help="결과 Excel 경로")
    ap.add_argument("--json-out", default=None, help="결과 JSON 경로")
    ap.add_argument("--visualize", action="store_true", help="원본 파일에 색상 표시")
    ap.add_argument("--case-col", default="Case No.", help="Case NO 컬럼명")
    ap.add_argument("--no-backup", action="store_true", help="백업 생성 안함")
    args = ap.parse_args()

    p = Path(args.input)
    if p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(p, sheet_name=args.sheet)
    else:
        df = pd.read_csv(p, encoding="utf-8")

    # DataFrame인지 확인
    if not isinstance(df, pd.DataFrame):
        logger.error(f"데이터 로딩 실패: {type(df)}")
        return

    det = HybridAnomalyDetector(cfg)
    result = det.run(df, export_excel=args.excel_out, export_json=args.json_out)

    s = result["summary"]
    logger.info(f"총 이상치: {s['total']}")
    logger.info(f"유형별: {s['by_type']}")
    logger.info(f"심각도별: {s['by_severity']}")

    # 색상 시각화 옵션
    if args.visualize:
        try:
            from anomaly_visualizer import AnomalyVisualizer

            logger.info("🎨 원본 파일에 색상 표시 시작...")
            visualizer = AnomalyVisualizer(result["anomalies"])

            viz_result = visualizer.apply_anomaly_colors(
                excel_file=args.input,
                sheet_name=args.sheet or "Case List",
                case_col=args.case_col,
                create_backup=not args.no_backup,
            )

            if viz_result["success"]:
                # 색상 범례 추가
                visualizer.add_color_legend(args.input, args.sheet or "Case List")
                logger.info("ℹ️ 범례는 '색상 범례' 시트에만 작성되어 데이터 행에는 영향을 주지 않습니다.")
                logger.info(f"✅ 색상 표시 완료: {viz_result['message']}")
                logger.info(
                    f"  - 시간 역전: {viz_result['time_reversal_count']}건 (빨강)"
                )
                logger.info(
                    f"  - ML 이상치: {viz_result['ml_outlier_count']}건 (주황/노랑)"
                )
                logger.info(
                    f"  - 데이터 품질: {viz_result['data_quality_count']}건 (보라)"
                )
                if viz_result.get("backup_path"):
                    logger.info(f"  - 백업 파일: {viz_result['backup_path']}")
            else:
                logger.error(f"❌ 색상 표시 실패: {viz_result['message']}")

        except ImportError as e:
            logger.error(f"❌ AnomalyVisualizer 모듈을 찾을 수 없습니다: {e}")
        except Exception as e:
            logger.error(f"❌ 색상 표시 중 오류 발생: {e}")


if __name__ == "__main__":
    main()
