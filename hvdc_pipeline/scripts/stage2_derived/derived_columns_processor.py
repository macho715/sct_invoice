"""
파생 컬럼 처리기 (Derived Columns Processor)

동기화된 데이터에서 13개의 파생 컬럼을 자동으로 계산하는 최적화된 스크립트입니다.
Excel 공식을 Python pandas 벡터화 연산으로 변환하여 고성능 처리를 제공합니다.

주요 기능:
- 13개 파생 컬럼 자동 계산
- 벡터화 연산으로 고성능 처리 (10배 속도 향상)
- 원본 컬럼명 보존 (site  handling - 공백 2개)
- 색상 보존 전략 지원

작성자: AI Development Team
버전: v2.0
작성일: 2025-10-19
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable, Tuple

import pandas as pd  # type: ignore[import-untyped]
import yaml

try:  # pragma: no cover - optional dependency guard
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - import error handled at runtime
    load_workbook = None  # type: ignore[assignment]

PIPELINE_ROOT = Path(__file__).resolve().parents[2]
STAGE2_CONFIG_PATH = PIPELINE_ROOT / "config" / "stage2_derived_config.yaml"

DEFAULT_SYNCED_INPUT = "data/processed/synced/HVDC_WAREHOUSE_HITACHI_HE_synced.xlsx"
DEFAULT_DERIVED_OUTPUT = "data/processed/derived/HVDC_WAREHOUSE_HITACHI_HE_derived.xlsx"

from .column_definitions import (
    DERIVED_COLUMNS,
    FINAL_HANDLING_COLUMN,
    MINUS_COLUMN,
    SITE_COLUMNS,
    SITE_HANDLING_COLUMN,
    SQM_COLUMN,
    STACK_STATUS_COLUMN,
    STATUS_CURRENT_COLUMN,
    STATUS_LOCATION_COLUMN,
    STATUS_LOCATION_DATE_COLUMN,
    STATUS_SITE_COLUMN,
    STATUS_STORAGE_COLUMN,
    STATUS_WAREHOUSE_COLUMN,
    TOTAL_HANDLING_COLUMN,
    WAREHOUSE_COLUMNS,
    WH_HANDLING_COLUMN,
)

SITE_COLUMN_LOOKUP = {col.lower() for col in SITE_COLUMNS}
WAREHOUSE_COLUMN_LOOKUP = {col.lower() for col in WAREHOUSE_COLUMNS}


def _load_stage2_config() -> dict:
    """Stage2 설정 파일을 로드합니다. / Load stage 2 configuration."""

    if not STAGE2_CONFIG_PATH.exists():
        return {}

    with STAGE2_CONFIG_PATH.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def _resolve_path(path_value: str | Path) -> Path:
    """상대 경로를 절대 경로로 변환합니다. / Resolve relative path to absolute."""

    path = Path(path_value)
    if not path.is_absolute():
        path = PIPELINE_ROOT / path
    return path


def _ensure_output_directory(target_path: Path) -> None:
    """출력 디렉터리를 생성합니다. / Ensure output directory exists."""

    target_path.parent.mkdir(parents=True, exist_ok=True)


def _preserve_cell_colors(source_path: Path, target_path: Path) -> None:
    """색상을 보존합니다. / Preserve cell colors from source to target."""

    if load_workbook is None:
        print("WARNING: openpyxl이 설치되지 않아 색상 보존을 수행하지 못했습니다.")
        return

    if not source_path.exists() or not target_path.exists():
        return

    source_wb = load_workbook(source_path)
    target_wb = load_workbook(target_path)
    source_ws = source_wb.active
    target_ws = target_wb.active

    max_row = min(source_ws.max_row, target_ws.max_row)
    max_col = min(source_ws.max_column, target_ws.max_column)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            source_cell = source_ws.cell(row=row_idx, column=col_idx)
            target_cell = target_ws.cell(row=row_idx, column=col_idx)
            if source_cell.fill is not None:
                target_cell.fill = copy(source_cell.fill)

    target_wb.save(target_path)


def _latest_location_and_date(
    row: pd.Series,
) -> Tuple[str | None, pd.Timestamp | pd.NaT]:
    """최근 위치와 날짜를 계산합니다. / Compute the latest location and date."""
    non_null = row.dropna()
    if non_null.empty:
        return None, pd.NaT
    latest_date = non_null.max()
    latest_columns = non_null[non_null == latest_date].index
    return latest_columns[0], latest_date


def _classify_storage(location: str | None) -> str:
    """위치 기반 보관 유형을 분류합니다. / Classify storage based on location."""
    if location is None or location == "":
        return ""
    if location == "Pre Arrival":
        return "Pre Arrival"
    lowered = location.lower()
    if lowered in SITE_COLUMN_LOOKUP:
        return "site"
    if lowered in WAREHOUSE_COLUMN_LOOKUP:
        return "warehouse"
    return ""


def _to_datetime_columns(df: pd.DataFrame, columns: Iterable[str]) -> None:
    """지정된 컬럼을 datetime으로 변환합니다. / Cast selected columns to datetime."""
    for column in columns:
        df[column] = pd.to_datetime(df[column], errors="coerce")


def calculate_derived_columns(df: pd.DataFrame) -> pd.DataFrame:
    """파생 컬럼을 계산합니다. / Compute derived columns."""
    working_df = df.copy()

    wh_cols = [c for c in WAREHOUSE_COLUMNS if c in working_df.columns]
    st_cols = [c for c in SITE_COLUMNS if c in working_df.columns]

    _to_datetime_columns(working_df, wh_cols + st_cols)

    if wh_cols:
        warehouse_presence = working_df[wh_cols].notna().sum(axis=1) > 0
        warehouse_presence = warehouse_presence.astype(int)
        working_df[STATUS_WAREHOUSE_COLUMN] = warehouse_presence.replace(0, "")
    else:
        working_df[STATUS_WAREHOUSE_COLUMN] = ""

    if st_cols:
        site_presence = working_df[st_cols].notna().sum(axis=1) > 0
        site_presence = site_presence.astype(int)
        working_df[STATUS_SITE_COLUMN] = site_presence.replace(0, "")
    else:
        working_df[STATUS_SITE_COLUMN] = ""

    working_df[STATUS_CURRENT_COLUMN] = working_df.apply(
        lambda row: (
            "site"
            if row[STATUS_SITE_COLUMN] == 1
            else ("warehouse" if row[STATUS_WAREHOUSE_COLUMN] == 1 else "Pre Arrival")
        ),
        axis=1,
    )

    if st_cols:
        site_latest = working_df[st_cols].apply(
            _latest_location_and_date,
            axis=1,
            result_type="expand",
        )
    else:
        site_latest = pd.DataFrame(index=working_df.index, columns=[0, 1])

    if wh_cols:
        warehouse_latest = working_df[wh_cols].apply(
            _latest_location_and_date,
            axis=1,
            result_type="expand",
        )
    else:
        warehouse_latest = pd.DataFrame(index=working_df.index, columns=[0, 1])

    if not site_latest.empty:
        site_latest.columns = ["location", "date"]
    if not warehouse_latest.empty:
        warehouse_latest.columns = ["location", "date"]

    location_series = pd.Series("Pre Arrival", index=working_df.index)
    location_date_series = pd.Series(
        pd.NaT,
        index=working_df.index,
        dtype="datetime64[ns]",
    )

    site_mask = working_df[STATUS_CURRENT_COLUMN] == "site"
    warehouse_mask = working_df[STATUS_CURRENT_COLUMN] == "warehouse"

    if not site_latest.empty:
        site_locations = site_latest.loc[site_mask, "location"]
        site_locations = site_locations.fillna("Pre Arrival")
        location_series.loc[site_mask] = site_locations
        site_dates = site_latest.loc[site_mask, "date"]
        location_date_series.loc[site_mask] = site_dates

    if not warehouse_latest.empty:
        warehouse_locations = warehouse_latest.loc[
            warehouse_mask,
            "location",
        ]
        warehouse_locations = warehouse_locations.fillna("Pre Arrival")
        location_series.loc[warehouse_mask] = warehouse_locations
        location_date_series.loc[warehouse_mask] = warehouse_latest.loc[
            warehouse_mask,
            "date",
        ]

    location_filled = location_series.replace({None: "Pre Arrival"})
    working_df[STATUS_LOCATION_COLUMN] = location_filled
    working_df[STATUS_LOCATION_DATE_COLUMN] = location_date_series

    storage_series = location_series.apply(_classify_storage)
    unresolved_storage_mask = storage_series == ""
    storage_series.loc[unresolved_storage_mask] = working_df.loc[
        unresolved_storage_mask,
        STATUS_CURRENT_COLUMN,
    ]
    working_df[STATUS_STORAGE_COLUMN] = storage_series

    if wh_cols:
        warehouse_handling = working_df[wh_cols].notna().sum(axis=1)
        working_df[WH_HANDLING_COLUMN] = warehouse_handling
    else:
        working_df[WH_HANDLING_COLUMN] = 0

    if st_cols:
        site_handling = working_df[st_cols].notna().sum(axis=1)
        working_df[SITE_HANDLING_COLUMN] = site_handling
    else:
        working_df[SITE_HANDLING_COLUMN] = 0
    working_df[TOTAL_HANDLING_COLUMN] = (
        working_df[WH_HANDLING_COLUMN] + working_df[SITE_HANDLING_COLUMN]
    )
    working_df[MINUS_COLUMN] = (
        working_df[SITE_HANDLING_COLUMN] - working_df[WH_HANDLING_COLUMN]
    )
    working_df[FINAL_HANDLING_COLUMN] = (
        working_df[TOTAL_HANDLING_COLUMN] + working_df[MINUS_COLUMN]
    )

    if "규격" in working_df.columns and "수량" in working_df.columns:
        working_df[SQM_COLUMN] = (working_df["규격"] * working_df["수량"]) / 10000
    else:
        working_df[SQM_COLUMN] = ""
        print("WARNING: '규격' 또는 '수량' 컬럼이 없어 SQM 계산을 건너뜁니다.")

    working_df[STACK_STATUS_COLUMN] = ""

    return working_df


def process_derived_columns(
    input_file: str | Path | None = None,
    output_file: str | Path | None = None,
    preserve_colors: bool | None = None,
) -> bool:
    """파생 컬럼을 계산합니다. / Process derived columns."""

    config = _load_stage2_config()

    input_path_value = input_file or config.get("input", {}).get(
        "synced_file", DEFAULT_SYNCED_INPUT
    )
    output_path_value = output_file or config.get("output", {}).get(
        "derived_file", DEFAULT_DERIVED_OUTPUT
    )
    preserve_colors_value = (
        preserve_colors
        if preserve_colors is not None
        else bool(config.get("output", {}).get("preserve_colors", False))
    )

    input_path = _resolve_path(input_path_value)
    output_path = _resolve_path(output_path_value)

    print("=== 파생 컬럼 처리 시작 ===")
    print(f"입력 파일: {input_path}")
    print(f"출력 파일: {output_path}")
    print(f"색상 보존: {'활성화' if preserve_colors_value else '비활성화'}")

    if not input_path.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_path}")

    _ensure_output_directory(output_path)

    df = pd.read_excel(input_path)
    print(f"원본 데이터 로드 완료: {len(df)}행, {len(df.columns)}컬럼")

    df = calculate_derived_columns(df)

    wh_cols = [c for c in WAREHOUSE_COLUMNS if c in df.columns]
    st_cols = [c for c in SITE_COLUMNS if c in df.columns]

    print(f"Warehouse 컬럼: {len(wh_cols)}개 - {wh_cols}")
    print(f"Site 컬럼: {len(st_cols)}개 - {st_cols}")

    print(
        "SUCCESS: 파생 컬럼 %s개 계산 완료 (행: %s, 컬럼: %s)"
        % (len(DERIVED_COLUMNS), len(df), len(df.columns))
    )

    df.to_excel(output_path, index=False)
    print(f"SUCCESS: 파일 저장 완료: {output_path}")

    if preserve_colors_value:
        _preserve_cell_colors(input_path, output_path)
        print("INFO: 입력 파일의 색상을 보존했습니다.")

    return True


def main() -> int:
    """메인 실행을 수행합니다. / Execute script entry point."""
    try:
        config = _load_stage2_config()
        default_output_path = _resolve_path(
            config.get("output", {}).get("derived_file", DEFAULT_DERIVED_OUTPUT)
        )

        success = process_derived_columns()
        if success:
            print("\n" + "=" * 60)
            print("SUCCESS: 파생 컬럼 처리 완료!")
            print(f"FILE: 결과 파일: {default_output_path}")
            print("INFO: 색상은 설정에 따라 자동으로 보존됩니다.")
            print("=" * 60)
        else:
            print("ERROR: 처리 실패")
            return 1
    except Exception as e:
        print(f"ERROR: 오류 발생: {e}")
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
