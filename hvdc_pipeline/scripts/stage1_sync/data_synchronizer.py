# -*- coding: utf-8 -*-
"""
DataSynchronizer v2.9 (Hard-override Date + Colorize)
- Master always takes precedence
- Date columns: always write when Master has value; highlight only when logical date changed
- Non-date columns: overwrite when Master has non-null (configurable)
- New cases appended; entire row highlighted yellow
- ExcelFormatter is called after save to color cells/rows
"""

from __future__ import annotations
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# ===== Config =====
ORANGE = "FFC000"  # changed date cell
YELLOW = "FFFF00"  # new row
DATE_KEYS = [
    "ETD/ATD",
    "ETA/ATA",
    "DHL Warehouse",
    "DSV Indoor",
    "DSV Al Markaz",
    "DSV Outdoor",
    "AAA  Storage",
    "Hauler Indoor",
    "DSV MZP",
    "MOSB",
    "Shifting",
    "MIR",
    "SHU",
    "DAS",
    "AGI",
]
ALWAYS_OVERWRITE_NONDATE = True  # Master non-null overwrites


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(h).strip().lower())


def _is_date_col(col_name: str) -> bool:
    def norm(s: str) -> str:
        # 대소문자/공백/슬래시/하이픈 차이 제거 (예: "ETD/ATD" ≒ "etd atd")
        return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())

    cn = norm(col_name)
    return any(norm(k) == cn for k in DATE_KEYS)


def _to_date(val) -> Optional[pd.Timestamp]:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        if isinstance(val, pd.Timestamp):
            return val
        return pd.to_datetime(val, errors="coerce")
    except Exception:
        return None


@dataclass
class Change:
    row_index: int
    column_name: str
    old_value: Any
    new_value: Any
    change_type: str  # "date_update" | "field_update" | "new_record"


@dataclass
class ChangeTracker:
    """Minimal tracker compatible with ExcelFormatter coloring."""

    changes: List[Change] = field(default_factory=list)
    new_cases: Dict[str, Dict[str, Any]] = field(default_factory=dict)

    def add_change(self, **kw):
        # Accept dicts for compatibility
        self.changes.append(
            Change(
                row_index=int(kw.get("row_index", -1)),
                column_name=str(kw.get("column_name", "")),
                old_value=kw.get("old_value"),
                new_value=kw.get("new_value"),
                change_type=str(kw.get("change_type", "field_update")),
            )
        )

    def log_new_case(
        self, case_no: str, row_data: Dict[str, Any], row_index: Optional[int] = None
    ):
        self.new_cases[str(case_no)] = dict(row_data or {})
        if row_index is not None:
            self.add_change(
                row_index=row_index,
                column_name="",
                old_value=None,
                new_value=None,
                change_type="new_record",
            )


@dataclass
class SyncResult:
    success: bool
    message: str
    output_path: str
    stats: Dict[str, Any]


class DataSynchronizerV29:
    def __init__(self, date_keys: Optional[List[str]] = None) -> None:
        self.date_keys = date_keys or DATE_KEYS
        self.change_tracker = ChangeTracker()

    # ---- helper: dates equal ignoring format ----
    def _dates_equal(self, a, b) -> bool:
        da = _to_date(a)
        db = _to_date(b)
        if da is None and db is None:
            return True
        if da is None or db is None:
            return False
        # Handle pd.NaT
        if pd.isna(da) or pd.isna(db):
            return pd.isna(da) and pd.isna(db)
        return da.normalize() == db.normalize()

    def _case_col(self, df: pd.DataFrame) -> Optional[str]:
        # try flexible matches: 'case', 'case no', 'case_no'...
        patterns = [r"^case(\s*no\.?)?$", r"^case_no$", r"^sku$", r"^case$"]
        for col in df.columns:
            s = str(col).strip().lower()
            for p in patterns:
                if re.match(p, s):
                    return col
        return None

    def _build_index(self, df: pd.DataFrame, case_col: str) -> Dict[str, List[int]]:
        # ⚠️ 동일 Case No. 다수 행을 모두 업데이트하기 위해 multi-map 구성
        idx: Dict[str, List[int]] = {}
        import re

        series = df[case_col].fillna("").astype(str).str.strip().str.upper()
        # 정규식으로 특수문자 제거
        series = series.apply(lambda x: re.sub(r"[^A-Z0-9]", "", x))

        for i, v in enumerate(series.tolist()):
            if not v:
                continue
            idx.setdefault(v, []).append(i)
        return idx

    def _apply_updates(
        self, master: pd.DataFrame, wh: pd.DataFrame, case_col_m: str, case_col_w: str
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        stats = dict(updates=0, date_updates=0, field_updates=0, appends=0)
        wh_index = self._build_index(wh, case_col_w)
        dup_cases = {k: v for k, v in wh_index.items() if len(v) > 1}

        # map headers (case-insensitively) to enable cross-frame copy
        m_norm = {_norm_header(c): c for c in master.columns}
        w_norm = {_norm_header(c): c for c in wh.columns}

        # for columns that share the same normalized header, we consider them "aligned"
        aligned = [(m_norm[k], w_norm[k]) for k in sorted(set(m_norm) & set(w_norm))]

        for mi, mrow in master.iterrows():
            import re

            key = (
                re.sub(r"[^A-Z0-9]", "", str(mrow[case_col_m]).strip().upper())
                if pd.notna(mrow[case_col_m])
                else ""
            )
            if not key:
                continue

            if key not in wh_index:
                # append new
                append_row = {wcol: mrow[mcol] for (mcol, wcol) in aligned}
                wh = pd.concat([wh, pd.DataFrame([append_row])], ignore_index=True)
                new_index = len(wh) - 1
                stats["appends"] += 1
                # track new case for yellow highlight
                self.change_tracker.log_new_case(
                    case_no=key, row_data=append_row, row_index=new_index
                )
                continue

            target_rows = wh_index[key]  # 리스트 반환
            for wi in target_rows:  # 동일 Case No. 모든 행에 반영
                for mcol, wcol in aligned:
                    mval = mrow[mcol]
                    wval = (
                        wh.at[wi, wcol] if wi < len(wh) and wcol in wh.columns else None
                    )
                    is_date = _is_date_col(wcol)

                    if is_date:
                        if pd.notna(mval) and not self._dates_equal(mval, wval):
                            stats["updates"] += 1
                            stats["date_updates"] += 1
                            wh.at[wi, wcol] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=wcol,
                                old_value=wval,
                                new_value=mval,
                                change_type="date_update",
                            )
                        elif pd.notna(mval):
                            wh.at[wi, wcol] = mval
                    else:
                        if ALWAYS_OVERWRITE_NONDATE and pd.notna(mval):
                            if (wval is None) or (str(mval) != str(wval)):
                                stats["updates"] += 1
                                stats["field_updates"] += 1
                                wh.at[wi, wcol] = mval
                                self.change_tracker.add_change(
                                    row_index=wi,
                                    column_name=wcol,
                                    old_value=wval,
                                    new_value=mval,
                                    change_type="field_update",
                                )

        stats["warehouse_duplicate_cases"] = len(dup_cases)
        return wh, stats

    def synchronize(
        self, master_xlsx: str, warehouse_xlsx: str, output_path: Optional[str] = None
    ) -> SyncResult:
        try:
            # Load
            m_xl = pd.ExcelFile(master_xlsx)
            w_xl = pd.ExcelFile(warehouse_xlsx)
            m_df = pd.read_excel(m_xl, sheet_name=m_xl.sheet_names[0])
            w_df = pd.read_excel(w_xl, sheet_name=w_xl.sheet_names[0])

            m_case = self._case_col(m_df)
            w_case = self._case_col(w_df)
            if not (m_case and w_case):
                return SyncResult(
                    False,
                    "CASE NO column not found.",
                    output_path or warehouse_xlsx,
                    {},
                )

            updated_w_df, stats = self._apply_updates(m_df, w_df, m_case, w_case)

            # Save to output
            out = output_path or str(
                Path(warehouse_xlsx).with_name(
                    Path(warehouse_xlsx).stem + ".synced.xlsx"
                )
            )
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                updated_w_df.to_excel(
                    writer, sheet_name=w_xl.sheet_names[0], index=False
                )

            # Colorize
            try:
                from excel_formatter import (
                    ExcelFormatter,
                )  # expects same folder or PYTHONPATH
            except Exception:
                try:
                    from .excel_formatter import (
                        ExcelFormatter,
                    )  # package-style import fallback
                except Exception:
                    # Fallback: create minimal ExcelFormatter
                    class ExcelFormatter:
                        def __init__(
                            self,
                            change_tracker,
                            orange_hex="FFC000",
                            yellow_hex="FFFF00",
                        ):
                            self.ct = change_tracker
                            self.orange = PatternFill(
                                start_color=orange_hex,
                                end_color=orange_hex,
                                fill_type="solid",
                            )
                            self.yellow = PatternFill(
                                start_color=yellow_hex,
                                end_color=yellow_hex,
                                fill_type="solid",
                            )

                        def apply_formatting_inplace(
                            self, excel_file_path, sheet_name, header_row=1
                        ):
                            try:
                                wb = load_workbook(excel_file_path)
                                if sheet_name not in wb.sheetnames:
                                    return False
                                ws = wb[sheet_name]

                                # Build header map
                                header_map = {}
                                for c_idx, cell in enumerate(ws[header_row], start=1):
                                    if cell.value is None:
                                        continue
                                    header_map[str(cell.value).strip()] = c_idx

                                # Apply date changes (ORANGE)
                                for ch in getattr(self.ct, "changes", []) or []:
                                    if (
                                        str(getattr(ch, "change_type", ""))
                                        != "date_update"
                                    ):
                                        continue

                                    row_index = getattr(ch, "row_index", None)
                                    col_name = getattr(ch, "column_name", None)
                                    if row_index is None or col_name is None:
                                        continue

                                    excel_row = int(row_index) + header_row + 1
                                    col_idx = header_map.get(col_name)
                                    if col_idx is None:
                                        # Case-insensitive fallback
                                        norm = (
                                            str(col_name)
                                            .strip()
                                            .lower()
                                            .replace(" ", "_")
                                        )
                                        for k, v in header_map.items():
                                            if norm == str(k).strip().lower().replace(
                                                " ", "_"
                                            ):
                                                col_idx = v
                                                break
                                    if col_idx is None:
                                        continue

                                    ws.cell(row=excel_row, column=col_idx).fill = (
                                        self.orange
                                    )

                                # Apply new records (YELLOW)
                                painted_rows = set()
                                for ch in getattr(self.ct, "changes", []) or []:
                                    if (
                                        str(getattr(ch, "change_type", ""))
                                        == "new_record"
                                    ):
                                        row_index = getattr(ch, "row_index", None)
                                        if row_index is None:
                                            continue
                                        excel_row = int(row_index) + header_row + 1
                                        for c in ws[excel_row]:
                                            c.fill = self.yellow
                                        painted_rows.add(excel_row)

                                wb.save(excel_file_path)
                                return True
                            except Exception:
                                return False

            fmt = ExcelFormatter(
                self.change_tracker, orange_hex=ORANGE, yellow_hex=YELLOW
            )
            fmt.apply_formatting_inplace(
                out, sheet_name=w_xl.sheet_names[0], header_row=1
            )

            stats["output_file"] = out
            return SyncResult(True, "Sync & colorize done.", out, stats)
        except Exception as e:
            return SyncResult(False, f"Error: {e}", output_path or warehouse_xlsx, {})


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--master", required=True)
    ap.add_argument("--warehouse", required=True)
    ap.add_argument("--out", default="")
    args = ap.parse_args()

    sync = DataSynchronizerV29()
    res = sync.synchronize(args.master, args.warehouse, args.out or None)
    print("success:", res.success)
    print("message:", res.message)
    print("output:", res.output_path)
    print("stats:", res.stats)
