#!/usr/bin/env python3
"""
Excel File Structure Analyzer
인보이스 Excel 파일 구조 상세 분석 도구

Version: 1.0.0
Created: 2025-10-14
Author: MACHO-GPT v3.4-mini HVDC Project Enhancement
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Any
import logging

# 로깅 설정
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def analyze_excel_structure(excel_file_path: str) -> Dict[str, Any]:
    """Excel 파일 구조 상세 분석"""

    file_path = Path(excel_file_path)

    if not file_path.exists():
        logger.error(f"File not found: {excel_file_path}")
        return {"error": "File not found"}

    logger.info(f"Analyzing Excel file: {file_path.name}")

    analysis = {
        "file_info": {
            "file_name": file_path.name,
            "file_size_mb": round(file_path.stat().st_size / 1024 / 1024, 2),
            "file_type": file_path.suffix,
        },
        "sheets": {},
        "summary": {},
        "data_samples": {},
    }

    try:
        # openpyxl로 workbook 로드 (VBA 매크로 정보 포함)
        wb = openpyxl.load_workbook(excel_file_path, keep_vba=True, data_only=False)

        logger.info(f"Workbook loaded: {len(wb.sheetnames)} sheets")

        # VBA 매크로 확인
        analysis["vba_macros"] = {
            "has_vba": wb.vba_archive is not None,
            "vba_projects": (len(wb.vba_archive.namelist()) if wb.vba_archive else 0),
        }

        # 각 시트 분석
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 시트 정보
            sheet_info = {
                "sheet_name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "has_data": sheet.max_row > 1,
            }

            # 헤더 추출 (첫 번째 행)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            sheet_info["headers"] = headers
            sheet_info["column_count"] = len([h for h in headers if h is not None])

            # 데이터 행 수
            sheet_info["data_rows"] = sheet.max_row - 1  # 헤더 제외

            analysis["sheets"][sheet_name] = sheet_info

        # pandas로 상세 데이터 분석
        logger.info("Reading with pandas for detailed analysis...")
        df_dict = pd.read_excel(excel_file_path, sheet_name=None, engine="openpyxl")

        total_items = 0
        for sheet_name, df in df_dict.items():
            total_items += len(df)

            # 샘플 데이터 저장 (첫 3행)
            if len(df) > 0:
                sample_data = []
                for idx, row in df.head(3).iterrows():
                    sample_data.append(row.to_dict())

                analysis["data_samples"][sheet_name] = {
                    "row_count": len(df),
                    "columns": list(df.columns),
                    "sample_rows": sample_data,
                    "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                }

        # 전체 요약
        analysis["summary"] = {
            "total_sheets": len(wb.sheetnames),
            "total_items": total_items,
            "average_items_per_sheet": (
                round(total_items / len(wb.sheetnames), 1)
                if len(wb.sheetnames) > 0
                else 0
            ),
            "sheet_names": wb.sheetnames,
        }

        logger.info(
            f"Analysis complete: {total_items} total items across {len(wb.sheetnames)} sheets"
        )

    except Exception as e:
        logger.error(f"Error analyzing file: {e}")
        analysis["error"] = str(e)

    return analysis


def print_analysis_report(analysis: Dict[str, Any]):
    """분석 결과 출력"""

    print("\n" + "=" * 80)
    print("Excel File Structure Analysis")
    print("=" * 80)

    # 파일 정보
    file_info = analysis.get("file_info", {})
    print(f"\n[FILE INFO]")
    print(f"Name: {file_info.get('file_name', 'Unknown')}")
    print(f"Size: {file_info.get('file_size_mb', 0)} MB")
    print(f"Type: {file_info.get('file_type', 'Unknown')}")

    # VBA 정보
    vba_info = analysis.get("vba_macros", {})
    print(f"\n[VBA MACROS]")
    print(f"Has VBA: {vba_info.get('has_vba', False)}")
    print(f"VBA Projects: {vba_info.get('vba_projects', 0)}")

    # 요약
    summary = analysis.get("summary", {})
    print(f"\n[SUMMARY]")
    print(f"Total Sheets: {summary.get('total_sheets', 0)}")
    print(f"Total Items: {summary.get('total_items', 0)}")
    print(f"Average Items/Sheet: {summary.get('average_items_per_sheet', 0)}")

    # 시트별 상세
    print(f"\n[SHEET DETAILS]")
    sheets = analysis.get("sheets", {})
    for i, (sheet_name, sheet_info) in enumerate(list(sheets.items())[:10], 1):
        print(f"\n{i}. {sheet_name}")
        print(f"   Rows: {sheet_info.get('data_rows', 0)}")
        print(f"   Columns: {sheet_info.get('column_count', 0)}")
        headers = sheet_info.get("headers", [])
        if headers:
            non_null_headers = [h for h in headers if h is not None]
            print(f"   Headers: {', '.join(map(str, non_null_headers[:5]))}")

    if len(sheets) > 10:
        print(f"\n   ... and {len(sheets) - 10} more sheets")

    # 데이터 샘플
    data_samples = analysis.get("data_samples", {})
    if data_samples:
        print(f"\n[DATA SAMPLE - First Sheet]")
        first_sheet = list(data_samples.keys())[0]
        sample = data_samples[first_sheet]
        print(f"Sheet: {first_sheet}")
        print(f"Columns: {len(sample.get('columns', []))}")
        print(f"Sample columns: {', '.join(sample.get('columns', [])[:8])}")

    print("\n" + "=" * 80)


def main():
    """메인 실행 함수"""

    # 파일 경로
    excel_file = Path(
        rPath(__file__).parent.parent / "Data" / "DSV 202509" / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    )

    # 구조 분석
    analysis = analyze_excel_structure(str(excel_file))

    # 결과 출력
    print_analysis_report(analysis)

    # JSON 저장
    import json
    from datetime import datetime

    output_dir = Path(__file__).parent / "out"
    output_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = output_dir / f"excel_structure_analysis_{timestamp}.json"

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(analysis, f, indent=2, ensure_ascii=False, default=str)

    logger.info(f"Analysis saved to: {json_path}")

    return analysis


if __name__ == "__main__":
    main()
