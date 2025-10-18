#!/usr/bin/env python3
"""
Excel 구조 비교 검증
원본 INVOICE vs 생성된 검증 결과 비교
"""

import pandas as pd
import openpyxl
from pathlib import Path


def main():
    print("=" * 80)
    print("Excel Structure Comparison")
    print("=" * 80)

    # 1. 원본 INVOICE 파일의 MasterData
    print("\n[1. Original INVOICE - MasterData Sheet]")
    invoice_file = Path(
        rPath(__file__).parent.parent / "Data" / "DSV 202509" / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    )

    wb_invoice = openpyxl.load_workbook(invoice_file, data_only=False)
    ws_invoice = wb_invoice["MasterData"]

    print(f"File: {invoice_file.name}")
    print(f"Rows: {ws_invoice.max_row}")
    print(f"Columns: {ws_invoice.max_column}")

    print(f"\nColumn Structure:")
    for i, cell in enumerate(ws_invoice[1], 1):
        if cell.value:
            col_letter = cell.column_letter
            print(f"  {i:2d}. {col_letter:2s} | {cell.value}")

    # 2. 생성된 검증 결과 Excel
    print(f"\n[2. Generated Validation Report]")
    result_file = Path(
        r"C:\Users\minky\Downloads\HVDC_Invoice_Audit-20251012T195441Z-1-001\HVDC_Invoice_Audit\01_DSV_SHPT\Core_Systems\out\masterdata_validated_20251014_200508.xlsx"
    )

    if result_file.exists():
        df_result = pd.read_excel(result_file)
        print(f"File: {result_file.name}")
        print(f"Rows: {len(df_result)}")
        print(f"Columns: {len(df_result.columns)}")

        print(f"\nColumn Structure:")
        for i, col in enumerate(df_result.columns, 1):
            print(f"  {i:2d}. {col}")

        print(f"\n[3. Column Mapping]")
        print(f"\n[VBA Original - Columns 1-13]")
        for i, col in enumerate(df_result.columns[:13], 1):
            print(f"  {i:2d}. {col}")

        print(f"\n[Python Added - Columns 14-22]")
        for i, col in enumerate(df_result.columns[13:], 14):
            print(f"  {i:2d}. {col}")

        # 샘플 데이터 확인
        print(f"\n[4. Sample Row Verification]")
        print(f"\nRow 1 (MASTER DO FEE - should be PASS):")
        row1 = df_result.iloc[0]
        print(f"  DESCRIPTION: {row1['DESCRIPTION']}")
        print(f"  RATE: ${row1['RATE']}")
        print(f"  REV RATE (VBA): ${row1['REV RATE']}")
        print(f"  DIFFERENCE (VBA): ${row1['DIFFERENCE']}")
        print(f"  ---")
        print(f"  Ref_Rate_USD (Python): ${row1['Ref_Rate_USD']}")
        print(f"  Python_Delta: {row1['Python_Delta']}%")
        print(f"  CG_Band: {row1['CG_Band']}")
        print(f"  Validation_Status: {row1['Validation_Status']}")
        print(f"  Gate_Score: {row1['Gate_Score']}")

    else:
        print(f"[ERROR] Result file not found: {result_file}")

    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
