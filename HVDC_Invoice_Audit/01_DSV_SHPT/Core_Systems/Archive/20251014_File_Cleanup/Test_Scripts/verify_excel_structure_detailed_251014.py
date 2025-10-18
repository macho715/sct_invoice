#!/usr/bin/env python3
"""
Excel 구조 상세 검증 스크립트
MasterData와 개별 시트 구조를 상세 비교 분석

Version: 1.0.0
Created: 2025-10-14
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List


def verify_excel_structure():
    """Excel 파일 구조 상세 검증"""

    excel_file = Path(
        rPath(__file__).parent.parent / "Data" / "DSV 202509" / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    )

    print("=" * 80)
    print("Excel Structure Verification")
    print("=" * 80)

    # 1. MasterData 시트 검증
    print("\n[1. MasterData Sheet - VBA Output]")
    df_master = pd.read_excel(excel_file, sheet_name="MasterData")

    print(f"Rows: {len(df_master)}")
    print(f"Columns: {len(df_master.columns)}")
    print(f"\nColumn Names and Positions:")
    for i, col in enumerate(df_master.columns, 1):
        print(f"  {i:2d}. {col}")

    print(f"\nSample Data (Row 1):")
    row1 = df_master.iloc[0]
    for col, val in row1.items():
        print(f"  {col}: {val}")

    # 2. 개별 시트 검증 (SCT0126 예시)
    print(f"\n[2. Individual Sheet - SCT0126]")

    wb = openpyxl.load_workbook(excel_file, data_only=False)
    ws_sct = wb["SCT0126"]

    print(f"Max Row: {ws_sct.max_row}")
    print(f"Max Column: {ws_sct.max_column}")

    # 헤더 찾기
    print(f"\nSearching for 'S/No' header...")
    header_row = None
    for row_idx in range(1, min(20, ws_sct.max_row + 1)):
        for cell in ws_sct[row_idx]:
            if cell.value and "S/No" in str(cell.value):
                header_row = row_idx
                print(f"Found at Row {row_idx}, Column {cell.column_letter}")
                break
        if header_row:
            break

    if header_row:
        print(f"\nHeaders at Row {header_row}:")
        headers = []
        for i, cell in enumerate(ws_sct[header_row], 1):
            if cell.value:
                headers.append((i, cell.column_letter, cell.value))
                if i <= 15:
                    print(f"  {cell.column_letter:2s} ({i:2d}): {cell.value}")

        # 데이터 샘플
        print(f"\nData Sample (Rows {header_row+1} to {header_row+3}):")
        for row_idx in range(header_row + 1, min(header_row + 4, ws_sct.max_row + 1)):
            values = []
            for col_idx in range(1, min(len(headers) + 1, 11)):
                val = ws_sct.cell(row_idx, col_idx).value
                values.append(str(val)[:20] if val else "")
            print(f"  Row {row_idx}: {values[:5]}")

    # 3. MasterData vs 개별 시트 비교
    print(f"\n[3. MasterData vs Individual Sheets]")
    print(f"\nMasterData columns: {list(df_master.columns)}")
    print(f"\nIndividual sheet columns (from VBA):")
    print(f"  - Headers start at variable row (has 'S/No')")
    print(
        f"  - Columns include: S/No, RATE SOURCE, DESCRIPTION, RATE, Q'TY, TOTAL, etc."
    )

    # 4. 최종 보고서 구조 제안
    print(f"\n[4. Proposed Final Report Structure]")
    print(f"\nSheet: MasterData_Validated")
    print(f"\nColumns (22 total):")
    print(f"\n[VBA Original - 13 columns]")
    for i, col in enumerate(df_master.columns, 1):
        print(f"  {i:2d}. {col}")

    print(f"\n[Python Validation - 9 columns]")
    validation_cols = [
        "Validation_Status",
        "Ref_Rate_USD",
        "Python_Delta",
        "CG_Band",
        "Charge_Group",
        "Gate_Score",
        "Gate_Status",
        "PDF_Count",
        "Validation_Notes",
    ]
    for i, col in enumerate(validation_cols, 14):
        print(f"  {i:2d}. {col}")

    # 5. 컬럼 매핑 확인
    print(f"\n[5. Column Mapping for Validation]")
    print(f"\nKey columns for validation:")
    print(f"  - RATE SOURCE → Charge_Group 결정")
    print(f"  - DESCRIPTION → Contract rate lookup")
    print(f"  - RATE → Draft rate (검증 대상)")
    print(f"  - REV RATE → VBA calculated rate")
    print(f"  - DIFFERENCE → VBA delta (금액)")
    print(f"\nPython validation will add:")
    print(f"  - Ref_Rate_USD → Configuration lookup")
    print(f"  - Python_Delta → (RATE - Ref_Rate) / Ref_Rate * 100 (백분율)")
    print(f"  - CG_Band → COST-GUARD band classification")

    print("\n" + "=" * 80)
    print("[VERIFICATION COMPLETE]")
    print("=" * 80)


if __name__ == "__main__":
    verify_excel_structure()
