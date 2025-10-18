#!/usr/bin/env python3
"""
VBA-Integrated Excel Report Generator
Combines VBA-processed data with Python audit results for comprehensive reporting

Version: 1.0.0
Created: 2025-10-13
Author: MACHO-GPT v3.4-mini HVDC Project Enhancement
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

from vba_excel_analyzer import VBAExcelAnalyzer

# 로깅 설정
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class VBAIntegratedExcelReportGenerator:
    """
    VBA 통합 Excel 보고서 생성기

    Features:
    - VBA 처리 결과와 Python 감사 결과 통합
    - Formula 추출 결과 분석 시트
    - REV RATE 계산 검증 시트
    - MasterData 컴파일 결과 시트
    - Executive Dashboard
    - 데이터 검증 및 품질 보고서
    """

    def __init__(self):
        """초기화"""
        self.workbook = None
        self.vba_results = None
        self.python_audit_data = None
        self.integration_results = {}

        # 스타일 정의
        self.styles = self._define_styles()

    def generate_comprehensive_report(
        self,
        vba_excel_path: str,
        python_csv_path: str,
        python_json_path: str,
        output_dir: str = "Results/Sept_2025/Reports",
    ) -> Dict[str, str]:
        """
        종합 VBA 통합 보고서 생성

        Args:
            vba_excel_path: VBA 처리된 Excel 파일 경로
            python_csv_path: Python 감사 CSV 파일 경로
            python_json_path: Python 감사 JSON 파일 경로
            output_dir: 출력 디렉토리

        Returns:
            Dict: 생성된 보고서 파일 경로들
        """
        logger.info("🚀 VBA 통합 Excel 보고서 생성 시작")

        try:
            # 1. VBA 분석 결과 로드
            self._load_vba_analysis(vba_excel_path)

            # 2. Python 감사 결과 로드
            self._load_python_audit_data(python_csv_path, python_json_path)

            # 3. 데이터 통합 및 검증
            self._integrate_and_validate_data()

            # 4. Excel 보고서 생성
            report_path = self._generate_excel_report(output_dir)

            logger.info("✅ VBA 통합 Excel 보고서 생성 완료")
            return {"vba_integrated_report": report_path}

        except Exception as e:
            logger.error(f"❌ VBA 통합 보고서 생성 실패: {str(e)}")
            return {"error": str(e)}

    def _load_vba_analysis(self, vba_excel_path: str):
        """VBA 분석 결과 로드"""
        logger.info("📊 VBA 분석 결과 로드 중...")

        try:
            analyzer = VBAExcelAnalyzer(vba_excel_path)
            self.vba_results = analyzer.analyze_vba_file()

            logger.info(
                f"  ✅ VBA 시트: {self.vba_results['summary']['total_sheets_analyzed']}개"
            )
            logger.info(
                f"  ✅ Formula: {self.vba_results['summary']['formulas_extracted']}개"
            )
            logger.info(
                f"  ✅ 계산: {self.vba_results['summary']['calculations_analyzed']}개"
            )

        except Exception as e:
            logger.error(f"❌ VBA 분석 로드 실패: {e}")
            raise

    def _load_python_audit_data(self, csv_path: str, json_path: str):
        """Python 감사 결과 로드"""
        logger.info("📋 Python 감사 결과 로드 중...")

        try:
            # CSV 데이터 로드
            self.python_audit_data = {
                "csv_data": pd.read_csv(csv_path),
                "json_data": None,
            }

            # JSON 데이터 로드 (있는 경우)
            if Path(json_path).exists():
                with open(json_path, "r", encoding="utf-8") as f:
                    self.python_audit_data["json_data"] = json.load(f)

            logger.info(f"  ✅ CSV 항목: {len(self.python_audit_data['csv_data'])}개")

        except Exception as e:
            logger.error(f"❌ Python 감사 데이터 로드 실패: {e}")
            raise

    def _integrate_and_validate_data(self):
        """데이터 통합 및 검증"""
        logger.info("🔗 데이터 통합 및 검증 중...")

        self.integration_results = {
            "formula_integration": self._integrate_formula_data(),
            "rev_rate_integration": self._integrate_rev_rate_data(),
            "master_data_integration": self._integrate_master_data(),
            "validation_summary": self._create_validation_summary(),
        }

    def _integrate_formula_data(self) -> Dict[str, Any]:
        """Formula 데이터 통합"""
        logger.info("  📋 Formula 데이터 통합...")

        vba_formulas = self.vba_results["vba_results"]["formula_extraction"][
            "formula_data"
        ]

        # Formula 데이터프레임 생성
        formula_df_data = []
        for formula_item in vba_formulas:
            formula_df_data.append(
                {
                    "Sheet": formula_item["sheet"],
                    "Row": formula_item["row"],
                    "Formula": formula_item["formula"],
                    "Formula_Type": self._classify_formula(formula_item["formula"]),
                    "Validation_Status": (
                        "VALID" if formula_item["formula"].strip() else "INVALID"
                    ),
                }
            )

        formula_df = pd.DataFrame(formula_df_data)

        # 통계 생성
        formula_stats = {
            "total_formulas": len(formula_df),
            "valid_formulas": len(
                formula_df[formula_df["Validation_Status"] == "VALID"]
            ),
            "formula_types": formula_df["Formula_Type"].value_counts().to_dict(),
            "sheets_with_formulas": formula_df["Sheet"].nunique(),
        }

        return {"formula_dataframe": formula_df, "formula_statistics": formula_stats}

    def _classify_formula(self, formula: str) -> str:
        """Formula 분류"""
        formula_upper = formula.upper()

        if "VLOOKUP" in formula_upper:
            return "VLOOKUP"
        elif "SUMIF" in formula_upper:
            return "SUMIF"
        elif "IF" in formula_upper:
            return "IF"
        elif "=" in formula_upper and any(
            op in formula_upper for op in ["+", "-", "*", "/"]
        ):
            return "CALCULATION"
        else:
            return "OTHER"

    def _integrate_rev_rate_data(self) -> Dict[str, Any]:
        """REV RATE 데이터 통합"""
        logger.info("  💰 REV RATE 데이터 통합...")

        vba_calculations = self.vba_results["vba_results"]["rev_rate_calculations"][
            "calculations"
        ]

        # REV RATE 데이터프레임 생성
        rev_rate_df_data = []
        for calc in vba_calculations:
            rev_rate_df_data.append(
                {
                    "Sheet": calc.get("sheet", ""),
                    "Row": calc.get("row", 0),
                    "Rate": calc.get("rate"),
                    "Quantity": calc.get("qty", 1),
                    "REV_RATE": calc.get("rev_rate"),
                    "REV_TOTAL": calc.get("rev_total"),
                    "TOTAL_USD": calc.get("total"),
                    "DIFFERENCE": calc.get("difference"),
                    "Validation_Status": self._validate_calculation(calc),
                }
            )

        rev_rate_df = pd.DataFrame(rev_rate_df_data)

        # 통계 생성
        validation_summary = self.vba_results["vba_results"]["rev_rate_calculations"][
            "validation_summary"
        ]

        return {
            "rev_rate_dataframe": rev_rate_df,
            "validation_summary": validation_summary,
            "calculation_accuracy": validation_summary.get("accuracy_rate", 0),
        }

    def _validate_calculation(self, calc: Dict) -> str:
        """개별 계산 검증"""
        try:
            rate = calc.get("rate")
            qty = calc.get("qty", 1)
            rev_rate = calc.get("rev_rate")
            rev_total = calc.get("rev_total")
            total = calc.get("total")
            difference = calc.get("difference")

            errors = []

            # REV RATE 검증
            if rate is not None and rev_rate is not None:
                expected_rev_rate = round(float(rate), 2)
                if abs(float(rev_rate) - expected_rev_rate) > 0.01:
                    errors.append("REV_RATE")

            # REV TOTAL 검증
            if rev_rate is not None and qty is not None and rev_total is not None:
                expected_rev_total = round(float(rev_rate) * float(qty), 2)
                if abs(float(rev_total) - expected_rev_total) > 0.01:
                    errors.append("REV_TOTAL")

            # DIFFERENCE 검증
            if rev_total is not None and total is not None and difference is not None:
                expected_difference = round(float(rev_total) - float(total), 2)
                if abs(float(difference) - expected_difference) > 0.01:
                    errors.append("DIFFERENCE")

            if errors:
                return f"ERROR: {', '.join(errors)}"
            else:
                return "VALID"

        except Exception:
            return "VALIDATION_ERROR"

    def _integrate_master_data(self) -> Dict[str, Any]:
        """MasterData 통합"""
        logger.info("  📊 MasterData 통합...")

        master_data = self.vba_results["vba_results"]["master_data"]
        python_csv = self.python_audit_data["csv_data"]

        # 통합 통계
        integration_stats = {
            "vba_master_rows": master_data["row_count"],
            "python_audit_rows": len(python_csv),
            "data_match_analysis": self._analyze_data_matching(),
        }

        return {
            "integration_statistics": integration_stats,
            "master_data_info": master_data,
        }

    def _analyze_data_matching(self) -> Dict[str, Any]:
        """데이터 매칭 분석"""
        vba_master_rows = self.vba_results["vba_results"]["master_data"]["row_count"]
        python_rows = len(self.python_audit_data["csv_data"])

        return {
            "vba_rows": vba_master_rows,
            "python_rows": python_rows,
            "match_rate": (
                min(vba_master_rows, python_rows) / max(vba_master_rows, python_rows)
                if max(vba_master_rows, python_rows) > 0
                else 0
            ),
            "data_consistency": (
                "HIGH"
                if abs(vba_master_rows - python_rows) <= 5
                else "MEDIUM" if abs(vba_master_rows - python_rows) <= 20 else "LOW"
            ),
        }

    def _create_validation_summary(self) -> Dict[str, Any]:
        """검증 요약 생성"""
        vba_validation = self.vba_results["validation_results"]

        return {
            "overall_vba_score": vba_validation["overall_score"],
            "formula_validation": vba_validation["formula_validation"],
            "calculation_accuracy": vba_validation["calculation_validation"][
                "accuracy_rate"
            ],
            "data_integrity_score": vba_validation["data_integrity"]["pass_rate"],
            "recommendations": self._generate_recommendations(),
        }

    def _generate_recommendations(self) -> List[str]:
        """개선 권장사항 생성"""
        recommendations = []

        vba_score = self.vba_results["validation_results"]["overall_score"]

        if vba_score < 0.7:
            recommendations.append("VBA 계산 정확도 개선 필요")

        calc_accuracy = self.vba_results["validation_results"][
            "calculation_validation"
        ]["accuracy_rate"]
        if calc_accuracy < 0.9:
            recommendations.append("REV RATE 계산 로직 검토 필요")

        formula_count = self.vba_results["summary"]["formulas_extracted"]
        if formula_count == 0:
            recommendations.append("Formula 추출 프로세스 점검 필요")

        return recommendations

    def _generate_excel_report(self, output_dir: str) -> str:
        """Excel 보고서 생성"""
        logger.info("📊 Excel 보고서 생성 중...")

        # 워크북 생성
        self.workbook = Workbook()

        # 기본 시트 제거
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # 시트 생성
        self._create_executive_dashboard()
        self._create_formula_analysis_sheet()
        self._create_rev_rate_analysis_sheet()
        self._create_master_data_sheet()
        self._create_validation_summary_sheet()
        self._create_integration_comparison_sheet()
        self._create_vba_processing_log()

        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(output_dir) / f"VBA_INTEGRATED_REPORT_{timestamp}.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self.workbook.save(output_path)

        logger.info(f"✅ Excel 보고서 저장: {output_path}")
        logger.info(f"📋 총 시트 수: {len(self.workbook.sheetnames)}")

        return str(output_path)

    def _create_executive_dashboard(self):
        """Executive Dashboard 시트 생성"""
        logger.info("  📊 Executive Dashboard 생성...")

        ws = self.workbook.create_sheet("Executive_Dashboard")

        # 헤더
        ws.merge_cells("A1:F1")
        ws["A1"] = "VBA-Integrated Invoice Audit Report - Executive Dashboard"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 기본 정보
        row = 3
        ws[f"A{row}"] = "Report Generation Date:"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row += 2

        # VBA 분석 요약
        ws[f"A{row}"] = "VBA Analysis Summary"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        summary = self.vba_results["summary"]

        dashboard_data = [
            ["Total Sheets Analyzed", summary["total_sheets_analyzed"]],
            ["Formulas Extracted", summary["formulas_extracted"]],
            ["Calculations Analyzed", summary["calculations_analyzed"]],
            ["MasterData Rows", summary["master_data_rows"]],
            ["Overall Validation Score", f"{summary['overall_validation_score']:.2%}"],
            ["Validation Status", summary["validation_status"]],
        ]

        for item, value in dashboard_data:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            row += 1

        row += 2

        # 통합 결과 요약
        ws[f"A{row}"] = "Integration Results Summary"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        integration_data = [
            [
                "Formula Integration",
                f"{len(self.integration_results['formula_integration']['formula_dataframe'])} formulas processed",
            ],
            [
                "REV RATE Validation",
                f"{self.integration_results['rev_rate_integration']['calculation_accuracy']:.2%} accuracy",
            ],
            [
                "Data Consistency",
                self.integration_results["master_data_integration"][
                    "integration_statistics"
                ]["data_match_analysis"]["data_consistency"],
            ],
            ["Python Audit Items", len(self.python_audit_data["csv_data"])],
        ]

        for item, value in integration_data:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            row += 1

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _create_formula_analysis_sheet(self):
        """Formula Analysis 시트 생성"""
        logger.info("  📋 Formula Analysis 시트 생성...")

        ws = self.workbook.create_sheet("Formula_Analysis")

        # 헤더
        ws["A1"] = "VBA Formula Extraction Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        # Formula 데이터
        formula_df = self.integration_results["formula_integration"][
            "formula_dataframe"
        ]

        if not formula_df.empty:
            # 데이터 추가
            for r in dataframe_to_rows(formula_df, index=False, header=True):
                ws.append(r)

            # 헤더 스타일 적용
            for col in range(1, len(formula_df.columns) + 1):
                ws.cell(row=3, column=col).font = Font(bold=True)
                ws.cell(row=3, column=col).fill = PatternFill(
                    start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
                )

        # 통계 추가
        stats = self.integration_results["formula_integration"]["formula_statistics"]

        stats_row = len(formula_df) + 6 if not formula_df.empty else 5

        ws[f"A{stats_row}"] = "Formula Statistics"
        ws[f"A{stats_row}"].font = Font(bold=True, size=12)

        stats_data = [
            ["Total Formulas", stats["total_formulas"]],
            ["Valid Formulas", stats["valid_formulas"]],
            ["Sheets with Formulas", stats["sheets_with_formulas"]],
        ]

        for i, (item, value) in enumerate(stats_data, 1):
            ws[f"A{stats_row + i}"] = item
            ws[f"B{stats_row + i}"] = value

        # 컬럼 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _create_rev_rate_analysis_sheet(self):
        """REV RATE Analysis 시트 생성"""
        logger.info("  💰 REV RATE Analysis 시트 생성...")

        ws = self.workbook.create_sheet("REV_RATE_Analysis")

        # 헤더
        ws["A1"] = "VBA REV RATE Calculation Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        # REV RATE 데이터
        rev_rate_df = self.integration_results["rev_rate_integration"][
            "rev_rate_dataframe"
        ]

        if not rev_rate_df.empty:
            # 데이터 추가 (상위 100행만)
            display_df = rev_rate_df.head(100)

            for r in dataframe_to_rows(display_df, index=False, header=True):
                ws.append(r)

            # 헤더 스타일 적용
            for col in range(1, len(display_df.columns) + 1):
                ws.cell(row=3, column=col).font = Font(bold=True)
                ws.cell(row=3, column=col).fill = PatternFill(
                    start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                )

            # 검증 상태에 따른 조건부 서식
            for row in range(4, min(104, len(display_df) + 4)):
                validation_cell = ws.cell(row=row, column=len(display_df.columns))
                if "ERROR" in str(validation_cell.value):
                    validation_cell.fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )
                elif "VALID" in str(validation_cell.value):
                    validation_cell.fill = PatternFill(
                        start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                    )

        # 검증 요약
        validation_summary = self.integration_results["rev_rate_integration"][
            "validation_summary"
        ]

        summary_row = min(104, len(rev_rate_df) + 6) if not rev_rate_df.empty else 5

        ws[f"A{summary_row}"] = "Validation Summary"
        ws[f"A{summary_row}"].font = Font(bold=True, size=12)

        summary_data = [
            ["Total Calculations", validation_summary["total_items"]],
            ["Valid Calculations", validation_summary["valid_calculations"]],
            ["Accuracy Rate", f"{validation_summary['accuracy_rate']:.2%}"],
            ["Calculation Errors", len(validation_summary["calculation_errors"])],
        ]

        for i, (item, value) in enumerate(summary_data, 1):
            ws[f"A{summary_row + i}"] = item
            ws[f"B{summary_row + i}"] = value

        # 컬럼 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width

    def _create_master_data_sheet(self):
        """MasterData 시트 생성"""
        logger.info("  📊 MasterData 시트 생성...")

        ws = self.workbook.create_sheet("MasterData_Analysis")

        # 헤더
        ws["A1"] = "VBA MasterData Compilation Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        # MasterData 정보
        master_data = self.vba_results["vba_results"]["master_data"]

        row = 3
        ws[f"A{row}"] = "MasterData Sheet Analysis"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        master_info = [
            ["Sheet Found", "Yes" if master_data["sheet_found"] else "No"],
            ["Total Rows", master_data["row_count"]],
            ["Column Count", len(master_data["column_mapping"])],
        ]

        for item, value in master_info:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            row += 1

        # 컬럼 매핑
        if master_data["column_mapping"]:
            row += 2
            ws[f"A{row}"] = "Column Mapping"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            ws[f"A{row}"] = "Column Index"
            ws[f"B{row}"] = "Column Name"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
            row += 1

            for col_idx, col_name in master_data["column_mapping"].items():
                ws[f"A{row}"] = col_idx
                ws[f"B{row}"] = col_name
                row += 1

        # 데이터 품질 요약
        if master_data["data_summary"]:
            row += 2
            ws[f"A{row}"] = "Data Quality Summary"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            ws[f"A{row}"] = "Column Name"
            ws[f"B{row}"] = "Non-Null Count"
            ws[f"C{row}"] = "Unique Count"

            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].font = Font(bold=True)
            row += 1

            for col_name, stats in master_data["data_summary"].items():
                ws[f"A{row}"] = col_name
                ws[f"B{row}"] = stats["non_null_count"]
                ws[f"C{row}"] = stats["unique_count"]
                row += 1

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15

    def _create_validation_summary_sheet(self):
        """Validation Summary 시트 생성"""
        logger.info("  ✅ Validation Summary 시트 생성...")

        ws = self.workbook.create_sheet("Validation_Summary")

        # 헤더
        ws["A1"] = "VBA Processing Validation Summary"
        ws["A1"].font = Font(size=14, bold=True)

        # 전체 검증 결과
        validation = self.integration_results["validation_summary"]

        row = 3
        ws[f"A{row}"] = "Overall Validation Results"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        overall_data = [
            ["Overall VBA Score", f"{validation['overall_vba_score']:.2%}"],
            [
                "Formula Validation",
                f"{validation['formula_validation']['valid_formulas']}/{validation['formula_validation']['total_formulas']} valid",
            ],
            ["Calculation Accuracy", f"{validation['calculation_accuracy']:.2%}"],
            ["Data Integrity Score", f"{validation['data_integrity_score']:.2%}"],
        ]

        for item, value in overall_data:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value

            # 점수에 따른 색상 적용
            if "%" in str(value):
                try:
                    percentage = float(str(value).replace("%", "")) / 100
                    if percentage >= 0.9:
                        ws[f"B{row}"].fill = PatternFill(
                            start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                        )
                    elif percentage >= 0.7:
                        ws[f"B{row}"].fill = PatternFill(
                            start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"
                        )
                    else:
                        ws[f"B{row}"].fill = PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                        )
                except:
                    pass

            row += 1

        # 권장사항
        if validation["recommendations"]:
            row += 2
            ws[f"A{row}"] = "Recommendations"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            for i, recommendation in enumerate(validation["recommendations"], 1):
                ws[f"A{row}"] = f"{i}. {recommendation}"
                row += 1

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _create_integration_comparison_sheet(self):
        """Integration Comparison 시트 생성"""
        logger.info("  🔗 Integration Comparison 시트 생성...")

        ws = self.workbook.create_sheet("Integration_Comparison")

        # 헤더
        ws["A1"] = "VBA vs Python Audit Data Comparison"
        ws["A1"].font = Font(size=14, bold=True)

        # 데이터 비교
        integration_stats = self.integration_results["master_data_integration"][
            "integration_statistics"
        ]
        match_analysis = integration_stats["data_match_analysis"]

        row = 3
        ws[f"A{row}"] = "Data Volume Comparison"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        comparison_data = [
            ["VBA MasterData Rows", match_analysis["vba_rows"]],
            ["Python Audit Items", match_analysis["python_rows"]],
            ["Data Match Rate", f"{match_analysis['match_rate']:.2%}"],
            ["Data Consistency Level", match_analysis["data_consistency"]],
        ]

        for item, value in comparison_data:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            row += 1

        # Python 감사 데이터 샘플 (상위 20행)
        row += 2
        ws[f"A{row}"] = "Python Audit Data Sample (Top 20 Items)"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        python_sample = self.python_audit_data["csv_data"].head(20)

        # 주요 컬럼만 선택
        key_columns = [
            "s_no",
            "sheet_name",
            "description",
            "unit_rate",
            "quantity",
            "total_usd",
            "status",
        ]
        available_columns = [col for col in key_columns if col in python_sample.columns]

        if available_columns:
            sample_data = python_sample[available_columns]

            # 헤더 추가
            for col_idx, col_name in enumerate(available_columns):
                ws.cell(row=row, column=col_idx + 1).value = col_name
                ws.cell(row=row, column=col_idx + 1).font = Font(bold=True)
            row += 1

            # 데이터 추가
            for _, data_row in sample_data.iterrows():
                for col_idx, col_name in enumerate(available_columns):
                    ws.cell(row=row, column=col_idx + 1).value = data_row[col_name]
                row += 1

        # 컬럼 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column].width = adjusted_width

    def _create_vba_processing_log(self):
        """VBA Processing Log 시트 생성"""
        logger.info("  📝 VBA Processing Log 시트 생성...")

        ws = self.workbook.create_sheet("VBA_Processing_Log")

        # 헤더
        ws["A1"] = "VBA Processing Audit Log"
        ws["A1"].font = Font(size=14, bold=True)

        # 처리 로그 정보
        file_info = self.vba_results["file_info"]

        row = 3
        ws[f"A{row}"] = "File Processing Information"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        file_data = [
            ["Source File", file_info["file_path"]],
            ["File Size (bytes)", file_info["file_size"]],
            ["Analysis Timestamp", file_info["analysis_timestamp"]],
            ["Total Sheets Processed", file_info["total_sheets"]],
        ]

        for item, value in file_data:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = value
            row += 1

        # 시트별 처리 결과
        row += 2
        ws[f"A{row}"] = "Sheet Processing Results"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # 헤더
        headers = ["Sheet Name", "Has Headers", "Data Rows", "Completeness", "Status"]
        for col_idx, header in enumerate(headers):
            ws.cell(row=row, column=col_idx + 1).value = header
            ws.cell(row=row, column=col_idx + 1).font = Font(bold=True)
        row += 1

        # 시트별 데이터
        sheet_analysis = self.vba_results["vba_results"]["sheet_analysis"]

        for sheet_name, analysis in sheet_analysis.items():
            ws.cell(row=row, column=1).value = sheet_name
            ws.cell(row=row, column=2).value = (
                "Yes" if analysis["has_headers"] else "No"
            )
            ws.cell(row=row, column=3).value = analysis["data_quality"].get(
                "non_empty_rows", 0
            )
            ws.cell(row=row, column=4).value = (
                f"{analysis['data_quality'].get('completeness', 0):.2%}"
            )

            # 상태 결정
            completeness = analysis["data_quality"].get("completeness", 0)
            if completeness >= 0.9:
                status = "EXCELLENT"
                ws.cell(row=row, column=5).fill = PatternFill(
                    start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                )
            elif completeness >= 0.7:
                status = "GOOD"
                ws.cell(row=row, column=5).fill = PatternFill(
                    start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"
                )
            else:
                status = "NEEDS_REVIEW"
                ws.cell(row=row, column=5).fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )

            ws.cell(row=row, column=5).value = status
            row += 1

        # 컬럼 너비 조정
        column_widths = [20, 12, 12, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

    def _define_styles(self) -> Dict[str, Any]:
        """스타일 정의"""
        return {
            "header": Font(bold=True, size=12),
            "title": Font(bold=True, size=14),
            "success": PatternFill(
                start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
            ),
            "warning": PatternFill(
                start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"
            ),
            "error": PatternFill(
                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
            ),
            "info": PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
            ),
        }


def main():
    """메인 실행 함수"""

    # 파일 경로 설정
    vba_excel_path = (
        "Data/DSV 202509/SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    )
    python_csv_path = (
        "Results/Sept_2025/shpt_sept_2025_enhanced_result_20251012_123701.csv"
    )
    python_json_path = (
        "Results/Sept_2025/shpt_sept_2025_enhanced_result_20251012_123701.json"
    )

    try:
        generator = VBAIntegratedExcelReportGenerator()
        results = generator.generate_comprehensive_report(
            vba_excel_path=vba_excel_path,
            python_csv_path=python_csv_path,
            python_json_path=python_json_path,
        )

        if "error" in results:
            print(f"❌ 보고서 생성 실패: {results['error']}")
        else:
            print("=" * 80)
            print("🎉 VBA 통합 Excel 보고서 생성 완료")
            print("=" * 80)
            print(f"📊 보고서 파일: {results['vba_integrated_report']}")

        return results

    except Exception as e:
        logger.error(f"❌ 메인 실행 실패: {e}")
        return {"error": str(e)}


if __name__ == "__main__":
    main()
