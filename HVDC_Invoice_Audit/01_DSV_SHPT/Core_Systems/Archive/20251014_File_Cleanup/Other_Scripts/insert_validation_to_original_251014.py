#!/usr/bin/env python3
"""
원본 INVOICE에 Python 검증 결과 삽입
_FINAL.xlsm 복사 후 MasterData 시트의 14-22번 컬럼에 검증 결과 삽입

Version: 1.0.0
Created: 2025-10-14
Author: MACHO-GPT v3.4-mini HVDC Project Enhancement
"""

import shutil
import sys
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
import logging

# Configuration Manager import
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "00_Shared"))
from config_manager import ConfigurationManager

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class ValidationInserter:
    """원본 Excel에 검증 결과 삽입"""

    def __init__(self):
        self.root = Path(__file__).parent.parent
        self.rate_dir = self.root.parent / "Rate"

        # Configuration Manager
        self.config_manager = ConfigurationManager(self.rate_dir)
        self.config_manager.load_all_configs()

        # 파일 경로
        self.original_file = (
            self.root
            / "Data"
            / "DSV 202509"
            / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_file = (
            self.root
            / "Results"
            / "Sept_2025"
            / f"SCNT_SHIPMENT_SEPT2025_VALIDATED_{timestamp}.xlsx"
        )

    def copy_original_file(self):
        """원본 파일 복사"""

        logger.info(f"Copying original file...")
        logger.info(f"  From: {self.original_file.name}")
        logger.info(f"  To: {self.output_file.name}")

        # xlsm을 xlsx로 복사 (openpyxl 호환성)
        shutil.copy2(self.original_file, self.output_file)

        logger.info(f"[OK] File copied")

    def load_validation_results(self) -> pd.DataFrame:
        """이전에 생성된 검증 결과 로드"""

        # 최신 검증 결과 CSV 로드
        out_dir = Path(__file__).parent / "out"
        csv_files = sorted(out_dir.glob("masterdata_validated_*.csv"))

        if not csv_files:
            logger.error("No validation results found!")
            return None

        latest_csv = csv_files[-1]
        logger.info(f"Loading validation results from: {latest_csv.name}")

        df = pd.read_csv(latest_csv)
        logger.info(f"  Loaded: {len(df)} rows × {len(df.columns)} columns")

        return df

    def insert_validation_columns(self, df_validation: pd.DataFrame):
        """검증 결과를 원본 Excel의 빈 컬럼에 삽입"""

        logger.info(f"\nInserting validation results into original file...")

        # Excel 파일 열기 (이미지 제외)
        wb = openpyxl.load_workbook(self.output_file, keep_vba=False)
        ws = wb["MasterData"]

        logger.info(f"  Original: {ws.max_row} rows × {ws.max_column} columns")

        # 헤더 삽입 (Row 1, Columns 14-22)
        validation_headers = [
            "Validation_Status",  # N (14)
            "Ref_Rate_USD",  # O (15)
            "Python_Delta",  # P (16)
            "CG_Band",  # Q (17)
            "Charge_Group",  # R (18)
            "Gate_Score",  # S (19)
            "Gate_Status",  # T (20)
            "PDF_Count",  # U (21)
            "Validation_Notes",  # V (22)
        ]

        # 헤더 작성
        for i, header in enumerate(validation_headers, 14):
            ws.cell(1, i, header)
            ws.cell(1, i).font = Font(bold=True)

        logger.info(f"  Headers inserted: columns 14-22")

        # 데이터 삽입 (Row 2부터)
        for row_idx in range(len(df_validation)):
            excel_row = row_idx + 2  # Excel은 1-based, 헤더가 1번 행

            # 14-22번 컬럼에 검증 결과 삽입
            ws.cell(excel_row, 14, df_validation.iloc[row_idx]["Validation_Status"])
            ws.cell(excel_row, 15, df_validation.iloc[row_idx]["Ref_Rate_USD"])
            ws.cell(excel_row, 16, df_validation.iloc[row_idx]["Python_Delta"])
            ws.cell(excel_row, 17, df_validation.iloc[row_idx]["CG_Band"])
            ws.cell(excel_row, 18, df_validation.iloc[row_idx]["Charge_Group"])
            ws.cell(excel_row, 19, df_validation.iloc[row_idx]["Gate_Score"])
            ws.cell(excel_row, 20, df_validation.iloc[row_idx]["Gate_Status"])
            ws.cell(excel_row, 21, df_validation.iloc[row_idx]["PDF_Count"])
            ws.cell(excel_row, 22, df_validation.iloc[row_idx]["Validation_Notes"])

            if (row_idx + 1) % 20 == 0:
                logger.info(f"  Inserted: {row_idx + 1}/102 rows")

        logger.info(f"  [OK] All {len(df_validation)} rows inserted")
        logger.info(f"  Final: {ws.max_row} rows × {ws.max_column} columns")

        # 저장
        wb.save(self.output_file)
        logger.info(f"\n[SAVED] Validated file: {self.output_file}")

        return wb

    def apply_conditional_formatting(self, wb: openpyxl.Workbook):
        """조건부 서식 적용"""

        logger.info(f"\nApplying conditional formatting...")

        ws = wb["MasterData"]

        # Validation_Status (N열, 14번)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        ws.conditional_formatting.add(
            f"N2:N{ws.max_row}",
            CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"N2:N{ws.max_row}",
            CellIsRule(operator="equal", formula=['"FAIL"'], fill=red_fill),
        )
        ws.conditional_formatting.add(
            f"N2:N{ws.max_row}",
            CellIsRule(operator="equal", formula=['"REVIEW_NEEDED"'], fill=yellow_fill),
        )

        # CG_Band (Q열, 17번)
        ws.conditional_formatting.add(
            f"Q2:Q{ws.max_row}",
            CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"Q2:Q{ws.max_row}",
            CellIsRule(operator="equal", formula=['"WARN"'], fill=yellow_fill),
        )
        orange_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
        ws.conditional_formatting.add(
            f"Q2:Q{ws.max_row}",
            CellIsRule(operator="equal", formula=['"HIGH"'], fill=orange_fill),
        )
        ws.conditional_formatting.add(
            f"Q2:Q{ws.max_row}",
            CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=red_fill),
        )

        # Gate_Status (T열, 20번)
        ws.conditional_formatting.add(
            f"T2:T{ws.max_row}",
            CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"T2:T{ws.max_row}",
            CellIsRule(operator="equal", formula=['"FAIL"'], fill=red_fill),
        )

        logger.info(f"  [OK] Conditional formatting applied")

        # 저장
        wb.save(self.output_file)

    def add_summary_sheet(self, wb: openpyxl.Workbook, df_validation: pd.DataFrame):
        """요약 시트 추가"""

        logger.info(f"\nAdding summary sheet...")

        # 새 시트 생성
        ws_summary = wb.create_sheet("Validation_Summary", 1)

        # 요약 데이터 작성
        summary_data = [
            ["DSV SHPT Invoice Validation Summary", ""],
            ["", ""],
            ["Total Items", len(df_validation)],
            ["Total Amount (USD)", f"${df_validation['TOTAL (USD)'].sum():,.2f}"],
            ["", ""],
            ["Validation Status", "Count"],
        ]

        # Validation Status 분포
        status_counts = df_validation["Validation_Status"].value_counts()
        for status, count in status_counts.items():
            summary_data.append([f"  {status}", count])

        summary_data.extend([["", ""], ["Charge Group", "Count"]])

        # Charge Group 분포
        cg_counts = df_validation["Charge_Group"].value_counts()
        for group, count in cg_counts.items():
            summary_data.append([f"  {group}", count])

        # Contract 상세
        contract_items = df_validation[df_validation["Charge_Group"] == "Contract"]
        summary_data.extend(
            [
                ["", ""],
                ["Contract Validation", ""],
                ["  Total Contract items", len(contract_items)],
                [
                    "  Items with ref_rate",
                    len(contract_items[contract_items["Ref_Rate_USD"].notna()]),
                ],
            ]
        )

        # 데이터 쓰기
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row_idx, col_idx, value)

        # 스타일 적용
        ws_summary.cell(1, 1).font = Font(size=14, bold=True)
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 15

        logger.info(f"  [OK] Summary sheet added")

        wb.save(self.output_file)

    def run(self):
        """전체 프로세스 실행"""

        logger.info("=" * 80)
        logger.info("Insert Validation to Original INVOICE")
        logger.info("=" * 80)

        # 1. 원본 파일 복사
        self.copy_original_file()

        # 2. 검증 결과 로드
        df_validation = self.load_validation_results()
        if df_validation is None:
            return

        # 3. 검증 결과 삽입
        wb = self.insert_validation_columns(df_validation)

        # 4. 조건부 서식 적용
        self.apply_conditional_formatting(wb)

        # 5. 요약 시트 추가
        self.add_summary_sheet(wb, df_validation)

        logger.info("\n" + "=" * 80)
        logger.info("[SUCCESS] Validation inserted to original INVOICE!")
        logger.info("=" * 80)
        logger.info(f"\n[OUTPUT] {self.output_file}")
        logger.info(f"\nStructure:")
        logger.info(f"  - MasterData: 103 rows × 22 columns")
        logger.info(f"    - Columns 1-13: VBA original")
        logger.info(f"    - Columns 14-22: Python validation")
        logger.info(f"  - Validation_Summary: Statistics")
        logger.info(f"  - Conditional formatting: Applied")


def main():
    """메인 실행"""

    inserter = ValidationInserter()
    inserter.run()


if __name__ == "__main__":
    main()

