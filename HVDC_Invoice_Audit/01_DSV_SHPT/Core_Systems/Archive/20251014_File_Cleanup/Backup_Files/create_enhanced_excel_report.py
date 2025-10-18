#!/usr/bin/env python3
"""
Enhanced Excel Report Generator for SHPT PDF Integrated Audit System (VBA 통합 이전 버전)
Documentation 기반 구조 설계 - PDF_INTEGRATION_GUIDE.md 준수

Version: 2.0.0 (VBA 제거)
Created: 2025-10-14
Author: MACHO-GPT v3.4-mini HVDC Project Enhancement
"""

import pandas as pd
import json
import ast
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# 로깅 설정
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class EnhancedExcelReportGenerator:
    """
    SHPT PDF 통합 감사 시스템을 위한 종합 Excel 보고서 생성기 (VBA 통합 이전 버전)

    Features:
    - Documentation 가이드라인 준수 (PDF_INTEGRATION_GUIDE.md)
    - PDF 통합 결과 완전 반영 (Gate-11~14 포함)
    - Cross-document 검증 상태 시각화
    - Demurrage Risk 분석 및 경고 표시
    - 증빙문서 매핑 상세 정보 제공
    - VBA 관련 기능 제거됨
    """

    def __init__(self):
        """초기화"""
        self.workbook = None
        self.data_df = None
        self.json_data = None

        # 컬럼 매핑 (Documentation 기준)
        self.column_mapping = self._define_column_mapping()

    def _define_column_mapping(self) -> Dict:
        """컬럼 매핑 정의 (Documentation 기반)"""
        return {
            # 기본 컬럼 (기존)
            "basic": [
                "s_no",
                "sheet_name",
                "description",
                "rate_source",
                "unit_rate",
                "quantity",
                "total_usd",
                "status",
                "flag",
                "delta_pct",
                "cg_band",
                "charge_group",
                "issues",
                "tolerance",
                "ref_rate_usd",
                "doc_aed",
                "gate_status",
                "gate_score",
                "gate_fails",
            ],
            # PDF 통합 컬럼 (Documentation 기준)
            "pdf_integration": [
                "pdf_validation_enabled",
                "pdf_parsed_files",
                "cross_doc_status",
                "cross_doc_issues",
                "demurrage_risk_level",
                "demurrage_days_overdue",
                "demurrage_estimated_cost",
                "gate_11_status",
                "gate_11_score",
                "gate_12_status",
                "gate_12_score",
                "gate_13_status",
                "gate_13_score",
                "gate_14_status",
                "gate_14_score",
                "supporting_docs_count",
                "supporting_docs_list",
                "evidence_types",
                "validation_timestamp",
            ],
        }

    def load_data(self, csv_path: str, json_path: Optional[str] = None) -> bool:
        """데이터 로드 (VBA 로직 제거됨)"""
        try:
            logger.info(f"Loading data from: {csv_path}")
            
            # CSV 데이터 로드
            if Path(csv_path).exists():
                self.data_df = pd.read_csv(csv_path)
                logger.info(f"CSV loaded: {len(self.data_df)} rows")
            else:
                logger.error(f"CSV file not found: {csv_path}")
                return False

            # JSON 데이터 로드 (있는 경우)
            if json_path and Path(json_path).exists():
                logger.info(f"Loading JSON data from: {json_path}")
                with open(json_path, "r", encoding="utf-8") as f:
                    self.json_data = json.load(f)

            # 데이터 향상
            if self.data_df is not None:
                self.data_df = self._enhance_dataframe()

            logger.info(f"Data loaded successfully: {len(self.data_df)} items")
            return True

        except Exception as e:
            logger.error(f"Failed to load data: {str(e)}")
            return False

    def _parse_supporting_docs(self, docs_str: str) -> List[Dict]:
        """supporting_docs_list 문자열을 파싱"""
        try:
            if pd.isna(docs_str) or docs_str == "":
                return []
            # 문자열을 리스트로 변환
            docs_list = ast.literal_eval(docs_str)
            return docs_list if isinstance(docs_list, list) else []
        except:
            return []

    def _parse_evidence_types(self, types_str: str) -> List[str]:
        """evidence_types 문자열을 파싱"""
        try:
            if pd.isna(types_str) or types_str == "":
                return []
            # 문자열을 리스트로 변환
            types_list = ast.literal_eval(types_str)
            return types_list if isinstance(types_list, list) else []
        except:
            return []

    def _enhance_dataframe(self) -> pd.DataFrame:
        """
        DataFrame에 PDF 통합 컬럼 추가 및 데이터 정제
        Documentation 기준에 따라 컬럼 확장
        """
        logger.info("Enhancing dataframe with PDF integration columns...")
        
        enhanced_df = self.data_df.copy()
        
        # PDF 통합 컬럼 추가 (없는 경우)
        pdf_columns = self.column_mapping["pdf_integration"]
        for col in pdf_columns:
            if col not in enhanced_df.columns:
                enhanced_df[col] = ""
        
        # supporting_docs_list 파싱 및 개수 계산
        if "supporting_docs_list" in enhanced_df.columns:
            enhanced_df["supporting_docs_parsed"] = enhanced_df["supporting_docs_list"].apply(
                self._parse_supporting_docs
            )
            enhanced_df["supporting_docs_count"] = enhanced_df["supporting_docs_parsed"].apply(len)
        
        # evidence_types 파싱
        if "evidence_types" in enhanced_df.columns:
            enhanced_df["evidence_types_parsed"] = enhanced_df["evidence_types"].apply(
                self._parse_evidence_types
            )
        
        logger.info(f"DataFrame enhanced: {enhanced_df.shape}")
        return enhanced_df

    def generate_excel_report(self, output_path: str) -> bool:
        """Excel 보고서 생성 (VBA 시트 제외)"""
        try:
            logger.info("Starting Excel report generation...")
            
            if self.data_df is None:
                logger.error("No data loaded")
                return False

            # 새 워크북 생성
            self.workbook = Workbook()
            
            # 기본 시트 제거
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)

            # 기본 시트들 생성
            logger.info("Creating main data sheet...")
            self.create_main_data_sheet()
            
            logger.info("Creating executive dashboard...")
            self.create_executive_dashboard()
            
            logger.info("Creating PDF integration summary...")
            self.create_pdf_integration_summary()
            
            logger.info("Creating gate analysis sheet...")
            self.create_gate_analysis_sheet()
            
            logger.info("Creating supporting docs mapping...")
            self.create_supporting_docs_mapping()

            # 파일 저장
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            self.workbook.save(output_path)

            logger.info(f"Excel report generated successfully: {output_path}")
            logger.info(f"Report contains {len(self.workbook.sheetnames)} sheets")

            return True

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {str(e)}")
            return False

    def create_main_data_sheet(self) -> None:
        """메인 데이터 시트 생성"""
        ws = self.workbook.create_sheet("Main Data")
        
        # 데이터를 시트에 추가
        for r in dataframe_to_rows(self.data_df, index=False, header=True):
            ws.append(r)
        
        # 테이블 생성
        if len(self.data_df) > 0:
            table = Table(displayName="MainDataTable", ref=f"A1:{ws.max_column_letter}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            ws.add_table(table)

    def create_executive_dashboard(self) -> None:
        """임원진 대시보드 시트 생성"""
        ws = self.workbook.create_sheet("Executive Dashboard")
        
        # 요약 정보
        ws['A1'] = "HVDC 인보이스 감사 요약"
        ws['A3'] = "총 항목 수"
        ws['B3'] = len(self.data_df)
        
        # 상태별 요약
        if 'status' in self.data_df.columns:
            status_summary = self.data_df['status'].value_counts()
            row = 5
            ws[f'A{row}'] = "상태별 요약"
            row += 1
            for status, count in status_summary.items():
                ws[f'A{row}'] = status
                ws[f'B{row}'] = count
                row += 1

    def create_pdf_integration_summary(self) -> None:
        """PDF 통합 요약 시트 생성"""
        ws = self.workbook.create_sheet("PDF Integration Summary")
        
        ws['A1'] = "PDF 통합 검증 요약"
        
        # PDF 검증 활성화 상태
        if 'pdf_validation_enabled' in self.data_df.columns:
            enabled_count = self.data_df['pdf_validation_enabled'].sum()
            ws['A3'] = "PDF 검증 활성화 항목"
            ws['B3'] = enabled_count
        
        # 게이트 상태 요약
        gate_columns = [col for col in self.data_df.columns if col.startswith('gate_') and col.endswith('_status')]
        row = 5
        ws[f'A{row}'] = "게이트 검증 요약"
        row += 1
        
        for gate_col in gate_columns:
            if gate_col in self.data_df.columns:
                pass_count = (self.data_df[gate_col] == 'PASS').sum()
                ws[f'A{row}'] = gate_col.replace('_', ' ').title()
                ws[f'B{row}'] = f"{pass_count}/{len(self.data_df)}"
                row += 1

    def create_gate_analysis_sheet(self) -> None:
        """게이트 분석 시트 생성"""
        ws = self.workbook.create_sheet("Gate Analysis")
        
        ws['A1'] = "게이트별 상세 분석"
        
        # 게이트 관련 컬럼 추출
        gate_columns = [col for col in self.data_df.columns if 'gate_' in col]
        
        if gate_columns:
            # 게이트 데이터만 추출
            gate_df = self.data_df[['s_no', 'sheet_name'] + gate_columns].copy()
            
            # 데이터를 시트에 추가
            for r in dataframe_to_rows(gate_df, index=False, header=True):
                ws.append(r)

    def create_supporting_docs_mapping(self) -> None:
        """증빙문서 매핑 시트 생성"""
        ws = self.workbook.create_sheet("Supporting Docs Mapping")
        
        ws['A1'] = "증빙문서 매핑"
        
        # 증빙문서 관련 컬럼 추출
        docs_columns = ['s_no', 'sheet_name', 'supporting_docs_count', 'supporting_docs_list', 'evidence_types']
        available_columns = [col for col in docs_columns if col in self.data_df.columns]
        
        if available_columns:
            docs_df = self.data_df[available_columns].copy()
            
            # 데이터를 시트에 추가
            for r in dataframe_to_rows(docs_df, index=False, header=True):
                ws.append(r)

    def create_comprehensive_report(self, csv_file: str, output_file: str, json_file: Optional[str] = None) -> str:
        """종합 보고서 생성 (VBA 로직 제거됨)"""
        try:
            logger.info("🚀 종합 Excel 보고서 생성 시작 (VBA 제거 버전)")
            
            # 데이터 로드
            if not self.load_data(csv_file, json_file):
                return "❌ 데이터 로드 실패"
            
            # Excel 보고서 생성
            if not self.generate_excel_report(output_file):
                return "❌ Excel 보고서 생성 실패"
            
            logger.info("✅ 종합 Excel 보고서 생성 완료")
            return f"✅ 보고서 생성 완료: {output_file}"
            
        except Exception as e:
            logger.error(f"종합 보고서 생성 실패: {str(e)}")
            return f"❌ 보고서 생성 실패: {str(e)}"


def main():
    """테스트 실행"""
    generator = EnhancedExcelReportGenerator()
    
    # 테스트 파일 경로
    csv_file = "Results/Sept_2025/shpt_sept_2025_enhanced_result_20251012_123701.csv"
    output_file = "Results/Sept_2025/Reports/enhanced_excel_report_simplified.xlsx"
    
    result = generator.create_comprehensive_report(csv_file, output_file)
    print(result)


if __name__ == "__main__":
    main()
