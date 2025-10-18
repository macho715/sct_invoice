#!/usr/bin/env python3
"""
Excel 최종 보고서 생성
SHPT Sept 2025 - Before/After Contract 검증 비교
"""

import pandas as pd
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def extract_port_from_description(description: str) -> str:
    """Description에서 Port 추출"""
    desc_upper = description.upper()

    if "ABU DHABI AIRPORT" in desc_upper or "AUH AIRPORT" in desc_upper:
        return "Abu Dhabi Airport"
    elif "DUBAI AIRPORT" in desc_upper:
        return "Dubai Airport"
    elif "KHALIFA PORT" in desc_upper or "KP" in desc_upper:
        return "Khalifa Port"
    elif "JEBEL ALI PORT" in desc_upper:
        return "Jebel Ali Port"
    else:
        return "Khalifa Port"  # 기본값


def extract_container_type(description: str) -> str:
    """Description에서 Container Type 추출"""
    desc_upper = description.upper()

    if "20DC" in desc_upper or "1 X 20DC" in desc_upper or "1X20DC" in desc_upper:
        return "20DC"
    elif "40HC" in desc_upper or "1 X 40HC" in desc_upper or "2 X 40HC" in desc_upper:
        return "40HC"
    elif "KG" in desc_upper or "CW:" in desc_upper:
        match = re.search(r"(\d+)\s*KG", desc_upper)
        if match:
            return f"{match.group(1)} KG"

    return "Container"


def extract_route(description: str) -> str:
    """Description에서 Transportation Route 추출"""
    desc_upper = description.upper()

    # FROM ... TO ... 패턴
    match = re.search(r"FROM\s+(.+?)\s+TO\s+(.+?)(?:\(|$)", desc_upper)
    if match:
        origin = match.group(1).strip()
        destination = match.group(2).strip()

        # 간소화
        origin = origin.replace("KHALIFA PORT", "KP").replace(
            "DSV MUSSAFAH YARD", "DSV"
        )
        destination = destination.replace("DSV MUSSAFAH YARD", "DSV").replace(
            "KHALIFA PORT", "KP"
        )

        return f"{origin}→{destination}"

    return "Unknown Route"


def generate_decision_logic(row):
    """
    Contract 항목의 Decision Logic 생성
    """
    description = row["description"]
    ref_rate = row["ref_rate_usd"]
    draft_rate = row["unit_rate"]
    delta = row["delta_pct"]
    status = row["status"]
    cg_band = row.get("cg_band", "N/A")

    if pd.isna(ref_rate) or ref_rate == 0:
        return "Ref rate not found in rate database; requires manual review"

    desc_upper = description.upper()

    # MASTER DO FEE
    if "MASTER DO FEE" in desc_upper:
        port = extract_port_from_description(description)
        if abs(delta) <= 2.0:
            return f"Contract rate for DO Fee at {port} = ${ref_rate:.2f}; matches draft → {status}"
        else:
            return f"Contract rate for DO Fee at {port} = ${ref_rate:.2f}; delta {delta:.2f}% (Band: {cg_band}) → {status}"

    # CUSTOMS CLEARANCE
    elif "CUSTOMS CLEARANCE" in desc_upper:
        if abs(delta) <= 2.0:
            return (
                f"Standard clearance rate = ${ref_rate:.2f}; matches draft → {status}"
            )
        else:
            return f"Standard clearance rate = ${ref_rate:.2f}; delta {delta:.2f}% → {status}"

    # TERMINAL HANDLING
    elif "TERMINAL HANDLING" in desc_upper:
        container_type = extract_container_type(description)
        if abs(delta) <= 2.0:
            return f"Contract THC for {container_type} = ${ref_rate:.2f}; matches draft → {status}"
        else:
            return f"Contract THC for {container_type} = ${ref_rate:.2f}; delta {delta:.2f}% (Band: {cg_band}) → {status}"

    # TRANSPORTATION
    elif "TRANSPORTATION" in desc_upper or "TRUCKING" in desc_upper:
        route = extract_route(description)
        if abs(delta) <= 2.0:
            return f"Lane {route}; std lane rate ${ref_rate:.2f} per truck; matches draft → {status}"
        else:
            return f"Lane {route}; std lane rate ${ref_rate:.2f} per truck; delta {delta:.2f}% → {status}"

    # 기타
    else:
        if abs(delta) <= 2.0:
            return f"Contract rate ${ref_rate:.2f}; matches draft → {status}"
        else:
            return f"Contract rate ${ref_rate:.2f}; draft ${draft_rate:.2f}; delta {delta:.2f}% (Band: {cg_band}) → {status}"


def format_attachments(supporting_docs_str):
    """
    JSON 문자열을 Excel용 포맷으로 변환
    """
    if pd.isna(supporting_docs_str) or supporting_docs_str == "[]":
        return ""

    try:
        # JSON 파싱
        docs_str = str(supporting_docs_str).replace("'", '"')
        docs = json.loads(docs_str)

        if not docs:
            return ""

        # doc_type별로 그룹화
        doc_types = {}
        for doc in docs:
            doc_type = doc.get("doc_type", "Other")
            file_name = doc.get("file_name", "Unknown")
            if doc_type not in doc_types:
                doc_types[doc_type] = []
            doc_types[doc_type].append(file_name)

        # 포맷팅
        formatted = []
        for doc_type in sorted(doc_types.keys()):
            files = doc_types[doc_type]
            if len(files) == 1:
                formatted.append(f"**{doc_type}:** {files[0]}")
            else:
                formatted.append(f"**{doc_type}:** {files[0]} (+{len(files)-1} more)")

        return "\n".join(formatted)

    except Exception as e:
        # 파싱 실패 시 간단히 처리
        if len(supporting_docs_str) > 100:
            return supporting_docs_str[:100] + "..."
        return supporting_docs_str


def create_comparison_report(df_after):
    """
    After CSV에서 Contract 항목만 추출하여 Comparison Report 생성
    """
    # Contract 항목만 필터링
    contract_items = df_after[df_after["charge_group"] == "Contract"].copy()

    # Ref Total 계산
    contract_items["ref_total"] = contract_items.apply(
        lambda row: (
            row["ref_rate_usd"] * row["quantity"]
            if pd.notna(row["ref_rate_usd"])
            else 0
        ),
        axis=1,
    )

    # Verdict 단순화
    contract_items["verdict"] = contract_items["status"].apply(
        lambda x: "PASS" if x == "PASS" else "FAIL/REVIEW"
    )

    # Notes (COST-GUARD Band)
    contract_items["notes_field"] = contract_items.apply(
        lambda row: (
            row["cg_band"]
            if pd.notna(row.get("cg_band")) and row["cg_band"] != "PASS"
            else ""
        ),
        axis=1,
    )

    # 새 DataFrame 생성
    comparison = pd.DataFrame(
        {
            "S/No": contract_items["s_no"],
            "Item (Description)": contract_items["description"],
            "Rate Source": contract_items["rate_source"],
            "Draft Rate (USD)": contract_items["unit_rate"].round(2),
            "Qty": contract_items["quantity"].astype(int),
            "Draft Total (USD)": contract_items["total_usd"].round(2),
            "Ref/Rev Rate (USD)": contract_items["ref_rate_usd"].round(2),
            "Ref Total (USD)": contract_items["ref_total"].round(2),
            "Δ%": contract_items["delta_pct"].round(2),
            "Verdict": contract_items["verdict"],
            "Notes": contract_items["notes_field"],
            "Attachments / Evidence": contract_items["supporting_docs_list"].apply(
                format_attachments
            ),
            "Decision Logic": contract_items.apply(generate_decision_logic, axis=1),
        }
    )

    return comparison


def apply_header_format(ws):
    """헤더 행 포맷팅"""
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border


def apply_conditional_formatting(ws, verdict_col="J"):
    """Verdict 컬럼에 조건부 서식 적용"""
    from openpyxl.styles import PatternFill

    # PASS = 녹색
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # FAIL = 빨간색
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 데이터 행에 적용
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{verdict_col}{row}"]
        if cell.value:
            if "PASS" in str(cell.value):
                cell.fill = pass_fill
            elif "FAIL" in str(cell.value) or "REVIEW" in str(cell.value):
                cell.fill = fail_fill


def adjust_column_widths(ws, sheet_name):
    """컬럼 너비 자동 조정"""
    if sheet_name == "Comparison_Report":
        column_widths = {
            "A": 8,  # S/No
            "B": 55,  # Item (Description)
            "C": 15,  # Rate Source
            "D": 15,  # Draft Rate
            "E": 8,  # Qty
            "F": 16,  # Draft Total
            "G": 16,  # Ref/Rev Rate
            "H": 16,  # Ref Total
            "I": 10,  # Δ%
            "J": 13,  # Verdict
            "K": 15,  # Notes
            "L": 45,  # Attachments
            "M": 65,  # Decision Logic
        }
    else:
        # Before/After Data 시트
        column_widths = {chr(65 + i): 15 for i in range(26)}  # A-Z
        column_widths["B"] = 50  # description
        column_widths["C"] = 50  # sheet_name 또는 description

    for col, width in column_widths.items():
        if col in [chr(65 + i) for i in range(26)]:  # A-Z only
            ws.column_dimensions[col].width = width


def freeze_header(ws):
    """첫 행 고정"""
    ws.freeze_panes = "A2"


def create_excel_report(df_before, df_after, comparison_df, output_path):
    """
    3개 시트로 Excel 보고서 생성
    """
    print("\nCreating Excel workbook...")
    wb = Workbook()

    # 시트 1: Before Data
    print("  - Writing Before_Data sheet...")
    ws1 = wb.active
    ws1.title = "Before_Data"
    for r in dataframe_to_rows(df_before, index=False, header=True):
        ws1.append(r)
    apply_header_format(ws1)
    adjust_column_widths(ws1, "Before_Data")
    freeze_header(ws1)

    # 시트 2: After Data
    print("  - Writing After_Data sheet...")
    ws2 = wb.create_sheet("After_Data")
    for r in dataframe_to_rows(df_after, index=False, header=True):
        ws2.append(r)
    apply_header_format(ws2)
    adjust_column_widths(ws2, "After_Data")
    freeze_header(ws2)

    # 시트 3: Comparison Report
    print("  - Writing Comparison_Report sheet...")
    ws3 = wb.create_sheet("Comparison_Report")
    for r in dataframe_to_rows(comparison_df, index=False, header=True):
        ws3.append(r)
    apply_header_format(ws3)
    apply_conditional_formatting(ws3)
    adjust_column_widths(ws3, "Comparison_Report")
    freeze_header(ws3)

    # 저장
    print(f"  - Saving to {output_path.name}...")
    wb.save(output_path)
    print(f"[OK] Excel report created successfully!")


def main():
    """메인 실행 함수"""
    print("=" * 80)
    print("SHPT Contract Validation Excel Report Generator")
    print("=" * 80)

    # 경로 설정
    base_dir = Path(__file__).parent.parent
    before_csv = (
        base_dir
        / "Results"
        / "Sept_2025"
        / "shpt_sept_2025_enhanced_result_20251012_123701.csv"
    )
    after_csv = (
        base_dir
        / "Results"
        / "Sept_2025"
        / "CSV"
        / "shpt_sept_2025_enhanced_result_20251013_002105.csv"
    )
    output_excel = (
        base_dir
        / "Results"
        / "Sept_2025"
        / "Reports"
        / "SHPT_SEPT_2025_CONTRACT_VALIDATION_REPORT.xlsx"
    )

    # CSV 존재 확인
    if not before_csv.exists():
        print(f"[ERROR] Before CSV not found: {before_csv}")
        return

    if not after_csv.exists():
        print(f"[ERROR] After CSV not found: {after_csv}")
        return

    # CSV 로드
    print("\nLoading CSV files...")
    print(f"  - Before: {before_csv.name}")
    df_before = pd.read_csv(before_csv)
    print(f"    OK Loaded {len(df_before)} items")

    print(f"  - After: {after_csv.name}")
    df_after = pd.read_csv(after_csv)
    print(f"    OK Loaded {len(df_after)} items")

    # Comparison Report 생성
    print("\nGenerating Comparison Report...")
    comparison_df = create_comparison_report(df_after)
    print(f"  OK Generated {len(comparison_df)} Contract items")

    # Excel 생성
    create_excel_report(df_before, df_after, comparison_df, output_excel)

    # 통계 출력
    print("\n" + "=" * 80)
    print("Report Statistics")
    print("=" * 80)
    print(f"Before Data: {len(df_before)} items")
    print(f"After Data: {len(df_after)} items")
    print(f"Contract Items: {len(comparison_df)} items")

    # Verdict 통계
    pass_count = len(comparison_df[comparison_df["Verdict"] == "PASS"])
    fail_count = len(comparison_df[comparison_df["Verdict"] != "PASS"])

    print(f"\n  - PASS: {pass_count} ({pass_count/len(comparison_df)*100:.1f}%)")
    print(f"  - FAIL/REVIEW: {fail_count} ({fail_count/len(comparison_df)*100:.1f}%)")

    # ref_rate 통계
    ref_rate_filled = comparison_df["Ref/Rev Rate (USD)"].notna().sum()
    print(
        f"\n  - Ref Rate Filled: {ref_rate_filled}/{len(comparison_df)} ({ref_rate_filled/len(comparison_df)*100:.1f}%)"
    )

    print(f"\nOutput Location:")
    print(f"   {output_excel}")
    print("\n[SUCCESS] Excel report generation complete!")
    print("=" * 80)


if __name__ == "__main__":
    main()
