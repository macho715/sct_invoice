#!/usr/bin/env python3
"""
Comprehensive Excel Report Generator
종합 Excel 보고서 생성기 - VBA, Python, PDF 통합 결과 포함

Version: 2.0.0
Created: 2025-10-13
Author: MACHO-GPT v3.4-mini HVDC Project Enhancement
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, DataBarRule
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional
import numpy as np

# 로깅 설정
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

class ComprehensiveExcelReportGenerator:
    """
    종합 Excel 보고서 생성기
    
    Features:
    - Executive Dashboard (경영진 요약)
    - VBA Analysis Results (VBA 분석 결과)
    - Python Audit Results (Python 감사 결과)
    - PDF Integration Summary (PDF 통합 요약)
    - Cross-Document Validation (교차 문서 검증)
    - Gate Validation Results (Gate 1-14 검증)
    - Compliance Check (규제 준수)
    - Anomaly Detection (이상 탐지)
    - Action Items (액션 아이템)
    - Detailed Statistics (상세 통계)
    """
    
    def __init__(self):
        """초기화"""
        self.logger = logger
        self.workbook = None
        self.styles = self._create_styles()
        
        # 색상 팔레트
        self.colors = {
            "primary": "4F81BD",      # 파란색
            "success": "70AD47",      # 녹색
            "warning": "FFC000",      # 주황색
            "danger": "C5504B",       # 빨간색
            "info": "5B9BD5",         # 하늘색
            "light": "F2F2F2",        # 회색
            "dark": "2F4F4F"          # 진한 회색
        }
    
    def _create_styles(self) -> Dict[str, NamedStyle]:
        """Excel 스타일 생성"""
        styles = {}
        
        # 헤더 스타일
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_style.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        styles["header"] = header_style
        
        # 서브헤더 스타일
        subheader_style = NamedStyle(name="subheader")
        subheader_style.font = Font(bold=True, size=10)
        subheader_style.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        subheader_style.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        subheader_style.alignment = Alignment(horizontal="center", vertical="center")
        styles["subheader"] = subheader_style
        
        # 데이터 스타일
        data_style = NamedStyle(name="data")
        data_style.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        data_style.alignment = Alignment(horizontal="left", vertical="center")
        styles["data"] = data_style
        
        # 숫자 스타일
        number_style = NamedStyle(name="number")
        number_style.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.number_format = "#,##0.00"
        styles["number"] = number_style
        
        # 퍼센트 스타일
        percent_style = NamedStyle(name="percent")
        percent_style.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        percent_style.alignment = Alignment(horizontal="right", vertical="center")
        percent_style.number_format = "0.0%"
        styles["percent"] = percent_style
        
        return styles
    
    def generate_comprehensive_report(self, validation_results: Dict[str, Any], 
                                    output_dir: str) -> str:
        """
        종합 Excel 보고서 생성
        
        Args:
            validation_results: 종합 검증 결과
            output_dir: 출력 디렉토리
            
        Returns:
            str: 생성된 Excel 파일 경로
        """
        self.logger.info("📊 종합 Excel 보고서 생성 시작")
        
        try:
            # 워크북 생성
            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)  # 기본 시트 제거
            
            # 스타일 등록
            for style in self.styles.values():
                if style.name not in self.workbook.named_styles:
                    self.workbook.add_named_style(style)
            
            # 시트별 생성
            self._create_executive_dashboard(validation_results)
            self._create_vba_analysis_sheet(validation_results)
            self._create_python_audit_sheet(validation_results)
            self._create_pdf_integration_sheet(validation_results)
            self._create_cross_validation_sheet(validation_results)
            self._create_gate_validation_sheet(validation_results)
            self._create_compliance_sheet(validation_results)
            self._create_anomaly_detection_sheet(validation_results)
            self._create_action_items_sheet(validation_results)
            self._create_statistics_sheet(validation_results)
            
            # 파일 저장
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(output_dir) / f"HVDC_Comprehensive_Validation_Report_{timestamp}.xlsx"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.workbook.save(output_path)
            self.logger.info(f"✅ 종합 Excel 보고서 생성 완료: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"❌ Excel 보고서 생성 실패: {e}")
            raise
    
    def _create_executive_dashboard(self, validation_results: Dict[str, Any]):
        """Executive Dashboard 시트 생성"""
        ws = self.workbook.create_sheet("Executive Dashboard")
        self.logger.info("  📊 Executive Dashboard 생성...")
        
        # 제목
        ws.merge_cells("A1:H1")
        ws["A1"] = "HVDC 9월 인보이스 검증 - Executive Dashboard"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color=self.colors["primary"], end_color=self.colors["primary"], fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # 검증 일시
        ws["A3"] = f"검증 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"].font = Font(size=12, italic=True)
        
        # 주요 KPI 카드
        summary = validation_results.get("validation_summary", {})
        
        # KPI 헤더
        kpi_headers = ["전체 신뢰도", "품질 등급", "Gate 통과율", "규제 준수", "검증 시간"]
        for i, header in enumerate(kpi_headers):
            cell = ws.cell(row=5, column=i+1, value=header)
            cell.style = "header"
        
        # KPI 값
        gate_results = validation_results.get("validation_results", {}).get("gate_validation", {})
        compliance_results = validation_results.get("validation_results", {}).get("compliance_check", {})
        
        kpi_values = [
            f"{summary.get('confidence_score', 0):.1%}",
            summary.get('quality_grade', 'UNKNOWN'),
            f"{gate_results.get('overall_pass_rate', 0):.1%}",
            f"{compliance_results.get('overall_compliance_score', 0):.1%}",
            f"{summary.get('total_validation_time', 0):.1f}초"
        ]
        
        for i, value in enumerate(kpi_values):
            cell = ws.cell(row=6, column=i+1, value=value)
            cell.style = "data"
            cell.font = Font(size=14, bold=True)
            
            # 조건부 색상
            if i == 0:  # 신뢰도
                score = summary.get('confidence_score', 0)
                if score >= 0.9:
                    cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
                elif score >= 0.7:
                    cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
        
        # 시스템 상태 요약
        ws["A8"] = "시스템 상태 요약"
        ws["A8"].font = Font(size=14, bold=True)
        
        # 시스템 상태 테이블
        system_headers = ["시스템", "상태", "세부사항"]
        for i, header in enumerate(system_headers):
            cell = ws.cell(row=9, column=i+1, value=header)
            cell.style = "subheader"
        
        system_results = validation_results.get("system_results", {})
        systems = [
            ("VBA Analysis", system_results.get("vba_analysis", {}).get("status", "UNKNOWN")),
            ("Python Audit", system_results.get("python_audit", {}).get("status", "UNKNOWN")),
            ("PDF Integration", system_results.get("pdf_integration", {}).get("status", "UNKNOWN")),
            ("Cross-System Integration", system_results.get("cross_system_integration", {}).get("status", "UNKNOWN"))
        ]
        
        for i, (system, status) in enumerate(systems, start=10):
            ws.cell(row=i, column=1, value=system).style = "data"
            status_cell = ws.cell(row=i, column=2, value=status)
            status_cell.style = "data"
            
            # 상태별 색상
            if status == "SUCCESS":
                status_cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
            elif status in ["FAILED", "ERROR"]:
                status_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
            elif status == "SKIPPED":
                status_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
        
        # 권장사항
        ws["A15"] = "권장사항"
        ws["A15"].font = Font(size=14, bold=True)
        
        recommendations = validation_results.get("recommendations", [])
        for i, rec in enumerate(recommendations, start=16):
            ws[f"A{i}"] = f"• {rec}"
            ws[f"A{i}"].style = "data"
        
        # 컬럼 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_vba_analysis_sheet(self, validation_results: Dict[str, Any]):
        """VBA Analysis Results 시트 생성"""
        ws = self.workbook.create_sheet("VBA Analysis")
        self.logger.info("  📋 VBA Analysis 시트 생성...")
        
        vba_results = validation_results.get("system_results", {}).get("vba_analysis", {})
        
        # 제목
        ws["A1"] = "VBA Excel 분석 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        if vba_results.get("status") != "SUCCESS":
            ws["A3"] = f"VBA 분석 실패: {vba_results.get('reason', 'Unknown error')}"
            return
        
        # 요약 정보
        ws["A3"] = "분석 요약"
        ws["A3"].font = Font(size=14, bold=True)
        
        summary_data = [
            ("분석된 시트 수", vba_results.get("sheets_analyzed", 0)),
            ("추출된 Formula 수", vba_results.get("formulas_extracted", 0)),
            ("분석된 계산 수", vba_results.get("calculations_analyzed", 0)),
            ("MasterData 행 수", vba_results.get("master_data_rows", 0)),
            ("검증 점수", vba_results.get("validation_score", "N/A")),
            ("검증 상태", vba_results.get("validation_status", "UNKNOWN"))
        ]
        
        # 요약 테이블
        headers = ["항목", "값"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1, value=header)
            cell.style = "header"
        
        for i, (item, value) in enumerate(summary_data, start=5):
            ws.cell(row=i, column=1, value=item).style = "data"
            ws.cell(row=i, column=2, value=value).style = "data"
        
        # 상세 결과 (있는 경우)
        detailed_results = vba_results.get("detailed_results", {})
        if detailed_results:
            ws["A12"] = "상세 VBA 결과"
            ws["A12"].font = Font(size=14, bold=True)
            
            # Formula 추출 결과
            formulas = detailed_results.get("vba_results", {}).get("formula_extraction", {}).get("formulas", [])
            if formulas:
                ws["A14"] = "추출된 Formula (상위 10개)"
                ws["A14"].font = Font(size=12, bold=True)
                
                formula_headers = ["시트", "셀", "Formula"]
                for i, header in enumerate(formula_headers):
                    cell = ws.cell(row=15, column=i+1, value=header)
                    cell.style = "subheader"
                
                for i, formula in enumerate(formulas[:10], start=16):
                    ws.cell(row=i, column=1, value=formula.get("sheet", "")).style = "data"
                    ws.cell(row=i, column=2, value=formula.get("cell", "")).style = "data"
                    ws.cell(row=i, column=3, value=formula.get("formula", "")).style = "data"
        
        self._adjust_column_widths(ws)
    
    def _create_python_audit_sheet(self, validation_results: Dict[str, Any]):
        """Python Audit Results 시트 생성"""
        ws = self.workbook.create_sheet("Python Audit")
        self.logger.info("  📋 Python Audit 시트 생성...")
        
        python_results = validation_results.get("system_results", {}).get("python_audit", {})
        
        # 제목
        ws["A1"] = "Python 감사 시스템 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        if python_results.get("status") != "SUCCESS":
            ws["A3"] = f"Python 감사 결과 로드 실패: {python_results.get('error', 'Unknown error')}"
            return
        
        # CSV 데이터 요약
        csv_data = python_results.get("csv_data")
        if csv_data is not None:
            ws["A3"] = f"감사 항목 수: {len(csv_data)}"
            ws["A3"].font = Font(size=14, bold=True)
            
            # 상위 10개 항목 표시
            if len(csv_data) > 0:
                ws["A5"] = "감사 결과 샘플 (상위 10개)"
                ws["A5"].font = Font(size=12, bold=True)
                
                # 헤더
                if len(csv_data.columns) > 0:
                    for i, col in enumerate(csv_data.columns[:8]):  # 최대 8개 컬럼
                        cell = ws.cell(row=6, column=i+1, value=col)
                        cell.style = "header"
                    
                    # 데이터 (상위 10행)
                    for row_idx in range(min(10, len(csv_data))):
                        for col_idx, col in enumerate(csv_data.columns[:8]):
                            value = csv_data.iloc[row_idx, csv_data.columns.get_loc(col)]
                            cell = ws.cell(row=row_idx+7, column=col_idx+1, value=str(value))
                            cell.style = "data"
        
        # JSON 데이터 요약
        json_data = python_results.get("json_data")
        if json_data:
            ws["A18"] = "JSON 감사 데이터 구조"
            ws["A18"].font = Font(size=12, bold=True)
            
            # JSON 키 표시
            if isinstance(json_data, dict):
                for i, key in enumerate(list(json_data.keys())[:10], start=19):
                    ws[f"A{i}"] = f"• {key}"
                    ws[f"A{i}"].style = "data"
        
        self._adjust_column_widths(ws)
    
    def _create_pdf_integration_sheet(self, validation_results: Dict[str, Any]):
        """PDF Integration Summary 시트 생성"""
        ws = self.workbook.create_sheet("PDF Integration")
        self.logger.info("  📄 PDF Integration 시트 생성...")
        
        pdf_results = validation_results.get("system_results", {}).get("pdf_integration", {})
        
        # 제목
        ws["A1"] = "PDF 문서 통합 검증 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        if pdf_results.get("status") == "SKIPPED":
            ws["A3"] = f"PDF 통합 건너뜀: {pdf_results.get('reason', 'Disabled')}"
            return
        elif pdf_results.get("status") == "FAILED":
            ws["A3"] = f"PDF 통합 실패: {pdf_results.get('error', 'Unknown error')}"
            return
        
        # PDF 처리 요약
        ws["A3"] = "PDF 처리 요약"
        ws["A3"].font = Font(size=14, bold=True)
        
        summary_data = [
            ("전체 PDF 파일 수", pdf_results.get("total_pdfs", 0)),
            ("성공적으로 파싱된 파일", pdf_results.get("parsed_successfully", 0)),
            ("파싱 실패 파일", pdf_results.get("parsing_failures", 0)),
            ("파싱 성공률", f"{pdf_results.get('parsed_successfully', 0) / max(1, pdf_results.get('total_pdfs', 1)):.1%}")
        ]
        
        # 요약 테이블
        headers = ["항목", "값"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1, value=header)
            cell.style = "header"
        
        for i, (item, value) in enumerate(summary_data, start=5):
            ws.cell(row=i, column=1, value=item).style = "data"
            ws.cell(row=i, column=2, value=value).style = "data"
        
        # 처리된 문서 목록
        documents = pdf_results.get("documents", [])
        if documents:
            ws["A10"] = f"처리된 문서 목록 (상위 {min(10, len(documents))}개)"
            ws["A10"].font = Font(size=12, bold=True)
            
            doc_headers = ["파일명", "문서 유형", "신뢰도"]
            for i, header in enumerate(doc_headers):
                cell = ws.cell(row=11, column=i+1, value=header)
                cell.style = "subheader"
            
            for i, doc in enumerate(documents[:10], start=12):
                file_path = doc.get("file_path", "")
                file_name = Path(file_path).name if file_path else "Unknown"
                
                ws.cell(row=i, column=1, value=file_name).style = "data"
                ws.cell(row=i, column=2, value=doc.get("doc_type", "Unknown")).style = "data"
                
                confidence = doc.get("confidence", 0.0)
                conf_cell = ws.cell(row=i, column=3, value=f"{confidence:.1%}")
                conf_cell.style = "percent"
                
                # 신뢰도별 색상
                if confidence >= 0.8:
                    conf_cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
                elif confidence >= 0.6:
                    conf_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
        
        self._adjust_column_widths(ws)
    
    def _create_cross_validation_sheet(self, validation_results: Dict[str, Any]):
        """Cross-Document Validation 시트 생성"""
        ws = self.workbook.create_sheet("Cross Validation")
        self.logger.info("  🔗 Cross Validation 시트 생성...")
        
        integration_results = validation_results.get("system_results", {}).get("cross_system_integration", {})
        
        # 제목
        ws["A1"] = "시스템 간 교차 검증 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        # VBA-Python 통합 결과
        vba_integration = integration_results.get("vba_integration", {})
        if vba_integration.get("status") != "NO_DATA":
            ws["A3"] = "VBA-Python 데이터 매칭 결과"
            ws["A3"].font = Font(size=14, bold=True)
            
            vba_match_data = [
                ("VBA 계산 항목 수", vba_integration.get("total_vba_calculations", 0)),
                ("Python 항목 수", vba_integration.get("total_python_items", 0)),
                ("매칭된 항목 수", vba_integration.get("matched_items", 0)),
                ("매칭 정확도", f"{vba_integration.get('accuracy_score', 0):.1%}")
            ]
            
            # 매칭 결과 테이블
            headers = ["항목", "값"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=4, column=i+1, value=header)
                cell.style = "header"
            
            for i, (item, value) in enumerate(vba_match_data, start=5):
                ws.cell(row=i, column=1, value=item).style = "data"
                ws.cell(row=i, column=2, value=value).style = "data"
        
        # PDF-Invoice 통합 결과
        pdf_integration = integration_results.get("pdf_integration", {})
        if pdf_integration.get("status") != "NO_DATA":
            ws["A10"] = "PDF-Invoice 데이터 매칭 결과"
            ws["A10"].font = Font(size=14, bold=True)
            
            pdf_match_data = [
                ("PDF 문서 수", pdf_integration.get("total_pdf_documents", 0)),
                ("Invoice 항목 수", pdf_integration.get("total_invoice_items", 0)),
                ("MBL 매칭 수", pdf_integration.get("mbl_matches", 0)),
                ("Container 매칭 수", pdf_integration.get("container_matches", 0))
            ]
            
            # PDF 매칭 결과 테이블
            headers = ["항목", "값"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=11, column=i+1, value=header)
                cell.style = "header"
            
            for i, (item, value) in enumerate(pdf_match_data, start=12):
                ws.cell(row=i, column=1, value=item).style = "data"
                ws.cell(row=i, column=2, value=value).style = "data"
        
        # 데이터 일관성 결과
        consistency = integration_results.get("data_consistency", {})
        if consistency:
            ws["A17"] = "전체 데이터 일관성"
            ws["A17"].font = Font(size=14, bold=True)
            
            consistency_data = [
                ("전체 상태", consistency.get("overall_status", "UNKNOWN")),
                ("데이터 품질 점수", f"{consistency.get('data_quality_score', 0):.1%}"),
                ("VBA 상태", consistency.get("vba_status", "UNKNOWN")),
                ("Python 상태", consistency.get("python_status", "UNKNOWN")),
                ("PDF 상태", consistency.get("pdf_status", "UNKNOWN"))
            ]
            
            # 일관성 결과 테이블
            headers = ["항목", "상태"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=18, column=i+1, value=header)
                cell.style = "header"
            
            for i, (item, status) in enumerate(consistency_data, start=19):
                ws.cell(row=i, column=1, value=item).style = "data"
                status_cell = ws.cell(row=i, column=2, value=status)
                status_cell.style = "data"
                
                # 상태별 색상 (품질 점수 제외)
                if item != "데이터 품질 점수":
                    if status in ["SUCCESS", "HIGH_QUALITY"]:
                        status_cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
                    elif status in ["MEDIUM_QUALITY", "PARTIAL"]:
                        status_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
                    elif status in ["LOW_QUALITY", "FAILED"]:
                        status_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
        
        self._adjust_column_widths(ws)
    
    def _create_gate_validation_sheet(self, validation_results: Dict[str, Any]):
        """Gate Validation Results 시트 생성"""
        ws = self.workbook.create_sheet("Gate Validation")
        self.logger.info("  🚪 Gate Validation 시트 생성...")
        
        gate_results = validation_results.get("validation_results", {}).get("gate_validation", {})
        
        # 제목
        ws["A1"] = "Gate 1-14 검증 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        # 전체 요약
        ws["A3"] = "검증 요약"
        ws["A3"].font = Font(size=14, bold=True)
        
        summary_data = [
            ("전체 Gate 수", gate_results.get("total_gates", 14)),
            ("통과한 Gate 수", gate_results.get("passed_gates", 0)),
            ("실패한 Gate 수", gate_results.get("failed_gates", 0)),
            ("전체 통과율", f"{gate_results.get('overall_pass_rate', 0):.1%}")
        ]
        
        # 요약 테이블
        headers = ["항목", "값"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1, value=header)
            cell.style = "header"
        
        for i, (item, value) in enumerate(summary_data, start=5):
            ws.cell(row=i, column=1, value=item).style = "data"
            ws.cell(row=i, column=2, value=value).style = "data"
        
        # 상세 Gate 결과
        gate_details = gate_results.get("gate_details", {})
        if gate_details:
            ws["A10"] = "상세 Gate 검증 결과"
            ws["A10"].font = Font(size=14, bold=True)
            
            # Gate 결과 테이블 헤더
            detail_headers = ["Gate", "상태", "설명", "세부사항"]
            for i, header in enumerate(detail_headers):
                cell = ws.cell(row=11, column=i+1, value=header)
                cell.style = "header"
            
            # Gate 결과 데이터
            row_num = 12
            for gate_name, gate_info in gate_details.items():
                ws.cell(row=row_num, column=1, value=gate_name).style = "data"
                
                status = gate_info.get("status", "UNKNOWN")
                status_cell = ws.cell(row=row_num, column=2, value=status)
                status_cell.style = "data"
                
                # 상태별 색상
                if status == "PASS":
                    status_cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
                elif status == "FAIL":
                    status_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
                elif status in ["SKIP", "UNKNOWN"]:
                    status_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
                
                ws.cell(row=row_num, column=3, value=gate_info.get("description", "")).style = "data"
                
                # 세부사항 (간단히)
                details = gate_info.get("details", {})
                if isinstance(details, dict):
                    detail_text = ", ".join([f"{k}: {v}" for k, v in list(details.items())[:3]])
                else:
                    detail_text = str(details)[:50]  # 50자 제한
                
                ws.cell(row=row_num, column=4, value=detail_text).style = "data"
                
                row_num += 1
        
        self._adjust_column_widths(ws)
    
    def _create_compliance_sheet(self, validation_results: Dict[str, Any]):
        """Compliance Check 시트 생성"""
        ws = self.workbook.create_sheet("Compliance")
        self.logger.info("  📋 Compliance 시트 생성...")
        
        compliance_results = validation_results.get("validation_results", {}).get("compliance_check", {})
        
        # 제목
        ws["A1"] = "규제 준수 확인 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        # 전체 준수 점수
        overall_score = compliance_results.get("overall_compliance_score", 0)
        ws["A3"] = f"전체 규제 준수 점수: {overall_score:.1%}"
        ws["A3"].font = Font(size=14, bold=True)
        
        # 개별 규제 확인 결과
        ws["A5"] = "개별 규제 준수 상태"
        ws["A5"].font = Font(size=14, bold=True)
        
        # 규제 결과 테이블 헤더
        compliance_headers = ["규제 기관", "준수 상태", "세부사항"]
        for i, header in enumerate(compliance_headers):
            cell = ws.cell(row=6, column=i+1, value=header)
            cell.style = "header"
        
        # 규제별 결과
        regulations = [
            ("FANR (Nuclear)", compliance_results.get("fanr_compliance", {})),
            ("MOIAT (Electrical)", compliance_results.get("moiat_compliance", {})),
            ("DCD (Hazmat)", compliance_results.get("dcd_compliance", {}))
        ]
        
        for i, (reg_name, reg_info) in enumerate(regulations, start=7):
            ws.cell(row=i, column=1, value=reg_name).style = "data"
            
            status = reg_info.get("status", "UNKNOWN")
            status_cell = ws.cell(row=i, column=2, value=status)
            status_cell.style = "data"
            
            # 상태별 색상
            if status == "COMPLIANT":
                status_cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
            elif status == "NOT_REQUIRED":
                status_cell.fill = PatternFill(start_color=self.colors["info"], end_color=self.colors["info"], fill_type="solid")
            elif status == "NON_COMPLIANT":
                status_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
            
            # 세부사항
            details = reg_info.get("details", {})
            if isinstance(details, dict):
                detail_text = ", ".join([f"{k}: {v}" for k, v in list(details.items())[:2]])
            else:
                detail_text = str(details)[:100]  # 100자 제한
            
            ws.cell(row=i, column=3, value=detail_text).style = "data"
        
        # 준수 문제사항
        issues = compliance_results.get("issues", [])
        if issues:
            ws["A12"] = "준수 문제사항"
            ws["A12"].font = Font(size=14, bold=True)
            
            for i, issue in enumerate(issues, start=13):
                ws[f"A{i}"] = f"• {issue}"
                ws[f"A{i}"].style = "data"
                ws[f"A{i}"].fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
        
        self._adjust_column_widths(ws)
    
    def _create_anomaly_detection_sheet(self, validation_results: Dict[str, Any]):
        """Anomaly Detection 시트 생성"""
        ws = self.workbook.create_sheet("Anomaly Detection")
        self.logger.info("  🤖 Anomaly Detection 시트 생성...")
        
        anomaly_results = validation_results.get("validation_results", {}).get("anomaly_detection", {})
        
        # 제목
        ws["A1"] = "AI 기반 이상 탐지 결과"
        ws["A1"].font = Font(size=16, bold=True)
        
        if anomaly_results.get("status") == "DISABLED":
            ws["A3"] = "이상 탐지 기능이 비활성화되었습니다."
            return
        elif anomaly_results.get("status") == "FAILED":
            ws["A3"] = f"이상 탐지 실패: {anomaly_results.get('error', 'Unknown error')}"
            return
        
        # 이상 탐지 요약
        ws["A3"] = "이상 탐지 요약"
        ws["A3"].font = Font(size=14, bold=True)
        
        summary_data = [
            ("전체 이상 항목", anomaly_results.get("total_anomalies_detected", 0)),
            ("고위험 이상", anomaly_results.get("high_risk_anomalies", 0)),
            ("중위험 이상", anomaly_results.get("medium_risk_anomalies", 0)),
            ("저위험 이상", anomaly_results.get("low_risk_anomalies", 0))
        ]
        
        # 요약 테이블
        headers = ["항목", "개수"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1, value=header)
            cell.style = "header"
        
        for i, (item, count) in enumerate(summary_data, start=5):
            ws.cell(row=i, column=1, value=item).style = "data"
            count_cell = ws.cell(row=i, column=2, value=count)
            count_cell.style = "data"
            
            # 위험도별 색상
            if "고위험" in item and count > 0:
                count_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
            elif "중위험" in item and count > 0:
                count_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
        
        # 상세 이상 항목
        anomaly_details = anomaly_results.get("anomaly_details", [])
        if anomaly_details:
            ws["A10"] = "상세 이상 항목"
            ws["A10"].font = Font(size=14, bold=True)
            
            # 이상 항목 테이블 헤더
            detail_headers = ["유형", "위험도", "설명", "권장사항"]
            for i, header in enumerate(detail_headers):
                cell = ws.cell(row=11, column=i+1, value=header)
                cell.style = "header"
            
            # 이상 항목 데이터
            for i, anomaly in enumerate(anomaly_details, start=12):
                ws.cell(row=i, column=1, value=anomaly.get("type", "")).style = "data"
                
                risk_level = anomaly.get("risk_level", "UNKNOWN")
                risk_cell = ws.cell(row=i, column=2, value=risk_level)
                risk_cell.style = "data"
                
                # 위험도별 색상
                if risk_level == "HIGH":
                    risk_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
                elif risk_level == "MEDIUM":
                    risk_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
                elif risk_level == "LOW":
                    risk_cell.fill = PatternFill(start_color=self.colors["info"], end_color=self.colors["info"], fill_type="solid")
                
                ws.cell(row=i, column=3, value=anomaly.get("description", "")).style = "data"
                ws.cell(row=i, column=4, value=anomaly.get("recommendation", "")).style = "data"
        
        self._adjust_column_widths(ws)
    
    def _create_action_items_sheet(self, validation_results: Dict[str, Any]):
        """Action Items 시트 생성"""
        ws = self.workbook.create_sheet("Action Items")
        self.logger.info("  📝 Action Items 시트 생성...")
        
        # 제목
        ws["A1"] = "액션 아이템 목록"
        ws["A1"].font = Font(size=16, bold=True)
        
        # 액션 아이템 생성
        action_items = []
        
        # 권장사항을 액션 아이템으로 변환
        recommendations = validation_results.get("recommendations", [])
        for i, rec in enumerate(recommendations, 1):
            action_items.append({
                "ID": f"REC_{i:03d}",
                "우선순위": "HIGH" if "개선 필요" in rec else "MEDIUM",
                "카테고리": "검증 개선",
                "설명": rec,
                "상태": "OPEN",
                "담당자": "TBD",
                "마감일": (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"),
                "생성일": datetime.now().strftime("%Y-%m-%d")
            })
        
        # 이상 항목을 액션 아이템으로 추가
        anomaly_results = validation_results.get("validation_results", {}).get("anomaly_detection", {})
        anomalies = anomaly_results.get("anomaly_details", [])
        for i, anomaly in enumerate(anomalies, len(action_items) + 1):
            action_items.append({
                "ID": f"ANO_{i:03d}",
                "우선순위": anomaly.get("risk_level", "MEDIUM"),
                "카테고리": "이상 조사",
                "설명": anomaly.get("description", ""),
                "상태": "OPEN",
                "담당자": "TBD",
                "마감일": (datetime.now() + timedelta(days=3 if anomaly.get("risk_level") == "HIGH" else 7)).strftime("%Y-%m-%d"),
                "생성일": datetime.now().strftime("%Y-%m-%d")
            })
        
        if action_items:
            # 액션 아이템 테이블
            headers = list(action_items[0].keys())
            for i, header in enumerate(headers):
                cell = ws.cell(row=3, column=i+1, value=header)
                cell.style = "header"
            
            # 액션 아이템 데이터
            for row_idx, item in enumerate(action_items, start=4):
                for col_idx, (key, value) in enumerate(item.items()):
                    cell = ws.cell(row=row_idx, column=col_idx+1, value=value)
                    cell.style = "data"
                    
                    # 우선순위별 색상
                    if key == "우선순위":
                        if value == "HIGH":
                            cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
                        elif value == "MEDIUM":
                            cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
        else:
            ws["A3"] = "생성된 액션 아이템이 없습니다."
        
        self._adjust_column_widths(ws)
    
    def _create_statistics_sheet(self, validation_results: Dict[str, Any]):
        """Detailed Statistics 시트 생성"""
        ws = self.workbook.create_sheet("Statistics")
        self.logger.info("  📈 Statistics 시트 생성...")
        
        statistics = validation_results.get("statistics", {})
        
        # 제목
        ws["A1"] = "상세 검증 통계"
        ws["A1"].font = Font(size=16, bold=True)
        
        # 시간 통계
        ws["A3"] = "검증 시간 통계"
        ws["A3"].font = Font(size=14, bold=True)
        
        time_stats = [
            ("검증 시작 시간", statistics.get("start_time", "N/A")),
            ("검증 종료 시간", statistics.get("end_time", "N/A")),
            ("총 검증 시간", f"{statistics.get('total_validation_time', 0):.1f}초")
        ]
        
        # 시간 통계 테이블
        headers = ["항목", "값"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1, value=header)
            cell.style = "header"
        
        for i, (item, value) in enumerate(time_stats, start=5):
            ws.cell(row=i, column=1, value=item).style = "data"
            ws.cell(row=i, column=2, value=str(value)).style = "data"
        
        # 처리 통계
        ws["A9"] = "처리 통계"
        ws["A9"].font = Font(size=14, bold=True)
        
        processing_stats = [
            ("전체 인보이스 수", statistics.get("total_invoices", 0)),
            ("검증된 인보이스 수", statistics.get("validated_invoices", 0)),
            ("검증 실패 수", statistics.get("failed_validations", 0)),
            ("PDF 문서 처리 수", statistics.get("pdf_documents_processed", 0)),
            ("Gate 검증 통과 수", statistics.get("gate_validations_passed", 0)),
            ("규제 준수 확인 수", statistics.get("compliance_checks_passed", 0))
        ]
        
        # 처리 통계 테이블
        headers = ["항목", "개수"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=10, column=i+1, value=header)
            cell.style = "header"
        
        for i, (item, count) in enumerate(processing_stats, start=11):
            ws.cell(row=i, column=1, value=item).style = "data"
            ws.cell(row=i, column=2, value=count).style = "data"
        
        # 모듈 통합 상태
        integration_status = validation_results.get("integration_status", {})
        if integration_status:
            ws["A18"] = "모듈 통합 상태"
            ws["A18"].font = Font(size=14, bold=True)
            
            # 통합 상태 테이블
            headers = ["모듈", "상태"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=19, column=i+1, value=header)
                cell.style = "header"
            
            for i, (module, status) in enumerate(integration_status.items(), start=20):
                ws.cell(row=i, column=1, value=module).style = "data"
                status_cell = ws.cell(row=i, column=2, value=status)
                status_cell.style = "data"
                
                # 상태별 색상
                if status == "READY":
                    status_cell.fill = PatternFill(start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid")
                elif "ERROR" in status:
                    status_cell.fill = PatternFill(start_color=self.colors["danger"], end_color=self.colors["danger"], fill_type="solid")
                elif status in ["MODULE_NOT_AVAILABLE", "DISABLED"]:
                    status_cell.fill = PatternFill(start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid")
        
        self._adjust_column_widths(ws)
    
    def _adjust_column_widths(self, ws):
        """컬럼 너비 자동 조정"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if hasattr(cell, 'value') and cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
            ws.column_dimensions[column].width = adjusted_width


def main():
    """테스트용 메인 함수"""
    # 샘플 검증 결과로 테스트
    sample_results = {
        "validation_summary": {
            "overall_status": "SUCCESS",
            "confidence_score": 0.92,
            "quality_grade": "EXCELLENT",
            "total_validation_time": 45.3
        },
        "system_results": {
            "vba_analysis": {
                "status": "SUCCESS",
                "sheets_analyzed": 31,
                "formulas_extracted": 29,
                "calculations_analyzed": 553,
                "master_data_rows": 102,
                "validation_score": 0.89,
                "validation_status": "PASS"
            },
            "python_audit": {
                "status": "SUCCESS",
                "csv_items_count": 150
            },
            "pdf_integration": {
                "status": "SUCCESS",
                "total_pdfs": 25,
                "parsed_successfully": 23,
                "parsing_failures": 2
            },
            "cross_system_integration": {
                "status": "SUCCESS"
            }
        },
        "validation_results": {
            "gate_validation": {
                "status": "SUCCESS",
                "total_gates": 14,
                "passed_gates": 13,
                "failed_gates": 1,
                "overall_pass_rate": 0.93,
                "gate_details": {}
            },
            "compliance_check": {
                "status": "SUCCESS",
                "overall_compliance_score": 1.0,
                "fanr_compliance": {"status": "NOT_REQUIRED"},
                "moiat_compliance": {"status": "COMPLIANT"},
                "dcd_compliance": {"status": "COMPLIANT"}
            },
            "anomaly_detection": {
                "status": "SUCCESS",
                "total_anomalies_detected": 2,
                "high_risk_anomalies": 0,
                "medium_risk_anomalies": 1,
                "low_risk_anomalies": 1,
                "anomaly_details": []
            }
        },
        "statistics": {
            "start_time": datetime.now(),
            "total_validation_time": 45.3
        },
        "integration_status": {
            "vba_analyzer": "READY",
            "pdf_integration": "READY",
            "audit_system": "READY"
        },
        "recommendations": ["모든 검증 항목이 기준을 충족합니다"]
    }
    
    generator = ComprehensiveExcelReportGenerator()
    report_path = generator.generate_comprehensive_report(sample_results, "test_output")
    print(f"테스트 보고서 생성: {report_path}")


if __name__ == "__main__":
    main()
