#!/usr/bin/env python3
"""
모든 컬럼 확인 스크립트
원본 INVOICE의 전체 18개 컬럼 확인
"""

import openpyxl
from pathlib import Path


def main():
    invoice_file = Path(
        rPath(__file__).parent.parent / "Data" / "DSV 202509" / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    )

    wb = openpyxl.load_workbook(invoice_file, data_only=False)
    ws = wb["MasterData"]

    print("=" * 80)
    print("Original INVOICE - All Columns in MasterData")
    print("=" * 80)

    print(f"\nTotal: {ws.max_row} rows x {ws.max_column} columns")
    print(f"\nAll Columns (1-{ws.max_column}):")

    for i in range(1, ws.max_column + 1):
        cell = ws.cell(1, i)
        col_letter = cell.column_letter
        print(f"  {i:2d}. {col_letter:2s} | {cell.value}")

    # 샘플 데이터 (Row 2)
    print(f"\nSample Data (Row 2):")
    for i in range(1, ws.max_column + 1):
        cell = ws.cell(2, i)
        col_letter = cell.column_letter
        print(f"  {col_letter:2s} ({i:2d}): {cell.value}")

    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
