#!/usr/bin/env python3
"""
Report Generator
원본 MasterData + Python 검증 결과 통합 최종 Excel 보고서 생성

Version: 2.0.0
Created: 2025-10-14
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import logging
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def generate_final_report():
    """최종 검증 보고서 생성"""

    logger.info("=" * 80)
    logger.info("Final Validation Report Generation")
    logger.info("=" * 80)

    # 1. 원본 MasterData 로드
    logger.info("\n[1] Loading original MasterData...")
    invoice_file = Path(
        rPath(__file__).parent.parent
        / "Data"
        / "DSV 202509"
        / "SCNT SHIPMENT DRAFT INVOICE (SEPT 2025)_FINAL.xlsm"
    )

    df_original = pd.read_excel(invoice_file, sheet_name="MasterData")
    logger.info(
        f"  Loaded: {len(df_original)} rows × {len(df_original.columns)} columns"
    )

    # 2. 검증 결과 로드
    logger.info("\n[2] Loading validation results...")
    out_dir = Path(__file__).parent / "out"
    csv_files = sorted(out_dir.glob("masterdata_validated_*.csv"))

    if not csv_files:
        logger.error("No validation results found!")
        return

    latest_csv = csv_files[-1]
    logger.info(f"  From: {latest_csv.name}")

    df_validation = pd.read_csv(latest_csv)
    logger.info(
        f"  Loaded: {len(df_validation)} rows × {len(df_validation.columns)} columns"
    )

    # 3. 검증 결과만 추출 (Python 추가 컬럼 9개)
    validation_cols = [
        "Validation_Status",
        "Ref_Rate_USD",
        "Python_Delta",
        "CG_Band",
        "Charge_Group",
        "Gate_Score",
        "Gate_Status",
        "PDF_Count",
        "Validation_Notes",
    ]

    df_validation_only = df_validation[validation_cols]

    # 4. 원본 MasterData + Python 검증 결합
    logger.info("\n[3] Combining original + validation...")
    df_final = pd.concat([df_original, df_validation_only], axis=1)

    logger.info(f"  Final: {len(df_final)} rows × {len(df_final.columns)} columns")
    logger.info(f"    - VBA original: {len(df_original.columns)} columns")
    logger.info(f"    - Python validation: {len(validation_cols)} columns")

    # 5. Excel 파일 생성 (조건부 서식 포함)
    logger.info("\n[4] Generating Excel report with formatting...")

    output_dir = Path(__file__).parent.parent / "Results"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"SCNT_SHIPMENT_SEPT2025_VALIDATED_{timestamp}.xlsx"

    # ExcelWriter로 생성
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Sheet 1: MasterData_Validated
        df_final.to_excel(writer, sheet_name="MasterData_Validated", index=False)

        # Sheet 2: Validation_Summary
        summary_data = {
            "Metric": [
                "Total Items",
                "Total Amount (USD)",
                "",
                "Validation Status",
                "  PASS",
                "  FAIL",
                "  REVIEW_NEEDED",
                "",
                "Charge Group",
                "  Contract",
                "  AtCost",
                "  PortalFee",
                "  Other",
                "",
                "Contract Validation",
                "  Total Contract",
                "  With Ref Rate",
                "  Coverage %",
                "",
                "COST-GUARD Distribution",
                "  PASS",
                "  WARN",
                "  HIGH",
                "  CRITICAL",
            ],
            "Value": [""] * 24,
        }

        # 통계 계산
        summary_data["Value"][0] = len(df_final)
        summary_data["Value"][1] = f"${df_final['TOTAL (USD)'].sum():,.2f}"

        status_counts = df_final["Validation_Status"].value_counts()
        summary_data["Value"][4] = status_counts.get("PASS", 0)
        summary_data["Value"][5] = status_counts.get("FAIL", 0)
        summary_data["Value"][6] = status_counts.get("REVIEW_NEEDED", 0)

        cg_counts = df_final["Charge_Group"].value_counts()
        summary_data["Value"][9] = cg_counts.get("Contract", 0)
        summary_data["Value"][10] = cg_counts.get("AtCost", 0)
        summary_data["Value"][11] = cg_counts.get("PortalFee", 0)
        summary_data["Value"][12] = cg_counts.get("Other", 0)

        contract_items = df_final[df_final["Charge_Group"] == "Contract"]
        with_ref = len(contract_items[contract_items["Ref_Rate_USD"].notna()])
        summary_data["Value"][15] = len(contract_items)
        summary_data["Value"][16] = with_ref
        summary_data["Value"][17] = (
            f"{with_ref/len(contract_items)*100:.1f}%"
            if len(contract_items) > 0
            else "0%"
        )

        validated_contract = contract_items[contract_items["CG_Band"].notna()]
        cg_band_counts = validated_contract["CG_Band"].value_counts()
        summary_data["Value"][20] = cg_band_counts.get("PASS", 0)
        summary_data["Value"][21] = cg_band_counts.get("WARN", 0)
        summary_data["Value"][22] = cg_band_counts.get("HIGH", 0)
        summary_data["Value"][23] = cg_band_counts.get("CRITICAL", 0)

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Validation_Summary", index=False)

        # Sheet 3: VBA_vs_Python
        vba_python_comparison = df_final[
            [
                "No",
                "DESCRIPTION",
                "RATE",
                "REV RATE",
                "DIFFERENCE",
                "Ref_Rate_USD",
                "Python_Delta",
                "Validation_Status",
            ]
        ].copy()

        # VBA와 Python 차이 계산
        vba_python_comparison["Rate_Diff"] = (
            vba_python_comparison["REV RATE"] - vba_python_comparison["Ref_Rate_USD"]
        )

        df_summary.to_excel(writer, sheet_name="VBA_vs_Python", index=False)

    logger.info(f"[SAVED] {output_file}")

    # 6. 조건부 서식 적용 (재오픈)
    logger.info("\n[5] Applying conditional formatting...")

    wb = openpyxl.load_workbook(output_file)
    ws = wb["MasterData_Validated"]

    # 색상 정의
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    orange_fill = PatternFill(
        start_color="FFD699", end_color="FFD699", fill_type="solid"
    )

    # Validation_Status (N열, 14번)
    ws.conditional_formatting.add(
        f"N2:N{ws.max_row}",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"N2:N{ws.max_row}",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"N2:N{ws.max_row}",
        CellIsRule(operator="equal", formula=['"REVIEW_NEEDED"'], fill=yellow_fill),
    )

    # CG_Band (Q열, 17번)
    ws.conditional_formatting.add(
        f"Q2:Q{ws.max_row}",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"Q2:Q{ws.max_row}",
        CellIsRule(operator="equal", formula=['"WARN"'], fill=yellow_fill),
    )
    ws.conditional_formatting.add(
        f"Q2:Q{ws.max_row}",
        CellIsRule(operator="equal", formula=['"HIGH"'], fill=orange_fill),
    )
    ws.conditional_formatting.add(
        f"Q2:Q{ws.max_row}",
        CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=red_fill),
    )

    # Gate_Status (T열, 20번)
    ws.conditional_formatting.add(
        f"T2:T{ws.max_row}",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"T2:T{ws.max_row}",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=red_fill),
    )

    wb.save(output_file)
    logger.info(f"  [OK] Conditional formatting applied")

    # 7. 최종 보고
    logger.info("\n" + "=" * 80)
    logger.info("[SUCCESS] Final Validation Report Generated!")
    logger.info("=" * 80)
    logger.info(f"\nOutput File: {output_file}")
    logger.info(f"\nStructure:")
    logger.info(
        f"  Sheet 1: MasterData_Validated ({len(df_final)} rows × {len(df_final.columns)} columns)"
    )
    logger.info(f"    - Columns 1-13: VBA original (No ~ DIFFERENCE)")
    logger.info(f"    - Columns 14-22: Python validation")
    logger.info(f"  Sheet 2: Validation_Summary (Statistics)")
    logger.info(f"  Sheet 3: VBA_vs_Python (Comparison)")
    logger.info(f"\nConditional Formatting:")
    logger.info(f"  - Validation_Status (Green/Red/Yellow)")
    logger.info(f"  - CG_Band (Green/Yellow/Orange/Red)")
    logger.info(f"  - Gate_Status (Green/Red)")

    print("\n" + "=" * 80)
    print(f"[REPORT GENERATED] {output_file.name}")
    print("=" * 80)

    return output_file


if __name__ == "__main__":
    generate_final_report()
