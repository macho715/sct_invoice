"""Enhanced Excel Report Generator with Anomaly Detection and Risk Scoring Columns

기존 Excel 포맷을 유지하면서 새로운 열(Anomaly Score, Risk Score, Risk Level, Anomaly Details)을 추가합니다.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import json
import logging
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)


class EnhancedExcelReportGenerator:
    """Enhanced Excel Report Generator"""

    def __init__(self):
        self.red_fill = PatternFill(start_color="FFCCCC", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFFCC", fill_type="solid")
        self.green_fill = PatternFill(start_color="CCFFCC", fill_type="solid")
        self.header_font = Font(bold=True, size=11)
        self.normal_font = Font(size=10)

    def create_enhanced_excel_report(
        self,
        validation_results: pd.DataFrame,
        output_path: str,
        preserve_formatting: bool = True,
    ) -> str:
        """
        기존 Excel 포맷을 유지하면서 새로운 열 추가

        Args:
            validation_results: 검증 결과 DataFrame
            output_path: 출력 Excel 파일 경로
            preserve_formatting: 기존 서식 유지 여부

        Returns:
            생성된 파일 경로
        """
        logger.info(f"Creating enhanced Excel report: {output_path}")

        # 1. 데이터 전처리 및 열 매핑
        processed_df = self._prepare_data(validation_results)

        # 2. Excel 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enhanced_Validation_Results"

        # 3. 헤더 작성
        self._write_headers(ws, processed_df.columns.tolist())

        # 4. 데이터 작성
        self._write_data(ws, processed_df)

        # 5. 조건부 서식 적용
        if preserve_formatting:
            self._apply_conditional_formatting(ws, processed_df)

        # 6. 열 너비 조정
        self._adjust_column_widths(ws)

        # 7. 파일 저장
        wb.save(output_path)
        logger.info(f"Enhanced Excel report saved: {output_path}")

        return output_path

    def _prepare_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """검증 결과 데이터 전처리"""

        # 기존 열 순서 정의 (검증 결과의 실제 열 순서)
        existing_columns = [
            "s_no",
            "sheet_name",
            "description",
            "rate_source",
            "unit_rate",
            "quantity",
            "total_usd",
            "status",
            "flag",
            "delta_pct",
            "cg_band",
            "charge_group",
            "issues",
            "tolerance",
            "ref_rate_usd",
            "doc_aed",
            "gates",
            "gate_status",
            "gate_score",
            "gate_fails",
        ]

        # 새로 추가할 열들
        new_columns = [
            "anomaly_score",  # Anomaly Score (0-100)
            "risk_score",  # Risk Score (0-1.0)
            "risk_level",  # Risk Level (LOW/MEDIUM/HIGH/CRITICAL)
            "anomaly_details",  # Anomaly Details (JSON string)
            "risk_triggered",  # Risk Triggered (boolean)
            "risk_components",  # Risk Components (JSON string)
        ]

        # 기존 열들 중 존재하는 것만 선택
        available_existing = [col for col in existing_columns if col in df.columns]

        # 새 열들을 기존 열 사이에 삽입
        final_columns = []

        # 1. 기본 정보 열들
        basic_cols = [
            "s_no",
            "sheet_name",
            "description",
            "rate_source",
            "unit_rate",
            "quantity",
            "total_usd",
        ]
        final_columns.extend([col for col in basic_cols if col in df.columns])

        # 2. 검증 상태 열들
        status_cols = ["status", "flag", "delta_pct", "cg_band", "charge_group"]
        final_columns.extend([col for col in status_cols if col in df.columns])

        # 3. Gate 관련 열들
        gate_cols = ["gates", "gate_status", "gate_score", "gate_fails"]
        final_columns.extend([col for col in gate_cols if col in df.columns])

        # 3.5. PDF 확인 열들 (새로 추가)
        pdf_verification_cols = ["PDF_Verification_Status", "PDF_Verification_Details"]
        final_columns.extend(
            [col for col in pdf_verification_cols if col in df.columns]
        )

        # 4. 새로운 Anomaly & Risk 열들 삽입
        final_columns.extend(new_columns)

        # 5. 나머지 열들
        remaining_cols = [col for col in df.columns if col not in final_columns]
        final_columns.extend(remaining_cols)

        # 데이터 전처리
        processed_df = df.copy()

        # Anomaly Score 추출 (anomaly_detection JSON에서)
        processed_df["anomaly_score"] = processed_df["anomaly_detection"].apply(
            self._extract_anomaly_score
        )

        # Risk Score는 이미 있음
        processed_df["risk_score"] = processed_df["risk_score"].fillna(0.0)

        # Risk Level 계산
        processed_df["risk_level"] = processed_df["risk_score"].apply(
            self._calculate_risk_level
        )

        # Anomaly Details 정리
        processed_df["anomaly_details"] = processed_df["anomaly_detection"].apply(
            self._format_anomaly_details
        )

        # Risk Components 정리
        processed_df["risk_components"] = processed_df["risk_components"].apply(
            self._format_risk_components
        )

        # Risk Triggered는 이미 있음
        processed_df["risk_triggered"] = processed_df["risk_triggered"].fillna(False)

        # 최종 열 순서로 재정렬
        final_df = processed_df[final_columns]

        logger.info(
            f"Processed data: {len(final_df)} rows, {len(final_df.columns)} columns"
        )
        return final_df

    def _extract_anomaly_score(self, anomaly_data: Any) -> float:
        """anomaly_detection JSON에서 score 추출"""
        try:
            if isinstance(anomaly_data, str):
                data = json.loads(anomaly_data)
            elif isinstance(anomaly_data, dict):
                data = anomaly_data
            else:
                return 0.0

            return float(data.get("score", 0.0))
        except (json.JSONDecodeError, TypeError, ValueError):
            return 0.0

    def _calculate_risk_level(self, risk_score: float) -> str:
        """Risk Score를 기반으로 Risk Level 계산"""
        if risk_score >= 0.8:
            return "CRITICAL"
        elif risk_score >= 0.6:
            return "HIGH"
        elif risk_score >= 0.3:
            return "MEDIUM"
        else:
            return "LOW"

    def _format_anomaly_details(self, anomaly_data: Any) -> str:
        """anomaly_detection 데이터를 읽기 쉬운 형태로 포맷"""
        try:
            if isinstance(anomaly_data, str):
                data = json.loads(anomaly_data)
            elif isinstance(anomaly_data, dict):
                data = anomaly_data
            else:
                return "N/A"

            details = []
            if "model" in data:
                details.append(f"Model: {data['model']}")
            if "risk_level" in data:
                details.append(f"Risk: {data['risk_level']}")
            if "flagged" in data:
                details.append(f"Flagged: {data['flagged']}")
            if "details" in data and isinstance(data["details"], dict):
                if "reference" in data["details"]:
                    details.append(f"Ref: {data['details']['reference']}")
                if "difference" in data["details"]:
                    details.append(f"Diff: {data['details']['difference']}")

            return " | ".join(details) if details else "N/A"

        except (json.JSONDecodeError, TypeError, ValueError):
            return "N/A"

    def _format_risk_components(self, risk_components: Any) -> str:
        """risk_components 데이터를 읽기 쉬운 형태로 포맷"""
        try:
            if isinstance(risk_components, str):
                data = json.loads(risk_components)
            elif isinstance(risk_components, dict):
                data = risk_components
            else:
                return "N/A"

            components = []
            for key, value in data.items():
                if isinstance(value, (int, float)):
                    components.append(f"{key}: {value:.2f}")
                else:
                    components.append(f"{key}: {value}")

            return " | ".join(components) if components else "N/A"

        except (json.JSONDecodeError, TypeError, ValueError):
            return "N/A"

    def _write_headers(self, ws, columns: list):
        """헤더 행 작성"""
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_data(self, ws, df: pd.DataFrame):
        """데이터 행 작성"""
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                # NaN 값을 빈 문자열로 변환
                if pd.isna(value):
                    value = ""
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)

                ws.cell(row=row_idx, column=col_idx, value=value)

    def _apply_conditional_formatting(self, ws, df: pd.DataFrame):
        """조건부 서식 적용"""

        # Anomaly Score 열 찾기
        anomaly_score_col = None
        risk_score_col = None
        risk_level_col = None
        pdf_verification_status_col = None

        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == "anomaly_score":
                anomaly_score_col = col_idx
            elif col_name == "risk_score":
                risk_score_col = col_idx
            elif col_name == "risk_level":
                risk_level_col = col_idx
            elif col_name == "PDF_Verification_Status":
                pdf_verification_status_col = col_idx

        # Anomaly Score 조건부 서식 (>80 빨강, 60-80 노랑, <60 초록)
        if anomaly_score_col:
            col_letter = openpyxl.utils.get_column_letter(anomaly_score_col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="greaterThan", formula=["80"], fill=self.red_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(
                    operator="between", formula=["60", "80"], fill=self.yellow_fill
                ),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="lessThan", formula=["60"], fill=self.green_fill),
            )

        # Risk Score 조건부 서식 (>0.8 빨강, 0.5-0.8 노랑, <0.5 초록)
        if risk_score_col:
            col_letter = openpyxl.utils.get_column_letter(risk_score_col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="greaterThan", formula=["0.8"], fill=self.red_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(
                    operator="between", formula=["0.5", "0.8"], fill=self.yellow_fill
                ),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="lessThan", formula=["0.5"], fill=self.green_fill),
            )

        # Risk Level 색상 코딩
        if risk_level_col:
            col_letter = openpyxl.utils.get_column_letter(risk_level_col)
            # CRITICAL: 빨강
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=["CRITICAL"], fill=self.red_fill),
            )
            # HIGH: 주황색
            high_fill = PatternFill(start_color="FFA500", fill_type="solid")
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=["HIGH"], fill=high_fill),
            )
            # MEDIUM: 노랑
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=["MEDIUM"], fill=self.yellow_fill),
            )
            # LOW: 초록
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=["LOW"], fill=self.green_fill),
            )

        # PDF Verification Status 색상 코딩
        if pdf_verification_status_col:
            col_letter = openpyxl.utils.get_column_letter(pdf_verification_status_col)
            # "일치": 초록
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=['"일치"'], fill=self.green_fill),
            )
            # "PDF 확인됨": 초록
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(
                    operator="equal", formula=['"PDF 확인됨"'], fill=self.green_fill
                ),
            )
            # "불일치": 빨강
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=['"불일치"'], fill=self.red_fill),
            )
            # "PDF 없음": 노랑
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(
                    operator="equal", formula=['"PDF 없음"'], fill=self.yellow_fill
                ),
            )
            # "검토 필요": 노랑
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(
                    operator="equal", formula=['"검토 필요"'], fill=self.yellow_fill
                ),
            )
            # "추출 실패": 주황
            orange_fill = PatternFill(start_color="FFA500", fill_type="solid")
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=['"추출 실패"'], fill=orange_fill),
            )
            # "확인 불필요": 회색
            gray_fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="equal", formula=['"확인 불필요"'], fill=gray_fill),
            )

    def _adjust_column_widths(self, ws):
        """열 너비 자동 조정"""
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # 최대 너비 제한 (50자)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def create_enhanced_excel_report(
    validation_results: pd.DataFrame, output_path: str, preserve_formatting: bool = True
) -> str:
    """
    Enhanced Excel Report 생성 (편의 함수)

    Args:
        validation_results: 검증 결과 DataFrame
        output_path: 출력 Excel 파일 경로
        preserve_formatting: 기존 서식 유지 여부

    Returns:
        생성된 파일 경로
    """
    generator = EnhancedExcelReportGenerator()
    return generator.create_enhanced_excel_report(
        validation_results, output_path, preserve_formatting
    )


if __name__ == "__main__":
    # 테스트 코드
    import pandas as pd

    # 샘플 데이터 생성
    sample_data = {
        "s_no": [1, 2, 3],
        "description": ["Test Item 1", "Test Item 2", "Test Item 3"],
        "anomaly_score": [85.5, 15.2, 72.3],
        "risk_score": [0.85, 0.25, 0.65],
        "risk_level": ["CRITICAL", "LOW", "MEDIUM"],
    }

    df = pd.DataFrame(sample_data)

    # 보고서 생성
    output_path = "test_enhanced_report.xlsx"
    create_enhanced_excel_report(df, output_path)
    print(f"Test report created: {output_path}")
