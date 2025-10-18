#!/usr/bin/env python3
"""
CRITICAL 16건 분석 및 4개 버킷 분류
A. 단거리(≤10km), B. 특수차량(HAZMAT/CICPA), C. 과소청구(delta<0), D. 순수 과다청구
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def classify_critical_bucket(row):
    """CRITICAL 항목을 4개 버킷으로 분류"""
    # A. 단거리 (distance_km ≤ 10)
    if pd.notna(row.get("distance_km")) and row["distance_km"] <= 10:
        return "A. SHORT_RUN"

    # B. 특수차량 (HAZMAT/CICPA)
    vehicle = str(row.get("vehicle", "")).upper()
    if "HAZMAT" in vehicle or "CICPA" in vehicle:
        return "B. SPECIAL_VEHICLE"

    # C. 과소청구 (delta_pct < 0)
    if pd.notna(row.get("delta_pct")) and row["delta_pct"] < 0:
        return "C. UNDER_CHARGE"

    # D. 순수 과다청구
    return "D. OVER_CHARGE"


def main():
    print("=" * 80)
    print("CRITICAL 16 Items Analysis and Bucket Classification")
    print("=" * 80)

    # Load latest result
    results_dir = Path(__file__).parent / "Results" / "Sept_2025" / "CSV"
    latest_csv = results_dir / "domestic_sept_2025_result_20251013_013624.csv"

    if not latest_csv.exists():
        print(f"[ERROR] Result file not found: {latest_csv}")
        return

    print(f"\nLoading: {latest_csv.name}")
    df = pd.read_csv(latest_csv)
    print(f"  Total items: {len(df)}")

    # Filter CRITICAL only
    critical_df = df[df["cg_band"] == "CRITICAL"].copy()
    print(f"  CRITICAL items: {len(critical_df)}")

    if len(critical_df) == 0:
        print("[WARN] No CRITICAL items found!")
        return

    # Classify into buckets
    print("\nClassifying into buckets...")
    critical_df["bucket"] = critical_df.apply(classify_critical_bucket, axis=1)
    critical_df["delta_sign"] = critical_df["delta_pct"].apply(
        lambda x: "UNDER" if pd.notna(x) and x < 0 else "OVER"
    )
    critical_df["is_short"] = critical_df["distance_km"].apply(
        lambda x: "Y" if pd.notna(x) and x <= 10 else "N"
    )

    # Bucket summary
    bucket_counts = critical_df["bucket"].value_counts()
    print("\nBucket Distribution:")
    for bucket, count in bucket_counts.items():
        print(f"  {bucket}: {count} items")

    # Prepare detailed analysis
    analysis_cols = [
        "shipment_ref",
        "origin",
        "origin_norm",
        "destination",
        "destination_norm",
        "vehicle",
        "vehicle_norm",
        "unit",
        "distance_km",
        "rate_usd",
        "ref_rate_usd",
        "delta_pct",
        "delta_sign",
        "is_short",
        "bucket",
        "similarity",
        "ref_method",
        "flags",
    ]

    # Select available columns
    available_cols = [col for col in analysis_cols if col in critical_df.columns]
    triage_df = critical_df[available_cols].copy()

    # Sort by bucket
    triage_df = triage_df.sort_values(["bucket", "delta_pct"], ascending=[True, False])

    # Create Excel output
    output_dir = Path(__file__).parent / "Results" / "Sept_2025" / "Reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_excel = output_dir / "CRITICAL_16_TRIAGE.xlsx"

    print(f"\nCreating Excel report: {output_excel.name}")
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["CRITICAL 16 Items - Bucket Classification"],
        [""],
        ["Bucket", "Count", "Description"],
        [
            "A. SHORT_RUN",
            bucket_counts.get("A. SHORT_RUN", 0),
            "Distance <= 10km (Min-Fare applicable)",
        ],
        [
            "B. SPECIAL_VEHICLE",
            bucket_counts.get("B. SPECIAL_VEHICLE", 0),
            "HAZMAT/CICPA (Rate multiplier applicable)",
        ],
        [
            "C. UNDER_CHARGE",
            bucket_counts.get("C. UNDER_CHARGE", 0),
            "Negative delta (PENDING_REVIEW buffer)",
        ],
        [
            "D. OVER_CHARGE",
            bucket_counts.get("D. OVER_CHARGE", 0),
            "Pure overcharge (Requires manual review)",
        ],
        [""],
        ["Total CRITICAL", len(critical_df), ""],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # Format header
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws_summary[1]:
        cell.font = Font(bold=True, size=14)

    for cell in ws_summary[3]:
        cell.fill = header_fill
        cell.font = header_font

    # Adjust column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 50

    # Sheet 2: Detailed Triage
    ws_detail = wb.create_sheet("Detailed_Triage")

    for r in dataframe_to_rows(triage_df, index=False, header=True):
        ws_detail.append(r)

    # Format header
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Freeze header
    ws_detail.freeze_panes = "A2"

    # Sheet 3: By Bucket
    for bucket_name in [
        "A. SHORT_RUN",
        "B. SPECIAL_VEHICLE",
        "C. UNDER_CHARGE",
        "D. OVER_CHARGE",
    ]:
        bucket_data = triage_df[triage_df["bucket"] == bucket_name]
        if len(bucket_data) > 0:
            sheet_name = bucket_name.split(".")[1].strip()[
                :31
            ]  # Excel sheet name limit
            ws_bucket = wb.create_sheet(sheet_name)

            for r in dataframe_to_rows(bucket_data, index=False, header=True):
                ws_bucket.append(r)

            # Format header
            for cell in ws_bucket[1]:
                cell.fill = header_fill
                cell.font = header_font

            ws_bucket.freeze_panes = "A2"

    # Save
    wb.save(output_excel)
    print(f"[OK] Excel report saved: {output_excel}")

    # Print detailed summary
    print("\n" + "=" * 80)
    print("Bucket Details")
    print("=" * 80)

    for bucket_name in [
        "A. SHORT_RUN",
        "B. SPECIAL_VEHICLE",
        "C. UNDER_CHARGE",
        "D. OVER_CHARGE",
    ]:
        bucket_data = triage_df[triage_df["bucket"] == bucket_name]
        if len(bucket_data) > 0:
            print(f"\n{bucket_name} ({len(bucket_data)} items):")
            for idx, row in bucket_data.iterrows():
                print(
                    f"  - {row['shipment_ref']}: {row.get('origin_norm', '')} → {row.get('destination_norm', '')}"
                )
                print(
                    f"    Vehicle: {row.get('vehicle_norm', '')}, Distance: {row.get('distance_km', 'N/A')}km, Delta: {row.get('delta_pct', 'N/A'):.1f}%"
                )

    print("\n" + "=" * 80)
    print("Next Steps:")
    print("=" * 80)
    print("1. Review CRITICAL_16_TRIAGE.xlsx")
    print("2. Extract NormalizationMap aliases (Step 2)")
    print("3. Add Min-Fare lanes to ApprovedLaneMap (Step 4)")
    print("4. Add HAZMAT/CICPA lanes to ApprovedLaneMap (Step 5)")
    print("5. Re-run validation")
    print("=" * 80)


if __name__ == "__main__":
    main()
