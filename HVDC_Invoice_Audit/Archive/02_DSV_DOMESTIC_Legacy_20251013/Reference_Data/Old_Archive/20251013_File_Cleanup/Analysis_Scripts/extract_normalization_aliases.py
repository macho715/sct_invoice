#!/usr/bin/env python3
"""
CRITICAL 항목에서 NormalizationMap 별칭 후보 추출
origin/destination 정규화가 안 된 항목을 찾아 상위 10개 추천
"""

import pandas as pd
from pathlib import Path
import re


def _tok(s: str):
    """토큰 추출"""
    return set(re.findall(r"[A-Za-z0-9]+", str(s).upper()))


def token_set_sim(a: str, b: str) -> float:
    """Token-set similarity"""
    A, B = _tok(a), _tok(b)
    if not A or not B:
        return 0.0
    return len(A & B) / len(A | B)


def find_best_canonical(
    raw_place: str, canonical_set: set, min_sim: float = 0.60
) -> tuple:
    """가장 유사한 Canonical 찾기"""
    best_match = None
    best_score = 0.0

    for canonical in canonical_set:
        score = token_set_sim(raw_place, canonical)
        if score > best_score:
            best_score = score
            best_match = canonical

    if best_score >= min_sim:
        return best_match, best_score
    return None, 0.0


def main():
    print("=" * 80)
    print("NormalizationMap Alias Extraction (Top 10)")
    print("=" * 80)

    # Load latest result
    results_dir = Path(__file__).parent / "Results" / "Sept_2025" / "CSV"
    latest_csv = results_dir / "domestic_sept_2025_result_20251013_013624.csv"

    print(f"\nLoading: {latest_csv.name}")
    df = pd.read_csv(latest_csv)

    # Filter CRITICAL
    critical_df = df[df["cg_band"] == "CRITICAL"].copy()
    print(f"  CRITICAL items: {len(critical_df)}")

    # Check normalization status
    unmapped_origins = []
    unmapped_destinations = []

    for _, row in critical_df.iterrows():
        origin = str(row.get("origin", "")).strip()
        origin_norm = str(row.get("origin_norm", "")).strip()
        dest = str(row.get("destination", "")).strip()
        dest_norm = str(row.get("destination_norm", "")).strip()

        # Check if normalization happened
        if origin and origin_norm:
            if origin.upper() != origin_norm.upper():
                # Normalized (already mapped)
                pass
            else:
                # Not normalized (needs alias)
                unmapped_origins.append(origin)

        if dest and dest_norm:
            if dest.upper() != dest_norm.upper():
                # Normalized (already mapped)
                pass
            else:
                # Not normalized (needs alias)
                unmapped_destinations.append(dest)

    # Count frequency
    origin_freq = pd.Series(unmapped_origins).value_counts()
    dest_freq = pd.Series(unmapped_destinations).value_counts()

    print(f"\nUnmapped Origins: {len(origin_freq)}")
    print(f"Unmapped Destinations: {len(dest_freq)}")

    # Get top candidates
    all_places = list(origin_freq.head(10).index) + list(dest_freq.head(10).index)
    all_places = list(set(all_places))  # Remove duplicates

    # Extract existing canonical places from results
    canonical_origins = set(df["origin_norm"].dropna().unique())
    canonical_dests = set(df["destination_norm"].dropna().unique())
    canonical_all = canonical_origins | canonical_dests

    # Remove exact matches
    canonical_all = {c.strip() for c in canonical_all if c.strip()}

    print(f"\nExisting Canonical places: {len(canonical_all)}")

    # Generate recommendations
    print("\n" + "=" * 80)
    print("Top 10 Alias Recommendations")
    print("=" * 80)

    recommendations = []

    for raw_place in all_places[:15]:  # Check top 15 to get 10 valid
        if len(recommendations) >= 10:
            break

        # Skip if already in canonical
        if raw_place.upper() in {c.upper() for c in canonical_all}:
            continue

        # Find best match
        best_canonical, score = find_best_canonical(raw_place, canonical_all)

        if best_canonical and score >= 0.60:
            recommendations.append(
                {
                    "raw_place": raw_place,
                    "normalized": best_canonical,
                    "similarity": round(score, 3),
                    "source": "auto_matched",
                }
            )
            print(
                f"{len(recommendations):2d}. {raw_place:30s} → {best_canonical:30s} (sim: {score:.3f})"
            )
        else:
            # Manual review needed
            recommendations.append(
                {
                    "raw_place": raw_place,
                    "normalized": raw_place,  # Keep as-is for now
                    "similarity": 0.0,
                    "source": "manual_review_needed",
                }
            )
            print(
                f"{len(recommendations):2d}. {raw_place:30s} → [MANUAL REVIEW NEEDED]"
            )

    # Save to CSV
    output_dir = Path(__file__).parent / "Results" / "Sept_2025" / "Reports"
    output_csv = output_dir / "NORMALIZATION_ALIASES_TOP10.csv"

    alias_df = pd.DataFrame(recommendations)
    alias_df.to_csv(output_csv, index=False)

    print(f"\n[OK] Aliases saved to: {output_csv}")

    # Create Excel template for manual addition
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "NormalizationMap_Add"

    # Header
    ws.append(["raw_place", "normalized", "similarity", "source", "notes"])

    # Format header
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for rec in recommendations:
        ws.append(
            [
                rec["raw_place"],
                rec["normalized"],
                rec["similarity"],
                rec["source"],
                (
                    "Review and approve"
                    if rec["source"] == "manual_review_needed"
                    else "Auto-suggested"
                ),
            ]
        )

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30

    output_excel = output_dir / "NORMALIZATION_ALIASES_TOP10.xlsx"
    wb.save(output_excel)

    print(f"[OK] Excel template saved to: {output_excel}")

    print("\n" + "=" * 80)
    print("Next Steps:")
    print("=" * 80)
    print("1. Review NORMALIZATION_ALIASES_TOP10.xlsx")
    print("2. Manually verify/adjust 'normalized' column")
    print(
        "3. Copy approved rows to DOMESTIC_with_distances.xlsx → NormalizationMap sheet"
    )
    print("4. Format: raw_place | normalized")
    print("=" * 80)


if __name__ == "__main__":
    main()
