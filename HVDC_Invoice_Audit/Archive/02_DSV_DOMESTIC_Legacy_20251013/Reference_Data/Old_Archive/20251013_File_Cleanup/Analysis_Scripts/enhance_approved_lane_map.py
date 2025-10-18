#!/usr/bin/env python3
"""
ApprovedLaneMap 보강 스크립트
1. 현재 100 lanes 확인
2. Min-Fare 레인 추가 (단거리 ≤10km)
3. HAZMAT/CICPA 특수요율 레인 추가
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl import load_workbook


def create_min_fare_lanes():
    """Min-Fare 레인 생성 (≤10km 구간)"""
    min_fare_lanes = [
        {
            "origin": "DSV Mussafah Yard",
            "destination": "Same-City-Short",
            "vehicle": "FLATBED",
            "unit": "per truck",
            "median_rate_usd": 200.0,
            "median_distance_km": 5.0,
            "lane_id": "MIN_FARE_FB",
            "notes": "Min fare for short-run ≤10km",
        },
        {
            "origin": "DSV Mussafah Yard",
            "destination": "Same-City-Short",
            "vehicle": "LOWBED",
            "unit": "per truck",
            "median_rate_usd": 600.0,
            "median_distance_km": 5.0,
            "lane_id": "MIN_FARE_LB",
            "notes": "Min fare for short-run ≤10km",
        },
        {
            "origin": "DSV Mussafah Yard",
            "destination": "Same-City-Short",
            "vehicle": "3 TON PU",
            "unit": "per truck",
            "median_rate_usd": 150.0,
            "median_distance_km": 5.0,
            "lane_id": "MIN_FARE_3T",
            "notes": "Min fare for short-run ≤10km",
        },
        {
            "origin": "DSV Mussafah Yard",
            "destination": "Same-City-Short",
            "vehicle": "7 TON PU",
            "unit": "per truck",
            "median_rate_usd": 200.0,
            "median_distance_km": 5.0,
            "lane_id": "MIN_FARE_7T",
            "notes": "Min fare for short-run ≤10km",
        },
    ]

    return pd.DataFrame(min_fare_lanes)


def create_special_vehicle_lanes(base_lanes_df):
    """기존 레인에 HAZMAT/CICPA 버전 추가"""
    special_lanes = []

    # 주요 레인만 선택 (상위 10개 정도)
    if len(base_lanes_df) > 0:
        top_lanes = base_lanes_df.head(10)

        for _, lane in top_lanes.iterrows():
            base_rate = lane["median_rate_usd"]

            # HAZMAT version (×1.15)
            hazmat_lane = lane.copy()
            hazmat_lane["vehicle"] = f"{lane['vehicle']} HAZMAT"
            hazmat_lane["median_rate_usd"] = round(base_rate * 1.15, 2)
            hazmat_lane["lane_id"] = f"{lane.get('lane_id', 'LANE')}_HAZMAT"
            hazmat_lane["notes"] = (
                f"HAZMAT surcharge (×1.15) applied to {lane['vehicle']}"
            )
            special_lanes.append(hazmat_lane)

            # CICPA version (×1.08)
            cicpa_lane = lane.copy()
            cicpa_lane["vehicle"] = f"{lane['vehicle']} CICPA"
            cicpa_lane["median_rate_usd"] = round(base_rate * 1.08, 2)
            cicpa_lane["lane_id"] = f"{lane.get('lane_id', 'LANE')}_CICPA"
            cicpa_lane["notes"] = (
                f"CICPA surcharge (×1.08) applied to {lane['vehicle']}"
            )
            special_lanes.append(cicpa_lane)

    return pd.DataFrame(special_lanes) if special_lanes else pd.DataFrame()


def main():
    print("=" * 80)
    print("ApprovedLaneMap Enhancement")
    print("=" * 80)

    # File path
    data_dir = Path(__file__).parent
    excel_file = data_dir / "DOMESTIC_with_distances.xlsx"

    if not excel_file.exists():
        print(f"[ERROR] File not found: {excel_file}")
        return

    print(f"\nLoading: {excel_file.name}")

    # Load workbook
    try:
        wb = load_workbook(excel_file)
        print(f"  Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"[ERROR] Failed to load workbook: {e}")
        return

    # Check if ApprovedLaneMap exists
    if "ApprovedLaneMap" not in wb.sheetnames:
        print("[ERROR] ApprovedLaneMap sheet not found!")
        return

    # Load ApprovedLaneMap
    approved_df = pd.read_excel(excel_file, sheet_name="ApprovedLaneMap")
    print(f"\nCurrent ApprovedLaneMap:")
    print(f"  Total lanes: {len(approved_df)}")
    print(f"  Columns: {list(approved_df.columns)}")

    # Backup original
    output_dir = Path(__file__).parent / "Results" / "Sept_2025" / "Reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    backup_file = (
        output_dir
        / f"ApprovedLaneMap_BACKUP_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    approved_df.to_excel(backup_file, index=False)
    print(f"  Backup saved: {backup_file.name}")

    # Create Min-Fare lanes
    print("\nCreating Min-Fare lanes...")
    min_fare_df = create_min_fare_lanes()
    print(f"  Min-Fare lanes created: {len(min_fare_df)}")

    # Create special vehicle lanes
    print("\nCreating HAZMAT/CICPA lanes...")
    special_df = create_special_vehicle_lanes(approved_df)
    print(f"  Special vehicle lanes created: {len(special_df)}")

    # Combine all
    enhanced_df = pd.concat([approved_df, min_fare_df, special_df], ignore_index=True)

    # Generate key column if not exists
    if "key" not in enhanced_df.columns:
        enhanced_df["key"] = (
            enhanced_df["origin"].str.strip().str.upper()
            + "||"
            + enhanced_df["destination"].str.strip().str.upper()
            + "||"
            + enhanced_df["vehicle"].str.strip().str.upper()
            + "||"
            + enhanced_df["unit"].str.strip().str.lower()
        )

    # Remove duplicates (keep last)
    enhanced_df = enhanced_df.drop_duplicates(subset=["key"], keep="last")

    print(f"\nEnhanced ApprovedLaneMap:")
    print(f"  Total lanes: {len(enhanced_df)} (was {len(approved_df)})")
    print(f"  Added: +{len(enhanced_df) - len(approved_df)} lanes")

    # Save enhanced version
    enhanced_file = output_dir / "ApprovedLaneMap_ENHANCED.xlsx"
    enhanced_df.to_excel(enhanced_file, index=False)
    print(f"\n[OK] Enhanced ApprovedLaneMap saved: {enhanced_file}")

    # Create summary
    print("\n" + "=" * 80)
    print("Enhancement Summary")
    print("=" * 80)
    print(f"Original lanes:      {len(approved_df)}")
    print(f"Min-Fare lanes:      +{len(min_fare_df)}")
    print(f"Special lanes:       +{len(special_df)}")
    print(f"Total (deduplicated): {len(enhanced_df)}")

    # Show sample Min-Fare lanes
    print("\nMin-Fare Lanes (sample):")
    print(
        min_fare_df[
            ["origin", "destination", "vehicle", "median_rate_usd", "lane_id"]
        ].to_string(index=False)
    )

    # Show sample Special lanes
    if len(special_df) > 0:
        print("\nSpecial Vehicle Lanes (sample):")
        print(
            special_df.head(5)[
                ["origin", "destination", "vehicle", "median_rate_usd", "lane_id"]
            ].to_string(index=False)
        )

    print("\n" + "=" * 80)
    print("Next Steps:")
    print("=" * 80)
    print("1. Review ApprovedLaneMap_ENHANCED.xlsx")
    print("2. MANUAL ACTION: Copy enhanced data back to DOMESTIC_with_distances.xlsx")
    print("   - Open DOMESTIC_with_distances.xlsx")
    print("   - Go to ApprovedLaneMap sheet")
    print("   - Replace with data from ApprovedLaneMap_ENHANCED.xlsx")
    print("   - Save the file")
    print("3. Run Step 6: Re-validation")
    print("=" * 80)


if __name__ == "__main__":
    main()
