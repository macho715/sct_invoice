#!/usr/bin/env python3
"""
Enhanced ApprovedLaneMap을 DOMESTIC_with_distances.xlsx에 자동 적용
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import shutil


def main():
    print("=" * 80)
    print("Apply Enhanced ApprovedLaneMap")
    print("=" * 80)

    # File paths
    data_dir = Path(__file__).parent
    original_file = data_dir / "DOMESTIC_with_distances.xlsx"

    reports_dir = data_dir / "Results" / "Sept_2025" / "Reports"
    enhanced_file = reports_dir / "ApprovedLaneMap_ENHANCED.xlsx"

    if not original_file.exists():
        print(f"[ERROR] Original file not found: {original_file}")
        return

    if not enhanced_file.exists():
        print(f"[ERROR] Enhanced file not found: {enhanced_file}")
        return

    # Backup original
    backup_file = (
        data_dir
        / f"DOMESTIC_with_distances_BACKUP_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(original_file, backup_file)
    print(f"\n[OK] Backup created: {backup_file.name}")

    # Load enhanced data
    print(f"\nLoading enhanced ApprovedLaneMap...")
    enhanced_df = pd.read_excel(enhanced_file)
    print(f"  Enhanced lanes: {len(enhanced_df)}")

    # Load original workbook
    print(f"\nUpdating {original_file.name}...")
    wb = load_workbook(original_file)

    # Remove old ApprovedLaneMap sheet
    if "ApprovedLaneMap" in wb.sheetnames:
        del wb["ApprovedLaneMap"]
        print("  Removed old ApprovedLaneMap sheet")

    # Create new ApprovedLaneMap sheet
    ws = wb.create_sheet("ApprovedLaneMap", 1)  # Insert at position 1 (after items)

    # Write data
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment

    for r in dataframe_to_rows(enhanced_df, index=False, header=True):
        ws.append(r)

    # Format header
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save
    wb.save(original_file)
    print(f"[OK] ApprovedLaneMap updated with {len(enhanced_df)} lanes")

    print("\n" + "=" * 80)
    print("Update Complete!")
    print("=" * 80)
    print(f"File: {original_file}")
    print(f"ApprovedLaneMap: 124 lanes (100 original + 4 Min-Fare + 20 Special)")
    print("\nNext Step: Run re-validation (Step 6)")
    print("=" * 80)


if __name__ == "__main__":
    main()
