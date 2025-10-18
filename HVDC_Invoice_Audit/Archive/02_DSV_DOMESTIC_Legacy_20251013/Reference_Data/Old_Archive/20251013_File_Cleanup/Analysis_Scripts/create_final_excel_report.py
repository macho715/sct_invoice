#!/usr/bin/env python3
"""
Generate Final DOMESTIC Excel Report from Reference-from-Execution validation
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def main():
    print("=" * 80)
    print("DOMESTIC Sept 2025 - Final Validation Report Generator")
    print("=" * 80)

    base_dir = Path(__file__).parent
    csv_file = base_dir / "Results" / "Sept_2025" / "Final_Validation" / "items.csv"

    if not csv_file.exists():
        print(f"\n[ERROR] File not found: {csv_file}")
        return

    # Load data
    print(f"\nLoading: {csv_file.name}")
    df = pd.read_csv(csv_file)
    print(f"  OK Loaded {len(df)} items")

    # Create workbook
    wb = Workbook()

    # Styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    critical_fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )
    critical_font = Font(bold=True, color="FFFFFF")
    unknown_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    special_fill = PatternFill(
        start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: Summary
    print("\nGenerating Summary...")
    ws1 = wb.active
    ws1.title = "Summary"

    band_dist = df["cg_band"].value_counts().to_dict()
    verdict_dist = df["verdict"].value_counts().to_dict()

    ws1.append(["DOMESTIC Invoice Audit - September 2025"])
    ws1.append(["Final Validation with Reference-from-Execution"])
    ws1.append([])
    ws1.append(["Key Metrics"])
    ws1.append(["Total Items", len(df)])
    ws1.append([])
    ws1.append(["COST-GUARD Band Distribution"])
    ws1.append(["Band", "Count", "%"])

    bands = ["PASS", "WARN", "HIGH", "CRITICAL", "UNKNOWN"]
    for band in bands:
        count = band_dist.get(band, 0)
        pct = f"{count/len(df)*100:.1f}%" if len(df) > 0 else "0%"
        ws1.append([band, count, pct])

    ws1.append([])
    ws1.append(["Verdict Distribution"])
    ws1.append(["Verdict", "Count", "%"])

    for verdict, count in sorted(verdict_dist.items()):
        pct = f"{count/len(df)*100:.1f}%" if len(df) > 0 else "0%"
        ws1.append([verdict, count, pct])

    ws1.append([])
    ws1.append(["Critical Items", band_dist.get("CRITICAL", 0)])
    ws1.append(
        [
            "Auto-Approved (VERIFIED + SPECIAL_PASS)",
            verdict_dist.get("VERIFIED", 0) + verdict_dist.get("SPECIAL_PASS", 0),
        ]
    )
    ws1.append(
        [
            "Manual Review Needed (PENDING_REVIEW + FAIL)",
            verdict_dist.get("PENDING_REVIEW", 0) + verdict_dist.get("FAIL", 0),
        ]
    )

    # Format
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"].font = Font(italic=True, size=11)
    ws1["A4"].font = Font(bold=True, size=12)
    ws1["A7"].font = Font(bold=True, size=12)
    ws1["A16"].font = Font(bold=True, size=12)

    for row in [8, 17]:
        for cell in ws1[row]:
            cell.fill = header_fill
            cell.font = header_font

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 15

    # Sheet 2: All Items
    print("Generating All Items...")
    ws2 = wb.create_sheet("All_Items")

    display_cols = [
        "S/N",
        "Shipment Reference#",
        "origin",
        "destination",
        "vehicle",
        "rate_usd",
        "ref_rate_usd",
        "delta_pct",
        "cg_band",
        "verdict",
        "ref_method",
    ]

    df_display = df[display_cols].copy()
    df_display["delta_pct"] = df_display["delta_pct"].apply(
        lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
    )

    for r in dataframe_to_rows(df_display, index=False, header=True):
        ws2.append(r)

    # Format headers
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Color code by band
    for row_idx in range(2, len(df) + 2):
        band = ws2[f"I{row_idx}"].value
        verdict = ws2[f"J{row_idx}"].value

        fill = None
        font = None

        if verdict == "SPECIAL_PASS":
            fill = special_fill
        elif band == "PASS":
            fill = pass_fill
        elif band == "WARN":
            fill = warn_fill
        elif band == "HIGH":
            fill = high_fill
        elif band == "CRITICAL":
            fill = critical_fill
            font = critical_font
        elif band == "UNKNOWN":
            fill = unknown_fill

        if fill or font:
            for col in range(1, 12):
                cell = ws2.cell(row=row_idx, column=col)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                cell.border = border

    ws2.freeze_panes = "A2"

    # Auto-width
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[column].width = min(max_length + 2, 50)

    # Sheet 3: CRITICAL Items
    print("Generating CRITICAL Items...")
    ws3 = wb.create_sheet("CRITICAL_Items")

    critical_df = df[df["cg_band"] == "CRITICAL"].copy()

    if not critical_df.empty:
        display_cols_critical = [
            "S/N",
            "Shipment Reference#",
            "origin",
            "destination",
            "vehicle",
            "# Trips",
            "rate_usd",
            "Amount (USD)",
            "ref_rate_usd",
            "delta_pct",
            "verdict",
            "ref_method",
            "Comments",
        ]

        critical_display = critical_df[display_cols_critical].copy()
        critical_display["delta_pct"] = critical_display["delta_pct"].apply(
            lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
        )

        for r in dataframe_to_rows(critical_display, index=False, header=True):
            ws3.append(r)

        # Format
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for row_idx in range(2, len(critical_df) + 2):
            verdict = ws3[f"K{row_idx}"].value

            fill = critical_fill if verdict != "SPECIAL_PASS" else special_fill
            font = critical_font if verdict != "SPECIAL_PASS" else None

            for col in range(1, 14):
                cell = ws3.cell(row=row_idx, column=col)
                cell.fill = fill
                if font:
                    cell.font = font
                cell.border = border

        ws3.freeze_panes = "A2"

        for col in ws3.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws3.column_dimensions[column].width = min(max_length + 2, 50)
    else:
        ws3.append(["No CRITICAL items found"])

    # Sheet 4: Special Pass
    print("Generating Special Pass Items...")
    ws4 = wb.create_sheet("Special_Pass")

    special_df = df[df["verdict"] == "SPECIAL_PASS"].copy()

    if not special_df.empty:
        display_cols_special = [
            "S/N",
            "Shipment Reference#",
            "origin",
            "destination",
            "vehicle",
            "rate_usd",
            "ref_rate_usd",
            "delta_pct",
            "cg_band",
            "ref_method",
        ]

        special_display = special_df[display_cols_special].copy()
        special_display["delta_pct"] = special_display["delta_pct"].apply(
            lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
        )

        for r in dataframe_to_rows(special_display, index=False, header=True):
            ws4.append(r)

        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row_idx in range(2, len(special_df) + 2):
            for col in range(1, 11):
                cell = ws4.cell(row=row_idx, column=col)
                cell.fill = special_fill

        ws4.freeze_panes = "A2"

        for col in ws4.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws4.column_dimensions[column].width = min(max_length + 2, 40)
    else:
        ws4.append(["No Special Pass items found"])

    # Save
    output_file = (
        base_dir / "Results" / "Sept_2025" / "DOMESTIC_SEPT_2025_FINAL_REPORT.xlsx"
    )
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_file)

    print(f"\n[OK] Final report saved: {output_file}")

    # Statistics
    print("\n" + "=" * 80)
    print("Final Validation Statistics")
    print("=" * 80)
    print(f"Total Items: {len(df)}")
    print(f"\nBand Distribution:")
    for band in bands:
        count = band_dist.get(band, 0)
        pct = count / len(df) * 100 if len(df) > 0 else 0
        print(f"  {band}: {count} ({pct:.1f}%)")
    print(
        f"\nCritical Items: {band_dist.get('CRITICAL', 0)} (Single Digit: {'✓' if band_dist.get('CRITICAL', 0) <= 9 else '✗'})"
    )
    print(
        f"Auto-Approved: {verdict_dist.get('VERIFIED', 0) + verdict_dist.get('SPECIAL_PASS', 0)} ({(verdict_dist.get('VERIFIED', 0) + verdict_dist.get('SPECIAL_PASS', 0))/len(df)*100:.1f}%)"
    )
    print("=" * 80)


if __name__ == "__main__":
    main()
