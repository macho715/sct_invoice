#!/usr/bin/env python3
"""
DOMESTIC September 2025 Excel Report Generator
44건 검증 결과를 SHPT 스타일 Excel 보고서로 생성
"""

import pandas as pd
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def format_route(origin, destination):
    """Route 포맷팅: Origin → Destination"""
    if pd.isna(origin) or pd.isna(destination):
        return "Unknown Route"
    return f"{origin} → {destination}"


def generate_decision_logic_domestic(row):
    """
    DOMESTIC 항목의 Decision Logic 생성 (SHPT 스타일)
    """
    origin = row.get("origin_norm", row.get("origin", "Unknown"))
    destination = row.get("destination_norm", row.get("destination", "Unknown"))
    vehicle = row.get("vehicle_norm", row.get("vehicle", "Unknown"))
    ref_rate = row.get("ref_rate_usd", None)
    draft_rate = row.get("rate_usd", 0)
    delta = row.get("delta_pct", 0)
    cg_band = row.get("cg_band", "N/A")
    decision = row.get("decision", "PENDING")
    similarity = row.get("similarity", 0)

    route = format_route(origin, destination)

    if pd.isna(ref_rate) or ref_rate == 0:
        return f"Lane {route} [{vehicle}]; ref rate not found in database; requires manual review"

    if abs(delta) <= 2.0:
        return f"Lane {route} [{vehicle}]; ref rate ${ref_rate:.2f} per truck; matches draft → {decision}"
    elif abs(delta) <= 10.0:
        return f"Lane {route} [{vehicle}]; ref rate ${ref_rate:.2f} per truck; delta {delta:.2f}% (Band: {cg_band}) → {decision}"
    else:
        if delta > 0:
            return f"Lane {route} [{vehicle}]; ref rate ${ref_rate:.2f}; overcharge suspect (+{delta:.2f}%) → {decision}"
        else:
            return f"Lane {route} [{vehicle}]; ref rate ${ref_rate:.2f}; undercharge suspect ({delta:.2f}%) → {decision}"


def find_dn_files(shipment_ref, supporting_docs_base):
    """
    Shipment Reference에 해당하는 DN 파일 찾기
    """
    if pd.isna(shipment_ref) or not supporting_docs_base.exists():
        return ""

    # Shipment Reference 정규화
    ref = str(shipment_ref).strip()

    # Supporting Documents 폴더 탐색
    for folder in supporting_docs_base.iterdir():
        if folder.is_dir() and ref in folder.name:
            # DN 파일 찾기
            dn_files = list(folder.glob("*.pdf"))
            if dn_files:
                formatted = []
                for dn_file in dn_files[:3]:  # 최대 3개만
                    formatted.append(f"**DN:** {dn_file.name}")
                if len(dn_files) > 3:
                    formatted.append(f"(+{len(dn_files)-3} more)")
                return "\n".join(formatted)

    return ""


def classify_issue_type(delta_pct, ref_rate):
    """Issue Type 분류"""
    if pd.isna(ref_rate) or ref_rate == 0:
        return "MISSING_REF"

    if pd.isna(delta_pct):
        return "UNKNOWN"

    if delta_pct > 10.0:
        return "OVERCHARGE"
    elif delta_pct < -10.0:
        return "UNDERCHARGE"
    elif abs(delta_pct) > 5.0:
        return "MODERATE_VARIANCE"
    else:
        return "NORMAL"


def determine_root_cause(row):
    """Root Cause 분석"""
    issue_type = classify_issue_type(
        row.get("delta_pct", 0), row.get("ref_rate_usd", 0)
    )
    delta = row.get("delta_pct", 0)
    similarity = row.get("similarity", 0)
    distance_km = row.get("distance_km", 0)

    if issue_type == "MISSING_REF":
        return "Reference rate not available in lane map"

    if similarity < 0.6:
        return (
            f"Low similarity match ({similarity:.2%}); route/vehicle mismatch possible"
        )

    if distance_km <= 10 and abs(delta) > 50:
        return "Short-run route (<10km); fixed cost minimum fare suspected"

    if issue_type == "OVERCHARGE":
        return f"Invoice rate significantly higher than reference (+{delta:.1f}%); verify contract terms"

    if issue_type == "UNDERCHARGE":
        return f"Invoice rate significantly lower than reference ({delta:.1f}%); verify special discount"

    return "Within acceptable variance"


def recommend_action(row):
    """Recommended Action"""
    issue_type = classify_issue_type(
        row.get("delta_pct", 0), row.get("ref_rate_usd", 0)
    )
    delta = row.get("delta_pct", 0)
    cg_band = row.get("cg_band", "N/A")

    if cg_band == "PASS":
        return "APPROVE"

    if cg_band == "WARN":
        return "APPROVE with note"

    if issue_type == "MISSING_REF":
        return "UPDATE lane map; manual review required"

    if issue_type == "OVERCHARGE":
        if abs(delta) > 100:
            return "REJECT; request rate justification from supplier"
        else:
            return "REVIEW contract terms; may approve with variance note"

    if issue_type == "UNDERCHARGE":
        return "REVIEW special discount authorization; may approve"

    return "REVIEW and decide"


def create_summary_data(df):
    """Summary 데이터 생성"""
    total_items = len(df)
    total_amount = (
        df["amount_usd"].sum() if "amount_usd" in df.columns else df["rate_usd"].sum()
    )

    # COST-GUARD 분포
    cg_counts = df["cg_band"].value_counts().to_dict()

    # Decision 분포
    decision_counts = df["decision"].value_counts().to_dict()

    summary_data = {
        "Total Items": total_items,
        "Total Amount (USD)": f"${total_amount:,.2f}",
        "PASS": cg_counts.get("PASS", 0),
        "WARN": cg_counts.get("WARN", 0),
        "HIGH": cg_counts.get("HIGH", 0),
        "CRITICAL": cg_counts.get("CRITICAL", 0),
        "Decision PASS": decision_counts.get("PASS", 0),
        "Decision PENDING_REVIEW": decision_counts.get("PENDING_REVIEW", 0),
        "Decision FAIL": decision_counts.get("FAIL", 0),
    }

    return pd.DataFrame([summary_data])


def create_all_items_report(df, supporting_docs_base):
    """All Items 시트 생성 (SHPT 스타일)"""
    all_items = pd.DataFrame(
        {
            "S/No": range(1, len(df) + 1),
            "Shipment Ref": df["shipment_ref"],
            "Route (Origin → Destination)": df.apply(
                lambda r: format_route(
                    r.get("origin_norm", r.get("origin")),
                    r.get("destination_norm", r.get("destination")),
                ),
                axis=1,
            ),
            "Vehicle Type": df.get("vehicle_norm", df.get("vehicle", "")),
            "Draft Rate (USD)": df["rate_usd"].round(2),
            "Trips": df.get("quantity", df.get("qty", 1)).astype(int),
            "Draft Total (USD)": (
                df["rate_usd"] * df.get("quantity", df.get("qty", 1))
            ).round(2),
            "Ref Rate (USD)": df["ref_rate_usd"].round(2),
            "Ref Total (USD)": (
                df["ref_rate_usd"] * df.get("quantity", df.get("qty", 1))
            ).round(2),
            "Δ%": df["delta_pct"].round(2),
            "CG Band": df["cg_band"],
            "Verdict": df["decision"],
            "Attachments": df["shipment_ref"].apply(
                lambda ref: find_dn_files(ref, supporting_docs_base)
            ),
            "Decision Logic": df.apply(generate_decision_logic_domestic, axis=1),
        }
    )

    return all_items


def create_critical_items_report(df, supporting_docs_base):
    """Critical Items 시트 생성"""
    critical = df[df["cg_band"] == "CRITICAL"].copy()

    if critical.empty:
        return pd.DataFrame()

    critical_report = pd.DataFrame(
        {
            "S/No": range(1, len(critical) + 1),
            "Shipment Ref": critical["shipment_ref"],
            "Route": critical.apply(
                lambda r: format_route(
                    r.get("origin_norm", r.get("origin")),
                    r.get("destination_norm", r.get("destination")),
                ),
                axis=1,
            ),
            "Vehicle": critical.get("vehicle_norm", critical.get("vehicle", "")),
            "Draft Rate": critical["rate_usd"].round(2),
            "Ref Rate": critical["ref_rate_usd"].round(2),
            "Delta %": critical["delta_pct"].round(2),
            "Issue Type": critical.apply(
                lambda r: classify_issue_type(r["delta_pct"], r["ref_rate_usd"]), axis=1
            ),
            "Root Cause": critical.apply(determine_root_cause, axis=1),
            "Recommended Action": critical.apply(recommend_action, axis=1),
            "Similarity Score": critical.get("similarity", 0).round(3),
            "Distance (km)": critical.get("distance_km", 0).round(1),
        }
    )

    return critical_report


def apply_header_format(ws):
    """헤더 행 포맷팅"""
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border


def apply_conditional_formatting(ws, verdict_col="L"):
    """Verdict 컬럼에 조건부 서식 적용"""
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        cell = ws[f"{verdict_col}{row}"]
        if cell.value:
            if "PASS" in str(cell.value):
                cell.fill = pass_fill
            elif "PENDING" in str(cell.value) or "REVIEW" in str(cell.value):
                cell.fill = warn_fill
            elif "FAIL" in str(cell.value):
                cell.fill = fail_fill


def adjust_column_widths(ws, sheet_name):
    """컬럼 너비 자동 조정"""
    if sheet_name == "All_Items":
        column_widths = {
            "A": 8,  # S/No
            "B": 30,  # Shipment Ref
            "C": 50,  # Route
            "D": 15,  # Vehicle
            "E": 15,  # Draft Rate
            "F": 8,  # Trips
            "G": 16,  # Draft Total
            "H": 16,  # Ref Rate
            "I": 16,  # Ref Total
            "J": 10,  # Δ%
            "K": 12,  # CG Band
            "L": 18,  # Verdict
            "M": 40,  # Attachments
            "N": 70,  # Decision Logic
        }
    elif sheet_name == "Critical_Items":
        column_widths = {
            "A": 8,  # S/No
            "B": 30,  # Shipment Ref
            "C": 50,  # Route
            "D": 15,  # Vehicle
            "E": 15,  # Draft Rate
            "F": 15,  # Ref Rate
            "G": 10,  # Delta %
            "H": 18,  # Issue Type
            "I": 50,  # Root Cause
            "J": 45,  # Recommended Action
            "K": 15,  # Similarity
            "L": 12,  # Distance
        }
    else:
        column_widths = {chr(65 + i): 20 for i in range(26)}

    for col, width in column_widths.items():
        if col in [chr(65 + i) for i in range(26)]:
            ws.column_dimensions[col].width = width


def freeze_header(ws):
    """첫 행 고정"""
    ws.freeze_panes = "A2"


def create_excel_report(df, summary_df, all_items_df, critical_df, output_path):
    """
    3개 시트로 Excel 보고서 생성
    """
    print("\nCreating DOMESTIC Excel workbook...")
    wb = Workbook()

    # 시트 1: Summary
    print("  - Writing Summary sheet...")
    ws1 = wb.active
    ws1.title = "Summary"
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws1.append(r)
    apply_header_format(ws1)
    adjust_column_widths(ws1, "Summary")

    # 시트 2: All Items
    print("  - Writing All_Items sheet...")
    ws2 = wb.create_sheet("All_Items")
    for r in dataframe_to_rows(all_items_df, index=False, header=True):
        ws2.append(r)
    apply_header_format(ws2)
    apply_conditional_formatting(ws2)
    adjust_column_widths(ws2, "All_Items")
    freeze_header(ws2)

    # 시트 3: Critical Items
    if not critical_df.empty:
        print("  - Writing Critical_Items sheet...")
        ws3 = wb.create_sheet("Critical_Items")
        for r in dataframe_to_rows(critical_df, index=False, header=True):
            ws3.append(r)
        apply_header_format(ws3)
        adjust_column_widths(ws3, "Critical_Items")
        freeze_header(ws3)

    # 저장
    print(f"  - Saving to {output_path.name}...")
    wb.save(output_path)
    print(f"[OK] Excel report created successfully!")


def main():
    """메인 실행 함수"""
    print("=" * 80)
    print("DOMESTIC Sept 2025 Excel Report Generator")
    print("=" * 80)

    # 경로 설정
    base_dir = Path(__file__).parent.parent
    csv_file = (
        base_dir
        / "Results"
        / "Sept_2025"
        / "CSV"
        / "domestic_sept_2025_result_20251013_012914.csv"
    )
    supporting_docs_base = (
        base_dir
        / "Data"
        / "DSV 202509"
        / "SCNT Domestic (Sept 2025) - Supporting Documents"
    )
    output_excel = (
        base_dir
        / "Results"
        / "Sept_2025"
        / "Reports"
        / "DOMESTIC_SEPT_2025_VALIDATION_REPORT.xlsx"
    )

    # CSV 존재 확인
    if not csv_file.exists():
        print(f"[ERROR] CSV not found: {csv_file}")
        return

    # CSV 로드
    print(f"\nLoading CSV file...")
    print(f"  - File: {csv_file.name}")
    df = pd.read_csv(csv_file)
    print(f"    OK Loaded {len(df)} items")

    # Summary 생성
    print("\nGenerating Summary...")
    summary_df = create_summary_data(df)
    print(f"  OK Summary generated")

    # All Items 생성
    print("\nGenerating All Items Report...")
    all_items_df = create_all_items_report(df, supporting_docs_base)
    print(f"  OK Generated {len(all_items_df)} items")

    # Critical Items 생성
    print("\nGenerating Critical Items Report...")
    critical_df = create_critical_items_report(df, supporting_docs_base)
    if not critical_df.empty:
        print(f"  OK Generated {len(critical_df)} critical items")
    else:
        print(f"  OK No critical items found")

    # Excel 생성
    create_excel_report(df, summary_df, all_items_df, critical_df, output_excel)

    # 통계 출력
    print("\n" + "=" * 80)
    print("Report Statistics")
    print("=" * 80)
    print(f"Total Items: {len(df)}")

    # CG Band 분포
    print(f"\nCOST-GUARD Band Distribution:")
    for band in ["PASS", "WARN", "HIGH", "CRITICAL"]:
        count = len(df[df["cg_band"] == band])
        pct = count / len(df) * 100 if len(df) > 0 else 0
        print(f"  {band}: {count} ({pct:.1f}%)")

    # Decision 분포
    print(f"\nDecision Distribution:")
    decision_counts = df["decision"].value_counts()
    for decision, count in decision_counts.items():
        pct = count / len(df) * 100
        print(f"  {decision}: {count} ({pct:.1f}%)")

    # ref_rate 통계
    ref_rate_filled = df["ref_rate_usd"].notna().sum()
    print(
        f"\nRef Rate Filled: {ref_rate_filled}/{len(df)} ({ref_rate_filled/len(df)*100:.1f}%)"
    )

    print(f"\nOutput Location:")
    print(f"   {output_excel}")
    print("\n[SUCCESS] DOMESTIC Excel report generation complete!")
    print("=" * 80)


if __name__ == "__main__":
    main()
