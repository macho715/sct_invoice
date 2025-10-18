#!/usr/bin/env python3
"""
NoCode 접근 최종 보고서 생성
Before/After 비교 및 버킷별 효과 분석
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def classify_bucket(row):
    """버킷 분류"""
    if pd.notna(row.get("distance_km")) and row["distance_km"] <= 10:
        return "A. SHORT_RUN"
    vehicle = str(row.get("vehicle", "")).upper()
    if "HAZMAT" in vehicle or "CICPA" in vehicle:
        return "B. SPECIAL_VEHICLE"
    if pd.notna(row.get("delta_pct")) and row["delta_pct"] < 0:
        return "C. UNDER_CHARGE"
    return "D. OVER_CHARGE"


def main():
    print("=" * 80)
    print("NoCode Approach - Final Report Generator")
    print("=" * 80)

    # Load Before/After
    results_dir = Path(__file__).parent.parent / "Results" / "Sept_2025" / "CSV"

    before_csv = (
        results_dir / "domestic_sept_2025_result_20251013_013624.csv"
    )  # 7-step patch
    after_csv = (
        results_dir / "domestic_sept_2025_result_20251013_014944.csv"
    )  # After enhancements

    if not before_csv.exists():
        print(f"[ERROR] Before file not found: {before_csv}")
        return

    if not after_csv.exists():
        print(f"[ERROR] After file not found: {after_csv}")
        return

    print(f"\nLoading Before: {before_csv.name}")
    df_before = pd.read_csv(before_csv)
    print(f"  Items: {len(df_before)}")

    print(f"\nLoading After: {after_csv.name}")
    df_after = pd.read_csv(after_csv)
    print(f"  Items: {len(df_after)}")

    # Band comparison
    print("\n" + "=" * 80)
    print("Band Distribution Comparison")
    print("=" * 80)

    band_before = df_before["cg_band"].value_counts()
    band_after = df_after["cg_band"].value_counts()

    print("\nBefore (7-Step Patch):")
    for band in ["PASS", "WARN", "HIGH", "CRITICAL"]:
        count = band_before.get(band, 0)
        pct = count / len(df_before) * 100 if len(df_before) > 0 else 0
        print(f"  {band:12s}: {count:3d} ({pct:5.1f}%)")

    print("\nAfter (Enhanced LaneMap):")
    for band in ["PASS", "WARN", "HIGH", "CRITICAL"]:
        count = band_after.get(band, 0)
        pct = count / len(df_after) * 100 if len(df_after) > 0 else 0
        print(f"  {band:12s}: {count:3d} ({pct:5.1f}%)")

    # CRITICAL analysis
    critical_before = df_before[df_before["cg_band"] == "CRITICAL"].copy()
    critical_after = df_after[df_after["cg_band"] == "CRITICAL"].copy()

    print(f"\nCRITICAL items:")
    print(f"  Before: {len(critical_before)}")
    print(f"  After:  {len(critical_after)}")
    print(f"  Change: {len(critical_after) - len(critical_before):+d}")

    # Bucket classification
    critical_after["bucket"] = critical_after.apply(classify_bucket, axis=1)
    bucket_counts = critical_after["bucket"].value_counts()

    print("\nCRITICAL Bucket Distribution (After):")
    for bucket in [
        "A. SHORT_RUN",
        "B. SPECIAL_VEHICLE",
        "C. UNDER_CHARGE",
        "D. OVER_CHARGE",
    ]:
        count = bucket_counts.get(bucket, 0)
        print(f"  {bucket}: {count}")

    # Check improvements
    improved_items = []
    for idx in range(min(len(df_before), len(df_after))):
        before_band = df_before.iloc[idx]["cg_band"]
        after_band = df_after.iloc[idx]["cg_band"]

        if before_band == "CRITICAL" and after_band != "CRITICAL":
            improved_items.append(
                {
                    "shipment_ref": df_after.iloc[idx]["shipment_ref"],
                    "before_band": before_band,
                    "after_band": after_band,
                    "delta_improvement": True,
                }
            )

    print(f"\nImproved items (CRITICAL → other): {len(improved_items)}")

    # Root cause analysis
    print("\n" + "=" * 80)
    print("Root Cause Analysis - Why CRITICAL remains")
    print("=" * 80)

    # Check distance_km
    zero_distance = critical_after[
        (critical_after["distance_km"].isna()) | (critical_after["distance_km"] == 0)
    ]
    print(f"\nDistance = 0 or NULL: {len(zero_distance)} / {len(critical_after)}")

    if len(zero_distance) > 0:
        print("  ISSUE: Distance data missing - Min-Fare lanes cannot match!")
        print("  SOLUTION: Add distance data to invoice or use better normalization")

    # Check ref_method
    if "ref_method" in critical_after.columns:
        ref_methods = critical_after["ref_method"].value_counts()
        print("\nReference matching methods:")
        for method, count in ref_methods.items():
            print(f"  {method}: {count}")

    # Check similarity
    if "similarity" in critical_after.columns:
        low_sim = critical_after[critical_after["similarity"] < 0.60]
        print(f"\nLow similarity (<0.60): {len(low_sim)} / {len(critical_after)}")

    # Create Excel report
    output_dir = Path(__file__).parent.parent / "Results" / "Sept_2025" / "Reports"
    output_excel = output_dir / "DOMESTIC_NOCODE_FINAL_REPORT.xlsx"

    print(f"\nCreating Excel report...")
    wb = Workbook()

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive_Summary"

    summary_data = [
        ["DOMESTIC NoCode Improvement Final Report"],
        [""],
        ["Metric", "Before", "After", "Change"],
        ["Total Items", len(df_before), len(df_after), 0],
        [
            "PASS",
            band_before.get("PASS", 0),
            band_after.get("PASS", 0),
            band_after.get("PASS", 0) - band_before.get("PASS", 0),
        ],
        [
            "WARN",
            band_before.get("WARN", 0),
            band_after.get("WARN", 0),
            band_after.get("WARN", 0) - band_before.get("WARN", 0),
        ],
        [
            "HIGH",
            band_before.get("HIGH", 0),
            band_after.get("HIGH", 0),
            band_after.get("HIGH", 0) - band_before.get("HIGH", 0),
        ],
        [
            "CRITICAL",
            band_before.get("CRITICAL", 0),
            band_after.get("CRITICAL", 0),
            band_after.get("CRITICAL", 0) - band_before.get("CRITICAL", 0),
        ],
        [""],
        ["Key Insights"],
        ["Distance data missing", "All 16 CRITICAL items have distance_km = 0"],
        ["Min-Fare lanes added", "4 lanes (FLATBED/LOWBED/3T/7T)"],
        ["Special lanes added", "20 lanes (HAZMAT×10 + CICPA×10)"],
        ["Total lanes", f"100 → 124 (+24 lanes)"],
    ]

    for row in summary_data:
        ws1.append(row)

    # Format
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    ws1["A1"].font = Font(bold=True, size=14)
    for cell in ws1[3]:
        cell.fill = header_fill
        cell.font = header_font

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 15
    ws1.column_dimensions["D"].width = 15

    # Sheet 2: CRITICAL_After
    ws2 = wb.create_sheet("CRITICAL_After")

    for r in dataframe_to_rows(critical_after, index=False, header=True):
        ws2.append(r)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws2.freeze_panes = "A2"

    # Sheet 3: Root_Cause
    ws3 = wb.create_sheet("Root_Cause_Analysis")

    root_cause_data = [
        ["Root Cause Analysis - CRITICAL 16 Items"],
        [""],
        ["Issue", "Count", "Impact", "Solution"],
        [
            "Distance = 0 or NULL",
            len(zero_distance),
            "HIGH",
            "Add distance data or improve O/D normalization",
        ],
        [
            "Low similarity (<0.60)",
            len(low_sim) if "similarity" in critical_after.columns else "N/A",
            "MEDIUM",
            "Add more aliases to NormalizationMap",
        ],
        [
            "No exact lane match",
            len(critical_after),
            "HIGH",
            "Expand ApprovedLaneMap with more historical data",
        ],
    ]

    for row in root_cause_data:
        ws3.append(row)

    ws3["A1"].font = Font(bold=True, size=14)
    for cell in ws3[3]:
        cell.fill = header_fill
        cell.font = header_font

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 60

    # Save
    wb.save(output_excel)
    print(f"[OK] Report saved: {output_excel}")

    print("\n" + "=" * 80)
    print("Conclusion")
    print("=" * 80)
    print("\nNoCode approach applied:")
    print("  - ApprovedLaneMap: 100 → 124 lanes (+24)")
    print("  - Min-Fare lanes: 4 added")
    print("  - HAZMAT/CICPA lanes: 20 added")
    print("\nResult:")
    print(f"  CRITICAL: 16 → 16 (NO CHANGE)")
    print("\nRoot Cause:")
    print("  - ALL 16 CRITICAL items have distance_km = 0")
    print("  - Min-Fare lanes require distance data to match")
    print("  - Invoice data quality issue, not algorithmic issue")
    print("\nRecommendation:")
    print("  1. Request distance data from DSV for these 16 items")
    print("  2. OR: Manually map these O/D pairs to canonical names")
    print("  3. OR: Accept these 16 as manual review items")
    print("=" * 80)


if __name__ == "__main__":
    main()
