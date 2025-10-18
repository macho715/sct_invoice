#!/usr/bin/env python3
"""
Duplicate File Identifier
중복/유사/버전 파일 식별 및 분석
"""

import pandas as pd
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def find_exact_duplicates(df):
    """파일명이 동일한 파일들 찾기"""
    duplicates = []

    filename_groups = df.groupby("filename")
    for filename, group in filename_groups:
        if len(group) > 1:
            for idx, row in group.iterrows():
                duplicates.append(
                    {
                        "filename": filename,
                        "location": row["directory"],
                        "full_path": row["relative_path"],
                        "size_kb": row["size_kb"],
                        "modified_date": row["modified_date"],
                        "hash": row.get("hash"),
                        "duplicate_type": "EXACT_FILENAME",
                    }
                )

    return pd.DataFrame(duplicates)


def find_version_files(df):
    """버전 파일들 찾기 (v2, enhanced, improved 등)"""
    version_patterns = ["_v2", "_v3", "_enhanced", "_improved", "_patched", "_updated"]
    version_files = []

    for idx, row in df.iterrows():
        filename = row["filename"].lower()

        for pattern in version_patterns:
            if pattern in filename:
                # Find base filename
                base_name = filename
                for p in version_patterns:
                    base_name = base_name.replace(p, "")

                version_files.append(
                    {
                        "base_name": base_name,
                        "version_filename": row["filename"],
                        "full_path": row["relative_path"],
                        "directory": row["directory"],
                        "version_indicator": pattern,
                        "size_kb": row["size_kb"],
                        "modified_date": row["modified_date"],
                        "category": row["category"],
                    }
                )
                break

    return pd.DataFrame(version_files)


def find_backup_files(df):
    """백업 파일들 찾기"""
    backup_files = df[df["category"] == "BACKUP"].copy()

    # Add analysis
    backup_analysis = []
    for idx, row in backup_files.iterrows():
        # Find original file
        original_name = row["filename"]
        for pattern in ["_backup_", "_BACKUP_", ".backup"]:
            original_name = original_name.replace(pattern, "_").replace("__", "_")

        # Remove timestamp patterns
        import re

        original_name = re.sub(r"_\d{8}_\d{6}", "", original_name)
        original_name = re.sub(r"_\d{14}", "", original_name)

        backup_analysis.append(
            {
                "backup_file": row["filename"],
                "estimated_original": original_name,
                "full_path": row["relative_path"],
                "size_kb": row["size_kb"],
                "modified_date": row["modified_date"],
                "directory": row["directory"],
            }
        )

    return pd.DataFrame(backup_analysis)


def find_hash_duplicates(df):
    """해시값이 동일한 파일들 찾기 (내용이 완전히 같은 파일)"""
    hash_duplicates = []

    # Only files with hash
    df_with_hash = df[df["hash"].notna()].copy()

    hash_groups = df_with_hash.groupby("hash")
    for hash_val, group in hash_groups:
        if len(group) > 1:
            for idx, row in group.iterrows():
                hash_duplicates.append(
                    {
                        "filename": row["filename"],
                        "location": row["directory"],
                        "full_path": row["relative_path"],
                        "size_kb": row["size_kb"],
                        "hash": hash_val,
                        "duplicate_count": len(group),
                    }
                )

    return pd.DataFrame(hash_duplicates)


def main():
    print("=" * 80)
    print("Duplicate File Identifier")
    print("=" * 80)

    # Load inventory
    root_dir = Path(__file__).parent
    inventory_file = root_dir / "FILE_INVENTORY.xlsx"

    if not inventory_file.exists():
        print(
            "[ERROR] FILE_INVENTORY.xlsx not found. Run create_file_inventory.py first."
        )
        return

    print(f"\nLoading inventory...")
    df = pd.read_excel(inventory_file, sheet_name="All_Files")
    print(f"  Total files: {len(df)}")

    # Find duplicates
    print("\n" + "=" * 80)
    print("Duplicate Analysis")
    print("=" * 80)

    print("\n[1/4] Finding exact filename duplicates...")
    exact_dup = find_exact_duplicates(df)
    print(
        f"  Exact filename duplicates: {len(exact_dup)} files in {len(exact_dup)//2 if len(exact_dup) > 0 else 0} groups"
    )

    print("\n[2/4] Finding version files...")
    version_files = find_version_files(df)
    print(f"  Version files: {len(version_files)}")

    print("\n[3/4] Finding backup files...")
    backup_files = find_backup_files(df)
    print(f"  Backup files: {len(backup_files)}")

    print("\n[4/4] Finding content duplicates (by hash)...")
    hash_dup = find_hash_duplicates(df)
    print(f"  Hash duplicates: {len(hash_dup)} files")

    # Create Excel report
    output_excel = root_dir / "DUPLICATE_ANALYSIS.xlsx"

    print(f"\nCreating Excel report...")
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["HVDC Invoice Audit - Duplicate Analysis"],
        [f"Analysis Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"],
        [""],
        ["Duplicate Type", "Count", "Action Required"],
        ["Exact Filename Duplicates", len(exact_dup), "Review and keep one copy"],
        [
            "Version Files (v2/enhanced/etc)",
            len(version_files),
            "Keep latest, archive old",
        ],
        ["Backup Files", len(backup_files), "Archive all backups"],
        ["Hash Duplicates (Same Content)", len(hash_dup), "Keep one, delete rest"],
        [""],
        ["Total Files Scanned", len(df), ""],
        [
            "Files Flagged for Review",
            len(exact_dup) + len(version_files) + len(backup_files) + len(hash_dup),
            "",
        ],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # Format
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary["A1"].font = Font(bold=True, size=14)
    for cell in ws_summary[4]:
        cell.fill = header_fill
        cell.font = header_font

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 40

    # Sheet 2: Exact Filename Duplicates
    if not exact_dup.empty:
        ws_exact = wb.create_sheet("Exact_Filename_Dup")
        for r in dataframe_to_rows(exact_dup, index=False, header=True):
            ws_exact.append(r)

        for cell in ws_exact[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_exact.freeze_panes = "A2"

    # Sheet 3: Version Files
    if not version_files.empty:
        ws_ver = wb.create_sheet("Version_Files")
        for r in dataframe_to_rows(version_files, index=False, header=True):
            ws_ver.append(r)

        for cell in ws_ver[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_ver.freeze_panes = "A2"

    # Sheet 4: Backup Files
    if not backup_files.empty:
        ws_bak = wb.create_sheet("Backup_Files")
        for r in dataframe_to_rows(backup_files, index=False, header=True):
            ws_bak.append(r)

        for cell in ws_bak[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_bak.freeze_panes = "A2"

    # Sheet 5: Hash Duplicates
    if not hash_dup.empty:
        ws_hash = wb.create_sheet("Hash_Duplicates")
        for r in dataframe_to_rows(hash_dup, index=False, header=True):
            ws_hash.append(r)

        for cell in ws_hash[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_hash.freeze_panes = "A2"

    # Save
    wb.save(output_excel)
    print(f"[OK] Duplicate analysis saved: {output_excel}")

    # Print recommendations
    print("\n" + "=" * 80)
    print("Recommendations")
    print("=" * 80)

    if len(backup_files) > 0:
        print(f"\n1. Archive {len(backup_files)} backup files:")
        for idx, row in backup_files.head(5).iterrows():
            print(f"   - {row['backup_file']}")
        if len(backup_files) > 5:
            print(f"   ... and {len(backup_files) - 5} more")

    if len(version_files) > 0:
        print(f"\n2. Review {len(version_files)} version files:")
        base_groups = version_files.groupby("base_name")
        for base_name, group in list(base_groups)[:5]:
            print(f"   Base: {base_name}")
            for idx, row in group.iterrows():
                print(f"     - {row['version_filename']} ({row['version_indicator']})")

    if len(hash_dup) > 0:
        print(f"\n3. Remove {len(hash_dup)} content-duplicate files (keep one copy)")

    print("\n" + "=" * 80)
    print("Next Steps:")
    print("=" * 80)
    print("1. Review DUPLICATE_ANALYSIS.xlsx")
    print("2. Run legacy file analysis (Step 3)")
    print("3. Create archive structure (Step 4)")
    print("=" * 80)


if __name__ == "__main__":
    main()
