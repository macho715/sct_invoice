#!/usr/bin/env python3
"""
Complete File Inventory Scanner
전체 시스템 파일을 스캔하고 카테고리별로 분류
"""

import os
import hashlib
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def get_file_hash(file_path):
    """파일 해시 계산 (중복 검사용)"""
    try:
        with open(file_path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()
    except:
        return None


def categorize_file(file_path):
    """파일 카테고리 분류"""
    name = file_path.name.lower()

    # Backup files
    if "backup" in name or "_backup_" in name:
        return "BACKUP"

    # Version indicators
    if any(v in name for v in ["_v2", "_v3", "enhanced", "improved", "patched"]):
        return "VERSION"

    # Legacy
    if "legacy" in str(file_path.parent).lower() or "old" in name:
        return "LEGACY"

    # Test files
    if name.startswith("test_") or "_test" in name:
        return "TEST"

    # Sample files
    if "sample" in name or "example" in name:
        return "SAMPLE"

    # Temporary files
    if "temp" in name or "tmp" in name:
        return "TEMP"

    # Core system files
    if "audit" in name or "validator" in name or "system" in name:
        return "CORE"

    # Analysis/utility scripts
    if any(
        word in name for word in ["analyze", "create", "extract", "check", "verify"]
    ):
        return "UTILITY"

    # Results/Reports
    if "result" in name or "report" in name:
        return "OUTPUT"

    # Documentation
    if file_path.suffix in [".md", ".txt", ".pdf"]:
        return "DOCUMENTATION"

    # Configuration
    if file_path.suffix in [".json", ".yaml", ".yml", ".ini"]:
        return "CONFIG"

    # Data files
    if file_path.suffix in [".xlsx", ".xls", ".xlsm", ".csv"]:
        if "result" in name or "report" in name:
            return "OUTPUT"
        return "DATA"

    return "OTHER"


def scan_directory(root_path):
    """디렉토리를 재귀적으로 스캔"""
    root = Path(root_path)
    file_list = []

    # Exclude patterns
    exclude_patterns = [
        "__pycache__",
        ".git",
        "node_modules",
        ".pytest_cache",
        "venv",
        "env",
    ]

    for file_path in root.rglob("*"):
        # Skip directories and excluded patterns
        if file_path.is_dir():
            continue

        if any(pattern in str(file_path) for pattern in exclude_patterns):
            continue

        # Get file info
        try:
            stat = file_path.stat()
            rel_path = file_path.relative_to(root)

            file_info = {
                "filename": file_path.name,
                "relative_path": str(rel_path),
                "full_path": str(file_path),
                "directory": str(file_path.parent.relative_to(root)),
                "extension": file_path.suffix,
                "size_bytes": stat.st_size,
                "size_kb": round(stat.st_size / 1024, 2),
                "modified_date": datetime.fromtimestamp(stat.st_mtime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                "category": categorize_file(file_path),
                "hash": (
                    get_file_hash(file_path)
                    if stat.st_size < 10 * 1024 * 1024
                    else None
                ),  # Skip large files
            }

            file_list.append(file_info)
        except Exception as e:
            print(f"[WARN] Cannot process {file_path}: {e}")

    return pd.DataFrame(file_list)


def main():
    print("=" * 80)
    print("File Inventory Scanner")
    print("=" * 80)

    # Root directory
    root_dir = Path(__file__).parent
    print(f"\nScanning: {root_dir}")

    # Scan all files
    print("\nScanning files...")
    df = scan_directory(root_dir)
    print(f"  Total files found: {len(df)}")

    # Statistics
    print("\n" + "=" * 80)
    print("File Statistics")
    print("=" * 80)

    print("\nBy Extension:")
    ext_counts = df["extension"].value_counts().head(10)
    for ext, count in ext_counts.items():
        ext_name = ext if ext else "(no extension)"
        print(f"  {ext_name:15s}: {count:4d}")

    print("\nBy Category:")
    cat_counts = df["category"].value_counts()
    for cat, count in cat_counts.items():
        print(f"  {cat:15s}: {count:4d}")

    print("\nBy Directory:")
    dir_counts = df["directory"].value_counts().head(15)
    for dir_name, count in dir_counts.items():
        print(f"  {dir_name[:60]:60s}: {count:4d}")

    # Total size
    total_size_mb = df["size_bytes"].sum() / (1024 * 1024)
    print(f"\nTotal Size: {total_size_mb:.2f} MB")

    # Create Excel report
    output_excel = root_dir / "FILE_INVENTORY.xlsx"

    print(f"\nCreating Excel report...")
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["HVDC Invoice Audit - File Inventory"],
        [f"Scan Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
        [""],
        ["Category", "Count", "Total Size (MB)"],
    ]

    for cat in cat_counts.index:
        cat_files = df[df["category"] == cat]
        cat_size = cat_files["size_bytes"].sum() / (1024 * 1024)
        summary_data.append([cat, len(cat_files), round(cat_size, 2)])

    summary_data.append([""])
    summary_data.append(["Total", len(df), round(total_size_mb, 2)])

    for row in summary_data:
        ws_summary.append(row)

    # Format
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary["A1"].font = Font(bold=True, size=14)
    for cell in ws_summary[4]:
        cell.fill = header_fill
        cell.font = header_font

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 20

    # Sheet 2: All Files
    ws_all = wb.create_sheet("All_Files")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_all.append(r)

    for cell in ws_all[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws_all.freeze_panes = "A2"

    # Sheet 3-N: By Category
    for category in cat_counts.index:
        cat_df = df[df["category"] == category].copy()
        cat_df = cat_df.sort_values("size_bytes", ascending=False)

        sheet_name = category[:31]  # Excel sheet name limit
        ws_cat = wb.create_sheet(sheet_name)

        for r in dataframe_to_rows(cat_df, index=False, header=True):
            ws_cat.append(r)

        for cell in ws_cat[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_cat.freeze_panes = "A2"

    # Save
    wb.save(output_excel)
    print(f"[OK] Inventory saved: {output_excel}")

    print("\n" + "=" * 80)
    print("Next Steps:")
    print("=" * 80)
    print("1. Review FILE_INVENTORY.xlsx")
    print("2. Run duplicate detection (Step 2)")
    print("3. Analyze legacy files (Step 3)")
    print("=" * 80)


if __name__ == "__main__":
    main()
