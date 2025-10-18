#!/usr/bin/env python3
"""
Legacy File Analyzer
레거시, 미사용, 구버전 파일 식별 및 분석
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def find_legacy_folder_files(df):
    """Legacy 폴더의 파일들"""
    legacy_files = df[df["category"] == "LEGACY"].copy()

    # Also check directory name
    legacy_in_path = df[
        df["directory"].str.contains("legacy", case=False, na=False)
    ].copy()

    # Combine
    all_legacy = pd.concat([legacy_files, legacy_in_path]).drop_duplicates(
        subset=["full_path"]
    )

    return all_legacy


def find_superseded_files(df):
    """신버전에 의해 대체된 구버전 파일들"""
    superseded = []

    # Core audit scripts
    audit_scripts = df[
        (df["extension"] == ".py")
        & (df["category"] == "CORE")
        & df["filename"].str.contains("audit", case=False, na=False)
    ].copy()

    # Group by base directory
    for directory in audit_scripts["directory"].unique():
        dir_files = audit_scripts[audit_scripts["directory"] == directory]

        if len(dir_files) > 1:
            # Sort by modification date
            dir_files_sorted = dir_files.sort_values("modified_date", ascending=False)

            # Latest is active, rest are superseded
            latest = dir_files_sorted.iloc[0]
            old_files = dir_files_sorted.iloc[1:]

            for idx, old_file in old_files.iterrows():
                superseded.append(
                    {
                        "superseded_file": old_file["filename"],
                        "superseded_by": latest["filename"],
                        "full_path": old_file["relative_path"],
                        "directory": old_file["directory"],
                        "modified_date": old_file["modified_date"],
                        "size_kb": old_file["size_kb"],
                    }
                )

    return pd.DataFrame(superseded)


def find_old_result_files(df):
    """오래된 결과 파일들 (최근 3개 런 제외)"""
    result_files = df[df["category"] == "OUTPUT"].copy()

    # Parse modification dates
    result_files["mod_datetime"] = pd.to_datetime(result_files["modified_date"])

    old_results = []

    # Group by directory and extension
    for directory in result_files["directory"].unique():
        for ext in [".csv", ".xlsx", ".json"]:
            dir_ext_files = result_files[
                (result_files["directory"] == directory)
                & (result_files["extension"] == ext)
            ].copy()

            if len(dir_ext_files) > 3:
                # Sort by date, keep latest 3
                dir_ext_files = dir_ext_files.sort_values(
                    "mod_datetime", ascending=False
                )
                old_files = dir_ext_files.iloc[3:]

                for idx, old_file in old_files.iterrows():
                    old_results.append(
                        {
                            "filename": old_file["filename"],
                            "full_path": old_file["relative_path"],
                            "directory": old_file["directory"],
                            "modified_date": old_file["modified_date"],
                            "size_kb": old_file["size_kb"],
                            "reason": f"Older than latest 3 {ext} files in directory",
                        }
                    )

    return pd.DataFrame(old_results)


def find_test_sample_files(df):
    """테스트 및 샘플 파일들"""
    test_sample = df[df["category"].isin(["TEST", "SAMPLE"])].copy()

    # Also check filename patterns
    pattern_files = df[
        df["filename"].str.contains("test|sample|example|demo", case=False, na=False)
    ].copy()

    all_test_sample = pd.concat([test_sample, pattern_files]).drop_duplicates(
        subset=["full_path"]
    )

    return all_test_sample


def main():
    print("=" * 80)
    print("Legacy File Analyzer")
    print("=" * 80)

    # Load inventory
    root_dir = Path(__file__).parent
    inventory_file = root_dir / "FILE_INVENTORY.xlsx"

    if not inventory_file.exists():
        print("[ERROR] FILE_INVENTORY.xlsx not found.")
        return

    print(f"\nLoading inventory...")
    df = pd.read_excel(inventory_file, sheet_name="All_Files")
    print(f"  Total files: {len(df)}")

    # Analyze legacy files
    print("\n" + "=" * 80)
    print("Legacy Analysis")
    print("=" * 80)

    print("\n[1/4] Finding Legacy folder files...")
    legacy_files = find_legacy_folder_files(df)
    print(f"  Legacy files: {len(legacy_files)}")

    print("\n[2/4] Finding superseded files...")
    superseded_files = find_superseded_files(df)
    print(f"  Superseded files: {len(superseded_files)}")

    print("\n[3/4] Finding old result files (>3 runs old)...")
    old_results = find_old_result_files(df)
    print(f"  Old result files: {len(old_results)}")

    print("\n[4/4] Finding test/sample files...")
    test_sample_files = find_test_sample_files(df)
    print(f"  Test/sample files: {len(test_sample_files)}")

    # Total candidates for archiving
    total_candidates = (
        len(legacy_files)
        + len(superseded_files)
        + len(old_results)
        + len(test_sample_files)
    )

    print(f"\nTotal archive candidates: {total_candidates}")

    # Create Excel report
    output_excel = root_dir / "LEGACY_FILES_REPORT.xlsx"

    print(f"\nCreating Excel report...")
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["HVDC Invoice Audit - Legacy Files Analysis"],
        [f"Analysis Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"],
        [""],
        ["Category", "Count", "Action"],
        ["Legacy Folder Files", len(legacy_files), "Archive"],
        ["Superseded Files", len(superseded_files), "Archive (keep latest only)"],
        ["Old Result Files (>3 runs)", len(old_results), "Archive"],
        ["Test/Sample Files", len(test_sample_files), "Archive"],
        [""],
        ["Total Archive Candidates", total_candidates, "Move to Archive/"],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # Format
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary["A1"].font = Font(bold=True, size=14)
    for cell in ws_summary[4]:
        cell.fill = header_fill
        cell.font = header_font

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 40

    # Additional sheets
    if not legacy_files.empty:
        ws_leg = wb.create_sheet("Legacy_Files")
        for r in dataframe_to_rows(legacy_files, index=False, header=True):
            ws_leg.append(r)
        for cell in ws_leg[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_leg.freeze_panes = "A2"

    if not superseded_files.empty:
        ws_sup = wb.create_sheet("Superseded_Files")
        for r in dataframe_to_rows(superseded_files, index=False, header=True):
            ws_sup.append(r)
        for cell in ws_sup[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_sup.freeze_panes = "A2"

    if not old_results.empty:
        ws_old = wb.create_sheet("Old_Results")
        for r in dataframe_to_rows(old_results, index=False, header=True):
            ws_old.append(r)
        for cell in ws_old[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_old.freeze_panes = "A2"

    if not test_sample_files.empty:
        ws_test = wb.create_sheet("Test_Sample")
        for r in dataframe_to_rows(test_sample_files, index=False, header=True):
            ws_test.append(r)
        for cell in ws_test[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_test.freeze_panes = "A2"

    # Save
    wb.save(output_excel)
    print(f"[OK] Legacy analysis saved: {output_excel}")

    print("\n" + "=" * 80)
    print("Next Steps:")
    print("=" * 80)
    print("1. Review LEGACY_FILES_REPORT.xlsx")
    print("2. Create archive structure (Step 4)")
    print("3. Move files to archive (Step 5)")
    print("=" * 80)


if __name__ == "__main__":
    main()
