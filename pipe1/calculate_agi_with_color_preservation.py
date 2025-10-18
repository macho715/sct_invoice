"""
AGI 이후 13개 컬럼 자동 계산 - 색상 보존 버전
openpyxl 사용하여 기존 색상 유지
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import sys
from pathlib import Path


def calculate_agi_columns_preserve_colors(excel_file):
    print(f"\n=== AGI 컬럼 계산 시작 (색상 보존) ===")
    print(f"파일: {excel_file}")

    # 1. pandas로 데이터 읽기 및 계산
    df = pd.read_excel(excel_file)
    print(f"원본 행 수: {len(df)}")
    print(f"원본 컬럼 수: {len(df.columns)}")

    # AGI 위치 찾기
    try:
        agi_idx = df.columns.get_loc("AGI")
        print(f"AGI 컬럼 위치: {agi_idx}")
    except KeyError:
        print("❌ AGI 컬럼을 찾을 수 없습니다!")
        return False

    # Warehouse와 Site 컬럼 식별
    warehouse_cols = [
        "DHL Warehouse",
        "DSV Indoor",
        "DSV Al Markaz",
        "Hauler Indoor",
        "DSV Outdoor",
        "DSV MZP",
        "HAULER",
        "JDN MZD",
        "MOSB",
        "AAA  Storage",
    ]
    site_cols = ["MIR", "SHU", "AGI", "DAS"]

    warehouse_names = [
        "DSV Indoor",
        "DSV Al Markaz",
        "Hauler Indoor",
        "DSV Outdoor",
        "DSV MZP",
        "HAULER",
        "JDN MZD",
        "MOSB",
        "AAA  Storage",
        "DHL Warehouse",
    ]
    site_names = ["mir", "shu", "agi", "das"]

    wh_cols = [c for c in warehouse_cols if c in df.columns]
    st_cols = [c for c in site_cols if c in df.columns]

    print(f"\nWarehouse 컬럼 ({len(wh_cols)}개): {wh_cols}")
    print(f"Site 컬럼 ({len(st_cols)}개): {st_cols}")

    # 13개 컬럼 계산
    print("\n컬럼 계산 중...")

    df["Status_WAREHOUSE"] = df[wh_cols].apply(
        lambda row: 1 if row.count() > 0 else "", axis=1
    )
    df["Status_SITE"] = df[st_cols].apply(
        lambda row: 1 if row.count() > 0 else "", axis=1
    )
    df["Status_Current"] = df.apply(
        lambda row: (
            "site"
            if row["Status_SITE"] == 1
            else ("warehouse" if row["Status_WAREHOUSE"] == 1 else "Pre Arrival")
        ),
        axis=1,
    )

    def get_latest_location(row):
        if row["Status_Current"] == "site":
            dates = row[st_cols]
            valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
            if len(valid_dates) > 0:
                return valid_dates.idxmax()
        elif row["Status_Current"] == "warehouse":
            dates = row[wh_cols]
            valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
            if len(valid_dates) > 0:
                return valid_dates.idxmax()
        return "Pre Arrival"

    df["Status_Location"] = df.apply(get_latest_location, axis=1)

    def get_latest_date(row):
        if row["Status_Current"] == "site":
            dates = row[st_cols]
            valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
            if len(valid_dates) > 0:
                return valid_dates.max()
        elif row["Status_Current"] == "warehouse":
            dates = row[wh_cols]
            valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
            if len(valid_dates) > 0:
                return valid_dates.max()
        return ""

    df["Status_Location_Date"] = df.apply(get_latest_date, axis=1)

    def classify_storage(row):
        loc = row["Status_Location"]
        if loc == "Pre Arrival":
            return "Pre Arrival"
        elif loc in warehouse_names:
            return "warehouse"
        elif loc.lower() in site_names:
            return "site"
        return ""

    df["Status_Storage"] = df.apply(classify_storage, axis=1)
    df["wh handling"] = df[wh_cols].apply(
        lambda row: sum(pd.to_numeric(row, errors="coerce").notna()), axis=1
    )
    df["site handling"] = df[st_cols].apply(
        lambda row: sum(pd.to_numeric(row, errors="coerce").notna()), axis=1
    )
    df["total handling"] = df["wh handling"] + df["site handling"]
    df["minus"] = df["site handling"] - df["wh handling"]
    df["final handling"] = df["total handling"] + df["minus"]

    if "규격" in df.columns and "수량" in df.columns:
        df["SQM"] = (df["규격"] * df["수량"]) / 10000
    else:
        df["SQM"] = ""

    df["Stack_Status"] = ""

    print(f"\n계산 완료: 13개 컬럼 추가됨")

    # 2. openpyxl로 원본 파일 열기 (색상 보존)
    print("\n색상 보존 처리 중...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 기존 색상 저장
    color_map = {}
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if (
                cell.fill
                and cell.fill.start_color
                and hasattr(cell.fill.start_color, "rgb")
            ):
                rgb = str(cell.fill.start_color.rgb)
                if rgb not in ["00000000", "FFFFFFFF", None]:  # 기본 색상 제외
                    color_map[(row_idx, col_idx)] = cell.fill

    print(f"기존 색상 {len(color_map)}개 저장됨")

    # 3. 새 데이터를 임시 파일로 저장
    temp_file = excel_file.replace(".xlsx", "_temp.xlsx")
    df.to_excel(temp_file, index=False)

    # 4. 임시 파일을 열어서 색상 복원
    wb_new = openpyxl.load_workbook(temp_file)
    ws_new = wb_new.active

    # 색상 복원
    for (row_idx, col_idx), fill in color_map.items():
        if row_idx <= ws_new.max_row and col_idx <= ws_new.max_column:
            ws_new.cell(row=row_idx, column=col_idx).fill = fill

    # 5. 원본 파일로 저장
    wb_new.save(excel_file)
    print(f"\n✅ 파일 저장 완료 (색상 보존): {excel_file}")

    # 임시 파일 삭제
    Path(temp_file).unlink()

    return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python calculate_agi_with_color_preservation.py <Excel파일>")
        sys.exit(1)

    excel_file = sys.argv[1]
    if not Path(excel_file).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {excel_file}")
        sys.exit(1)

    success = calculate_agi_columns_preserve_colors(excel_file)
    sys.exit(0 if success else 1)
