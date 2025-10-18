"""모든 처리를 한 번에 - 색상 보존하면서 AGI 계산 및 복사"""

from pathlib import Path

import openpyxl
import pandas as pd

from pipe1.agi_columns import (
    DERIVED_COLUMNS,
    FINAL_HANDLING_COLUMN,
    MINUS_COLUMN,
    SITE_COLUMNS,
    SITE_HANDLING_COLUMN,
    SQM_COLUMN,
    STACK_STATUS_COLUMN,
    STATUS_CURRENT_COLUMN,
    STATUS_LOCATION_COLUMN,
    STATUS_LOCATION_DATE_COLUMN,
    STATUS_SITE_COLUMN,
    STATUS_STORAGE_COLUMN,
    STATUS_WAREHOUSE_COLUMN,
    TOTAL_HANDLING_COLUMN,
    WH_HANDLING_COLUMN,
    WAREHOUSE_COLUMNS,
)


# 1. pandas로 AGI 컬럼 계산
print("=== AGI 컬럼 계산 중 ===")
df = pd.read_excel("HVDC WAREHOUSE_HITACHI(HE).synced.xlsx")

warehouse_cols = [c for c in WAREHOUSE_COLUMNS if c in df.columns]
site_cols = [c for c in SITE_COLUMNS if c in df.columns]
wh_cols = warehouse_cols
st_cols = site_cols

df[STATUS_WAREHOUSE_COLUMN] = df[wh_cols].apply(
    lambda row: 1 if row.count() > 0 else "", axis=1
)
df[STATUS_SITE_COLUMN] = df[st_cols].apply(lambda row: 1 if row.count() > 0 else "", axis=1)
df[STATUS_CURRENT_COLUMN] = df.apply(
    lambda row: (
        "site"
        if row[STATUS_SITE_COLUMN] == 1
        else ("warehouse" if row[STATUS_WAREHOUSE_COLUMN] == 1 else "Pre Arrival")
    ),
    axis=1,
)


def get_latest_location(row):
    if row[STATUS_CURRENT_COLUMN] == "site":
        dates = row[st_cols]
        valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
        if len(valid_dates) > 0:
            return valid_dates.idxmax()
    elif row[STATUS_CURRENT_COLUMN] == "warehouse":
        dates = row[wh_cols]
        valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
        if len(valid_dates) > 0:
            return valid_dates.idxmax()
    return "Pre Arrival"


df[STATUS_LOCATION_COLUMN] = df.apply(get_latest_location, axis=1)


def get_latest_date(row):
    if row[STATUS_CURRENT_COLUMN] == "site":
        dates = row[st_cols]
        valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
        if len(valid_dates) > 0:
            return valid_dates.max()
    elif row[STATUS_CURRENT_COLUMN] == "warehouse":
        dates = row[wh_cols]
        valid_dates = dates[pd.to_datetime(dates, errors="coerce").notna()]
        if len(valid_dates) > 0:
            return valid_dates.max()
    return ""


df[STATUS_LOCATION_DATE_COLUMN] = df.apply(get_latest_date, axis=1)

warehouse_names = [
    "DSV Indoor",
    "DSV Al Markaz",
    "Hauler Indoor",
    "DSV Outdoor",
    "DSV MZP",
    "HAULER",
    "JDN MZD",
    "MOSB",
    "AAA  Storage",
    "DHL Warehouse",
]
site_names = ["mir", "shu", "agi", "das"]


def classify_storage(row):
    loc = row[STATUS_LOCATION_COLUMN]
    if loc == "Pre Arrival":
        return "Pre Arrival"
    elif loc in warehouse_names:
        return "warehouse"
    elif loc.lower() in site_names:
        return "site"
    return ""


df[STATUS_STORAGE_COLUMN] = df.apply(classify_storage, axis=1)
df[WH_HANDLING_COLUMN] = df[wh_cols].apply(
    lambda row: sum(pd.to_numeric(row, errors="coerce").notna()), axis=1
)
df[SITE_HANDLING_COLUMN] = df[st_cols].apply(
    lambda row: sum(pd.to_numeric(row, errors="coerce").notna()), axis=1
)
df[TOTAL_HANDLING_COLUMN] = df[WH_HANDLING_COLUMN] + df[SITE_HANDLING_COLUMN]
df[MINUS_COLUMN] = df[SITE_HANDLING_COLUMN] - df[WH_HANDLING_COLUMN]
df[FINAL_HANDLING_COLUMN] = df[TOTAL_HANDLING_COLUMN] + df[MINUS_COLUMN]

if "규격" in df.columns and "수량" in df.columns:
    df[SQM_COLUMN] = (df["규격"] * df["수량"]) / 10000
else:
    df[SQM_COLUMN] = ""

df[STACK_STATUS_COLUMN] = ""

print(f"✅ AGI 컬럼 {len(DERIVED_COLUMNS)}개 계산 완료")

# 2. openpyxl로 색상 보존하면서 저장
print("\n=== 색상 보존 처리 중 ===")
wb_orig = openpyxl.load_workbook("HVDC WAREHOUSE_HITACHI(HE).synced.xlsx")
ws_orig = wb_orig.active

# 색상 저장
colors = {}
for row_idx in range(1, ws_orig.max_row + 1):
    for col_idx in range(1, ws_orig.max_column + 1):
        cell = ws_orig.cell(row=row_idx, column=col_idx)
        if cell.fill and cell.fill.start_color:
            rgb = getattr(cell.fill.start_color, "rgb", None)
            if rgb and rgb not in ["00000000", "FFFFFFFF"]:
                colors[(row_idx, col_idx)] = cell.fill

print(f"기존 색상 {len(colors)}개 저장됨")

# 3. 새 데이터를 임시 파일로 저장
df.to_excel("_temp.xlsx", index=False)

# 4. 색상 복원
wb_new = openpyxl.load_workbook("_temp.xlsx")
ws_new = wb_new.active

for (row_idx, col_idx), fill in colors.items():
    if row_idx <= ws_new.max_row and col_idx <= ws_new.max_column:
        ws_new.cell(row=row_idx, column=col_idx).fill = fill

# 5. 최종 파일로 저장
wb_new.save("HVDC WAREHOUSE_HITACHI(HE).xlsx")
print(f"✅ 색상 보존된 최종 파일 생성")

# 임시 파일 삭제
Path("_temp.xlsx").unlink()

print("\n" + "=" * 60)
print("✅ 전체 처리 완료!")
print("=" * 60)
